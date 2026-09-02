"""
excel_exporter_plt_regulares.py — Generador de Reporte Excel Profesional Multi-Hoja QA/QC para PLT Regulares.
Estructura y estilos 100% calcados del motor de auditoría de Mapeo Geomecánico (plt_excel_exporter.py).
Incluye:
- 📊 Dashboard Ejecutivo con 5 Tarjetas KPI, tablas de distribución, Top Desviaciones y BarChart nativo de Excel.
- 📋 Catálogo de Errores con ordenamiento por frecuencia descendente (errores arriba) e hipervínculos dinámicos.
- 📑 Detalle de Incidencias plano con autofiltros y formato condicional.
- Pestañas individuales dedicadas por categoría de error con botones de navegación '⬅ Volver al Catálogo'.
"""

from collections import Counter, defaultdict
from datetime import datetime
import os
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, Any, List

from app.core.rules_plt_regulares import CATEGORIES_REGISTRY_PLT_REGULARES


def get_safe_sheet_name(name: str, index: int = 1) -> str:
    """Genera un nombre de pestaña seguro y recortado para Excel (<= 31 caracteres)."""
    clean_name = str(name).replace(":", " ").replace("/", " ").replace("\\", " ").replace("?", "").replace("*", "").replace("[", "").replace("]", "")
    clean_name = " ".join(clean_name.split())
    if len(clean_name) > 26:
        clean_name = clean_name[:26].strip()
    return f"{index:02d}. {clean_name}"[:31]


def safe_int(val, default=0):
    try:
        return int(val)
    except Exception:
        return default


def safe_float(val, default=0.0):
    try:
        return float(val)
    except Exception:
        return default


def get_styles():
    """Paleta corporativa ejecutiva estándar de GEMA."""
    return {
        "font_title": Font(name="Segoe UI", size=16, bold=True, color="1B365D"),
        "font_subtitle": Font(name="Segoe UI", size=10, italic=True, color="555555"),
        "font_section": Font(name="Segoe UI", size=11, bold=True, color="1B365D"),
        "font_header": Font(name="Segoe UI", size=10, bold=True, color="FFFFFF"),
        "font_bold": Font(name="Segoe UI", size=10, bold=True, color="000000"),
        "font_regular": Font(name="Segoe UI", size=10, color="000000"),
        "font_kpi_lbl": Font(name="Segoe UI", size=9, bold=True, color="555555"),
        "font_kpi_blue": Font(name="Segoe UI", size=18, bold=True, color="1B365D"),
        "font_kpi_green": Font(name="Segoe UI", size=18, bold=True, color="375623"),
        "font_kpi_red": Font(name="Segoe UI", size=18, bold=True, color="C00000"),
        "font_kpi_orange": Font(name="Segoe UI", size=18, bold=True, color="C65911"),
        "font_link": Font(name="Segoe UI", size=10, bold=True, color="1B365D", underline="single"),
        "font_back_link": Font(name="Segoe UI", size=10, bold=True, color="1B365D", underline="single"),
        
        "fill_primary": PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid"),
        "fill_green": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "fill_yellow": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "fill_orange": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        "fill_red": PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid"),
        "fill_zebra": PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid"),
        "fill_kpi_gray": PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid"),
        
        "border_thin": Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9")
        ),
        "border_kpi": Border(
            left=Side(style="thin", color="B0C4DE"),
            right=Side(style="thin", color="B0C4DE"),
            top=Side(style="thin", color="B0C4DE"),
            bottom=Side(style="thin", color="B0C4DE")
        ),
        
        "align_center": Alignment(horizontal="center", vertical="center"),
        "align_left": Alignment(horizontal="left", vertical="center"),
        "align_right": Alignment(horizontal="right", vertical="center")
    }


def write_kpi_card(ws, start_row, start_col, label, value, bg_fill, val_font, s):
    """Escribe una tarjeta KPI estándar de 2 columnas x 2 filas."""
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 1)
    c_lbl = ws.cell(row=start_row, column=start_col, value=label)
    c_lbl.font = s["font_kpi_lbl"]
    c_lbl.alignment = s["align_center"]

    ws.merge_cells(start_row=start_row + 1, start_column=start_col, end_row=start_row + 1, end_column=start_col + 1)
    c_val = ws.cell(row=start_row + 1, column=start_col, value=value)
    c_val.font = val_font
    c_val.alignment = s["align_center"]

    for r in range(start_row, start_row + 2):
        for c in range(start_col, start_col + 2):
            cell = ws.cell(row=r, column=c)
            cell.fill = bg_fill
            cell.border = s["border_kpi"]


def export_plt_regulares_to_excel(diag: Dict[str, Any], output_path: str):
    """Genera el libro Excel (.xlsx) completo de auditoría QA/QC idéntico a Mapeo."""
    s = get_styles()
    font_title = s["font_title"]
    font_subtitle = s["font_subtitle"]
    font_section = s["font_section"]
    font_header = s["font_header"]
    font_bold = s["font_bold"]
    font_regular = s["font_regular"]

    font_kpi_val_blue = s["font_kpi_blue"]
    font_kpi_val_green = s["font_kpi_green"]
    font_kpi_val_red = s["font_kpi_red"]
    font_kpi_val_orange = s["font_kpi_orange"]

    fill_primary = s["fill_primary"]
    fill_accent_green = s["fill_green"]
    fill_accent_yellow = s["fill_yellow"]
    fill_accent_orange = s["fill_orange"]
    fill_accent_red = s["fill_red"]
    fill_zebra = s["fill_zebra"]
    fill_kpi_gray = s["fill_kpi_gray"]

    border_thin = s["border_thin"]
    alignment_center = s["align_center"]
    alignment_left = s["align_left"]
    alignment_right = s["align_right"]

    wb = openpyxl.Workbook()

    total_filas = diag.get("total_rows", 0)
    total_alertas = diag.get("severity_counts", {}).get("ALERTA", 0)
    total_advertencias = diag.get("severity_counts", {}).get("ADVERTENCIA", 0)
    total_vacios = diag.get("severity_counts", {}).get("VACIO", 0)
    pct_integridad = diag.get("quality_index", 100.0)
    total_taladros = len(diag.get("drillhole_stats", {}))

    # =========================================================================
    # --- HOJA 1: 📊 DASHBOARD EJECUTIVO ---
    # =========================================================================
    ws_dash = wb.active
    ws_dash.title = "📊 Dashboard Ejecutivo"
    ws_dash.views.sheetView[0].showGridLines = True

    ws_dash.cell(row=2, column=2, value="SISTEMA DE AUDITORÍA GEOTÉCNICA — ENSAYOS PLT REGULARES").font = font_title
    ws_dash.cell(
        row=3, column=2,
        value=f"Dashboard Ejecutivo de Control de Calidad, Consistencia Geomecánica y Diámetros de Perforación (Logueo DDH) | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ).font = font_subtitle

    # 5 Tarjetas KPI Superiores
    write_kpi_card(ws_dash, 5, 2, "TALADROS EVALUADOS", f"{total_taladros:,}", fill_kpi_gray, font_kpi_val_blue, s)
    write_kpi_card(ws_dash, 5, 4, "MUESTRAS REGISTRADAS", f"{total_filas:,}", fill_kpi_gray, font_kpi_val_blue, s)
    write_kpi_card(ws_dash, 5, 6, "INTEGRIDAD GLOBAL", f"{pct_integridad:.2f}%", fill_accent_green, font_kpi_val_green, s)
    write_kpi_card(ws_dash, 5, 8, "ALERTAS CRÍTICAS", total_alertas, fill_accent_red, font_kpi_val_red, s)
    write_kpi_card(ws_dash, 5, 10, "ADVERTENCIAS", total_advertencias, fill_accent_orange, font_kpi_val_orange, s)

    # 1. Tabla: Desempeño por Campaña
    ws_dash.cell(row=9, column=2, value="DESEMPEÑO DE CONTROL POR CAMPAÑA").font = font_section
    headers_camp = ["Campaña", "Muestras", "Taladros Afectados", "Alertas (N)", "% Alertas", "Vacíos (N)", "% Vacíos", "Calidad %"]
    for idx, col in enumerate(headers_camp, start=2):
        cell = ws_dash.cell(row=10, column=idx, value=col)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = alignment_center
        cell.border = border_thin

    camp_stats = diag.get("campaign_stats", {})
    r_camp = 11
    for camp_name in sorted(camp_stats.keys()):
        c_data = camp_stats[camp_name]
        tot = c_data["total"]
        al = c_data["alertas"]
        va = c_data["vacios"]
        q_camp = max(0.0, (tot - (al + va)) / tot * 100.0) if tot > 0 else 100.0

        ws_dash.cell(row=r_camp, column=2, value=str(camp_name)).font = font_bold
        ws_dash.cell(row=r_camp, column=2).alignment = alignment_center

        ws_dash.cell(row=r_camp, column=3, value=safe_int(tot)).number_format = '#,##0'
        ws_dash.cell(row=r_camp, column=3).alignment = alignment_right

        ws_dash.cell(row=r_camp, column=4, value=safe_int(len([dh for dh, d in diag.get("drillhole_stats", {}).items() if d.get("alertas", 0) > 0]))).number_format = '#,##0'
        ws_dash.cell(row=r_camp, column=4).alignment = alignment_right

        ws_dash.cell(row=r_camp, column=5, value=safe_int(al)).number_format = '#,##0'
        ws_dash.cell(row=r_camp, column=5).alignment = alignment_right

        ws_dash.cell(row=r_camp, column=6, value=(al / max(1, tot))).number_format = '0.00%'
        ws_dash.cell(row=r_camp, column=6).alignment = alignment_right

        ws_dash.cell(row=r_camp, column=7, value=safe_int(va)).number_format = '#,##0'
        ws_dash.cell(row=r_camp, column=7).alignment = alignment_right

        ws_dash.cell(row=r_camp, column=8, value=(va / max(1, tot))).number_format = '0.00%'
        ws_dash.cell(row=r_camp, column=8).alignment = alignment_right

        c_q = ws_dash.cell(row=r_camp, column=9, value=q_camp / 100.0)
        c_q.number_format = '0.00%'
        c_q.alignment = alignment_right
        c_q.font = font_bold

        for col_idx in range(2, 10):
            ws_dash.cell(row=r_camp, column=col_idx).border = border_thin
            if r_camp % 2 == 0:
                ws_dash.cell(row=r_camp, column=col_idx).fill = fill_zebra
        r_camp += 1

    # 2. Tabla: Top 5 Taladros con Mayor Incidencia
    r_worst = r_camp + 2
    ws_dash.cell(row=r_worst, column=2, value="TOP 5 TALADROS CON MAYOR INCIDENCIA").font = font_section
    headers_worst = ["Taladro (DDH)", "Muestras", "Alertas (N)", "Advertencias (N)", "Vacíos (N)", "Salud %"]
    for idx, col in enumerate(headers_worst, start=2):
        cell = ws_dash.cell(row=r_worst + 1, column=idx, value=col)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = alignment_center
        cell.border = border_thin

    dh_sorted = sorted(
        diag.get("drillhole_stats", {}).items(),
        key=lambda x: (x[1].get("alertas", 0) + x[1].get("vacios", 0)),
        reverse=True
    )

    curr_w_r = r_worst + 2
    for dh_name, d_val in dh_sorted[:5]:
        t_dh = d_val["total"]
        al_dh = d_val["alertas"]
        ad_dh = d_val["advertencias"]
        va_dh = d_val["vacios"]
        score = max(0.0, (t_dh - (al_dh + va_dh)) / t_dh * 100.0) if t_dh > 0 else 100.0

        ws_dash.cell(row=curr_w_r, column=2, value=str(dh_name)).font = font_bold
        ws_dash.cell(row=curr_w_r, column=2).alignment = alignment_left

        ws_dash.cell(row=curr_w_r, column=3, value=safe_int(t_dh)).number_format = '#,##0'
        ws_dash.cell(row=curr_w_r, column=3).alignment = alignment_right

        ws_dash.cell(row=curr_w_r, column=4, value=safe_int(al_dh)).number_format = '#,##0'
        ws_dash.cell(row=curr_w_r, column=4).alignment = alignment_right

        ws_dash.cell(row=curr_w_r, column=5, value=safe_int(ad_dh)).number_format = '#,##0'
        ws_dash.cell(row=curr_w_r, column=5).alignment = alignment_right

        ws_dash.cell(row=curr_w_r, column=6, value=safe_int(va_dh)).number_format = '#,##0'
        ws_dash.cell(row=curr_w_r, column=6).alignment = alignment_right

        c_sc = ws_dash.cell(row=curr_w_r, column=7, value=score / 100.0)
        c_sc.number_format = '0.00%'
        c_sc.alignment = alignment_right
        c_sc.font = font_bold

        for col_idx in range(2, 8):
            ws_dash.cell(row=curr_w_r, column=col_idx).border = border_thin
            if curr_w_r % 2 == 0:
                ws_dash.cell(row=curr_w_r, column=col_idx).fill = fill_zebra
        curr_w_r += 1

    # 3. Tabla: Principales Desviaciones Críticas (Columna Derecha J)
    ws_dash.cell(row=9, column=10, value="PRINCIPALES DESVIACIONES CRÍTICAS").font = font_section
    headers_top = ["Regla de Consistencia", "Cantidad (N)", "% Incidencia"]
    for idx, col in enumerate(headers_top, start=10):
        cell = ws_dash.cell(row=10, column=idx, value=col)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = alignment_center
        cell.border = border_thin

    cat_counts_sorted = sorted(diag.get("category_counts", {}).items(), key=lambda x: x[1], reverse=True)
    top_deviations = cat_counts_sorted[:5]

    for idx, (cat_code, count_val) in enumerate(top_deviations):
        r_top = 11 + idx
        cat_obj = CATEGORIES_REGISTRY_PLT_REGULARES.get(cat_code)
        rule_label = cat_obj.name if cat_obj else cat_code

        ws_dash.cell(row=r_top, column=10, value=rule_label).font = font_regular
        ws_dash.cell(row=r_top, column=10).alignment = alignment_left

        ws_dash.cell(row=r_top, column=11, value=safe_int(count_val)).number_format = '#,##0'
        ws_dash.cell(row=r_top, column=11).alignment = alignment_right

        pct = count_val / max(1, sum(diag.get("category_counts", {}).values()))
        ws_dash.cell(row=r_top, column=12, value=pct).number_format = '0.00%'
        ws_dash.cell(row=r_top, column=12).alignment = alignment_right

        for col_idx in range(10, 13):
            ws_dash.cell(row=r_top, column=col_idx).border = border_thin
            if r_top % 2 == 0:
                ws_dash.cell(row=r_top, column=col_idx).fill = fill_zebra

    # Gráfica Nativa de Excel BarChart
    if len(top_deviations) > 0:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Frecuencia de Desviaciones Críticas Detectadas"
        chart.y_axis.title = "Cantidad de Ocurrencias"
        chart.x_axis.title = "Regla de Consistencia"
        max_r_chart = 10 + len(top_deviations)
        chart_data = Reference(ws_dash, min_col=11, min_row=10, max_row=max_r_chart)
        chart_cats = Reference(ws_dash, min_col=10, min_row=11, max_row=max_r_chart)
        chart.add_data(chart_data, titles_from_data=True)
        chart.set_categories(chart_cats)
        chart.legend = None
        chart.width = 16
        chart.height = 10
        ws_dash.add_chart(chart, "J18")

    # =========================================================================
    # --- HOJA 2: 📋 CATÁLOGO DE ERRORES ---
    # =========================================================================
    ws_cat = wb.create_sheet(title="📋 Catálogo de Errores")
    ws_cat.views.sheetView[0].showGridLines = True

    ws_cat.cell(row=2, column=2, value="REGISTRO MAESTRO DE REGLAS DE CONSISTENCIA PLT REGULARES").font = font_title
    ws_cat.cell(
        row=3, column=2,
        value="Catálogo canónico de inconsistencias geomecánicas ordenadas por frecuencia. Use los hipervínculos para auditar registros."
    ).font = font_subtitle

    all_campaigns = sorted(list(diag.get("campaign_stats", {}).keys()))

    headers_cat = ["N°", "Gravedad", "Regla de Consistencia Geomecánica", "Total Incidencias"] + [f"Campaña {c}" for c in all_campaigns] + ["Navegación / Vínculo"]
    for idx, col in enumerate(headers_cat, start=2):
        cell = ws_cat.cell(row=5, column=idx, value=col)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = alignment_center
        cell.border = border_thin

    anomalies = diag.get("anomalies", [])
    incidencias_por_cat = defaultdict(list)
    for anom in anomalies:
        incidencias_por_cat[anom["category_code"]].append(anom)

    # Construir lista completa de categorías con conteo y ordenar DESCENDENTE por count
    catalog_frequencies = []
    for cat_code, cat_obj in CATEGORIES_REGISTRY_PLT_REGULARES.items():
        matches = incidencias_por_cat.get(cat_code, [])
        catalog_frequencies.append({
            "code": cat_code,
            "obj": cat_obj,
            "count": len(matches),
            "matches": matches
        })

    # Ordenar: Errores (count > 0) ARRIBA de mayor a menor frecuencia
    catalog_frequencies = sorted(catalog_frequencies, key=lambda x: x["count"], reverse=True)

    active_sheets_mapping = {}
    r_cat = 6

    for c_idx, rule_item in enumerate(catalog_frequencies, start=1):
        cat_code = rule_item["code"]
        cat_obj = rule_item["obj"]
        cnt = rule_item["count"]
        matches = rule_item["matches"]

        ws_cat.cell(row=r_cat, column=2, value=c_idx).font = font_regular
        ws_cat.cell(row=r_cat, column=2).alignment = alignment_center
        ws_cat.cell(row=r_cat, column=2).border = border_thin

        c_sev = ws_cat.cell(row=r_cat, column=3, value=cat_obj.severity)
        c_sev.font = font_bold
        c_sev.alignment = alignment_center
        c_sev.border = border_thin
        if cat_obj.severity == "ALERTA":
            c_sev.fill = fill_accent_red
        elif cat_obj.severity == "ADVERTENCIA":
            c_sev.fill = fill_accent_orange
        else:
            c_sev.fill = fill_accent_yellow

        ws_cat.cell(row=r_cat, column=4, value=cat_obj.name).font = font_bold if cnt > 0 else font_regular
        ws_cat.cell(row=r_cat, column=4).border = border_thin

        c_cnt = ws_cat.cell(row=r_cat, column=5, value=cnt)
        c_cnt.font = font_bold
        c_cnt.alignment = alignment_right
        c_cnt.number_format = '#,##0'
        c_cnt.border = border_thin

        # Columnas por campaña
        for y_offset, camp_key in enumerate(all_campaigns):
            camp_cnt = sum(1 for m in matches if str(m.get("campana")) == str(camp_key))
            c_yr = ws_cat.cell(row=r_cat, column=6 + y_offset, value=camp_cnt)
            c_yr.font = font_regular
            c_yr.alignment = alignment_right
            c_yr.number_format = '#,##0'
            c_yr.border = border_thin
            if camp_cnt == 0:
                c_yr.font = Font(name="Segoe UI", size=9, color="AAAAAA")

        link_col = 6 + len(all_campaigns)
        c_link = ws_cat.cell(row=r_cat, column=link_col)
        if cnt > 0:
            tab_name = get_safe_sheet_name(cat_obj.name, c_idx)
            active_sheets_mapping[cat_code] = {
                "tab_name": tab_name,
                "rule_name": cat_obj.name,
                "severity": cat_obj.severity,
                "records": matches
            }
            c_link.value = f'=HYPERLINK("#{chr(39)}{tab_name}{chr(39)}!B2", "👉 Navegar a Registros")'
            c_link.font = s["font_link"]
            c_link.alignment = alignment_center
        else:
            c_link.value = "Limpio / 0 Incidencias"
            c_link.font = Font(name="Segoe UI", size=9, italic=True, color="2E7D32")
            c_link.alignment = alignment_center
            c_link.fill = fill_accent_green

        c_link.border = border_thin
        r_cat += 1

    last_col_letter = get_column_letter(5 + len(all_campaigns) + 1)
    ws_cat.auto_filter.ref = f"B5:{last_col_letter}{r_cat - 1}"

    # =========================================================================
    # --- HOJA: 🗂️ CELDAS ÚNICAS PLT (MUESTRAS Y ENSAYOS REGULARES) ---
    # =========================================================================
    ws_samples = wb.create_sheet(title="🗂️ Celdas Únicas PLT")
    ws_samples.views.sheetView[0].showGridLines = True

    ws_samples.cell(row=2, column=2, value="REGISTRO DE CELDAS Y MUESTRAS ÚNICAS EVALUADAS — PLT REGULARES").font = font_title
    ws_samples.cell(
        row=3, column=2,
        value="Consolidado exclusivo de muestras y tramos con observaciones, alertas o discrepancias QA/QC detectadas."
    ).font = font_subtitle

    c_back_samples = ws_samples.cell(row=2, column=16)
    c_back_samples.value = '=HYPERLINK("#' + "'📋 Catálogo de Errores'" + '!B2", "⬅ Volver al Catálogo de Errores")'
    c_back_samples.font = s["font_back_link"]
    c_back_samples.alignment = alignment_right

    headers_samples = [
        "Fila Excel", "Taladro", "Muestra", "Campaña", "Desde (m)", "Hasta (m)", "Longitud (m)",
        "Lito 1", "Lito 2", "Lito 3", "Tipo Litológico", "Factor K",
        "Carga P (kN)", "D (mm)", "Is(50) (MPa)", "UCS (MPa)", "ISRM (PLT)", "Estado QA/QC", "N° Incidencias"
    ]

    for c_idx, h_text in enumerate(headers_samples, start=2):
        cell = ws_samples.cell(row=5, column=c_idx, value=h_text)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = alignment_center
        cell.border = border_thin

    unique_samples = [s for s in diag.get("unique_samples_plt", []) if s.get("estado") != "CONFORME" or s.get("incidencias_cant", 0) > 0]
    r_samp = 6

    for s_item in unique_samples:
        ws_samples.cell(row=r_samp, column=2, value=s_item.get("row_index")).alignment = alignment_center
        ws_samples.cell(row=r_samp, column=3, value=s_item.get("taladro")).alignment = alignment_left
        ws_samples.cell(row=r_samp, column=4, value=s_item.get("muestra")).alignment = alignment_center
        ws_samples.cell(row=r_samp, column=5, value=s_item.get("campana")).alignment = alignment_center
        
        c_from = ws_samples.cell(row=r_samp, column=6, value=s_item.get("from_m"))
        c_from.alignment = alignment_right
        if s_item.get("from_m") is not None: c_from.number_format = '0.00'

        c_to = ws_samples.cell(row=r_samp, column=7, value=s_item.get("to_m"))
        c_to.alignment = alignment_right
        if s_item.get("to_m") is not None: c_to.number_format = '0.00'

        c_len = ws_samples.cell(row=r_samp, column=8, value=s_item.get("longitud_m"))
        c_len.alignment = alignment_right
        if s_item.get("longitud_m") is not None: c_len.number_format = '0.00'

        ws_samples.cell(row=r_samp, column=9, value=s_item.get("lito1")).alignment = alignment_center
        ws_samples.cell(row=r_samp, column=10, value=s_item.get("lito2")).alignment = alignment_center
        ws_samples.cell(row=r_samp, column=11, value=s_item.get("lito3")).alignment = alignment_center
        ws_samples.cell(row=r_samp, column=12, value=s_item.get("tipo_litologico")).alignment = alignment_left
        ws_samples.cell(row=r_samp, column=13, value=s_item.get("factor_k")).alignment = alignment_center

        c_p = ws_samples.cell(row=r_samp, column=14, value=s_item.get("p_kn"))
        c_p.alignment = alignment_right
        if s_item.get("p_kn") is not None: c_p.number_format = '#,##0.00'

        c_d = ws_samples.cell(row=r_samp, column=15, value=s_item.get("d_mm"))
        c_d.alignment = alignment_right
        if s_item.get("d_mm") is not None: c_d.number_format = '#,##0.00'

        c_is = ws_samples.cell(row=r_samp, column=16, value=s_item.get("is50_mpa"))
        c_is.alignment = alignment_right
        if s_item.get("is50_mpa") is not None: c_is.number_format = '0.00'

        c_ucs = ws_samples.cell(row=r_samp, column=17, value=s_item.get("ucs_mpa"))
        c_ucs.alignment = alignment_right
        if s_item.get("ucs_mpa") is not None: c_ucs.number_format = '0.00'

        ws_samples.cell(row=r_samp, column=18, value=s_item.get("isrm")).alignment = alignment_center
        
        c_st = ws_samples.cell(row=r_samp, column=19, value=s_item.get("estado"))
        c_st.alignment = alignment_center
        c_st.font = font_bold
        if s_item.get("estado") == "CONFORME":
            c_st.fill = fill_accent_green
        elif s_item.get("estado") == "NO CONFORME":
            c_st.fill = fill_accent_red
        else:
            c_st.fill = fill_accent_yellow

        c_cnt = ws_samples.cell(row=r_samp, column=20, value=s_item.get("incidencias_cant", 0))
        c_cnt.alignment = alignment_right
        c_cnt.number_format = '#,##0'

        for col_idx in range(2, 21):
            cell = ws_samples.cell(row=r_samp, column=col_idx)
            cell.border = border_thin
            if cell.font != font_bold:
                cell.font = font_regular
            if r_samp % 2 == 0 and col_idx != 19:
                cell.fill = fill_zebra

        r_samp += 1

    if r_samp > 6:
        ws_samples.auto_filter.ref = f"B5:T{r_samp - 1}"

    # Autoajuste de columnas en Celdas Únicas PLT
    for col in ws_samples.iter_cols(min_col=2, max_col=20, min_row=5, max_row=min(r_samp, 100)):
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_samples.column_dimensions[col_letter].width = max(11, min(max_len + 3, 26))

    # =========================================================================
    # --- HOJA OPCIONAL: 🗂️ CELDAS ÚNICAS LGG (SI SE CARGÓ EL LOGUEO) ---
    # =========================================================================
    unique_lgg = [l for l in diag.get("unique_runs_lgg", []) if l.get("estado") != "CONFORME" or l.get("muestras_plt_cant", 0) > 0]
    if unique_lgg:
        ws_lgg_sheet = wb.create_sheet(title="🗂️ Celdas Únicas LGG")
        ws_lgg_sheet.views.sheetView[0].showGridLines = True

        ws_lgg_sheet.cell(row=2, column=2, value="REGISTRO DE CORRIDAS ÚNICAS EVALUADAS — LOGUEO GENERAL (LGG)").font = font_title
        ws_lgg_sheet.cell(
            row=3, column=2,
            value="Listado de corridas geomecánicas oficiales de sondajes DDH cruzadas contra la planilla de ensayos PLT."
        ).font = font_subtitle

        c_back_lgg = ws_lgg_sheet.cell(row=2, column=11)
        c_back_lgg.value = '=HYPERLINK("#' + "'📋 Catálogo de Errores'" + '!B2", "⬅ Volver al Catálogo de Errores")'
        c_back_lgg.font = s["font_back_link"]
        c_back_lgg.alignment = alignment_right

        headers_lgg = [
            "N°", "Taladro", "Desde (m)", "Hasta (m)", "Longitud (m)",
            "Lito 1", "Lito 2", "Lito 3", "Resistencia ISRM", "Muestras PLT Asociadas"
        ]

        for c_idx, h_text in enumerate(headers_lgg, start=2):
            cell = ws_lgg_sheet.cell(row=5, column=c_idx, value=h_text)
            cell.font = font_header
            cell.fill = fill_primary
            cell.alignment = alignment_center
            cell.border = border_thin

        r_lgg = 6
        for idx_lgg, l_item in enumerate(unique_lgg, start=1):
            ws_lgg_sheet.cell(row=r_lgg, column=2, value=idx_lgg).alignment = alignment_center
            ws_lgg_sheet.cell(row=r_lgg, column=3, value=l_item.get("taladro")).alignment = alignment_left
            
            c_fd = ws_lgg_sheet.cell(row=r_lgg, column=4, value=l_item.get("desde_m"))
            c_fd.alignment = alignment_right
            if l_item.get("desde_m") is not None: c_fd.number_format = '0.00'

            c_th = ws_lgg_sheet.cell(row=r_lgg, column=5, value=l_item.get("hasta_m"))
            c_th.alignment = alignment_right
            if l_item.get("hasta_m") is not None: c_th.number_format = '0.00'

            c_ln = ws_lgg_sheet.cell(row=r_lgg, column=6, value=l_item.get("longitud_m"))
            c_ln.alignment = alignment_right
            if l_item.get("longitud_m") is not None: c_ln.number_format = '0.00'

            ws_lgg_sheet.cell(row=r_lgg, column=7, value=l_item.get("lito1")).alignment = alignment_center
            ws_lgg_sheet.cell(row=r_lgg, column=8, value=l_item.get("lito2")).alignment = alignment_center
            ws_lgg_sheet.cell(row=r_lgg, column=9, value=l_item.get("lito3")).alignment = alignment_center
            ws_lgg_sheet.cell(row=r_lgg, column=10, value=l_item.get("isrm")).alignment = alignment_center

            c_mcount = ws_lgg_sheet.cell(row=r_lgg, column=11, value=l_item.get("muestras_plt_cant", 0))
            c_mcount.alignment = alignment_right
            c_mcount.number_format = '#,##0'
            if l_item.get("muestras_plt_cant", 0) > 0:
                c_mcount.font = font_bold

            for col_idx in range(2, 12):
                cell = ws_lgg_sheet.cell(row=r_lgg, column=col_idx)
                cell.border = border_thin
                if cell.font != font_bold:
                    cell.font = font_regular
                if r_lgg % 2 == 0:
                    cell.fill = fill_zebra

            r_lgg += 1

        if r_lgg > 6:
            ws_lgg_sheet.auto_filter.ref = f"B5:K{r_lgg - 1}"

        for col in ws_lgg_sheet.iter_cols(min_col=2, max_col=11, min_row=5, max_row=min(r_lgg, 100)):
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_lgg_sheet.column_dimensions[col_letter].width = max(11, min(max_len + 3, 24))


    # =========================================================================
    # --- HOJA 3: 📑 DETALLE PLANO COMPLETO DE INCIDENCIAS ---
    # =========================================================================
    ws_detail = wb.create_sheet(title="📑 Detalle de Incidencias")
    ws_detail.views.sheetView[0].showGridLines = True

    ws_detail.cell(row=2, column=2, value="REGISTRO DETALLADO DE TODAS LAS INCIDENCIAS DETECTADAS").font = font_title
    ws_detail.cell(
        row=3, column=2,
        value="Base de datos plana de inconsistencias. Utilice los filtros en los encabezados para auditar registros específicos."
    ).font = font_subtitle

    headers_detail = [
        "ID", "Fila Excel", "Gravedad", "Taladro", "Muestra", "Campaña",
        "From (m)", "To (m)", "Columna de Falla", "Valor Actual", "Código Regla", "Mensaje de Inconsistencia Geomecánica"
    ]

    grid_heading_row = 5
    for idx, col in enumerate(headers_detail, start=2):
        cell = ws_detail.cell(row=grid_heading_row, column=idx, value=col)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = alignment_center
        cell.border = border_thin

    for d_idx, inc in enumerate(anomalies, start=1):
        r_d = grid_heading_row + d_idx
        ws_detail.cell(row=r_d, column=2, value=d_idx).alignment = alignment_center
        ws_detail.cell(row=r_d, column=3, value=safe_int(inc.get("row_index"))).alignment = alignment_center

        c_sev = ws_detail.cell(row=r_d, column=4, value=inc.get("severity", "ALERTA"))
        c_sev.font = font_bold
        c_sev.alignment = alignment_center
        if inc.get("severity") == "ALERTA":
            c_sev.fill = fill_accent_red
        elif inc.get("severity") == "ADVERTENCIA":
            c_sev.fill = fill_accent_orange
        else:
            c_sev.fill = fill_accent_yellow

        ws_detail.cell(row=r_d, column=5, value=str(inc.get("taladro", "—"))).alignment = alignment_left
        ws_detail.cell(row=r_d, column=6, value=str(inc.get("muestra", "—"))).alignment = alignment_center
        ws_detail.cell(row=r_d, column=7, value=str(inc.get("campana", "—"))).alignment = alignment_center
        
        ws_detail.cell(row=r_d, column=8, value=inc.get("from_m") if inc.get("from_m") is not None else "").alignment = alignment_right
        ws_detail.cell(row=r_d, column=9, value=inc.get("to_m") if inc.get("to_m") is not None else "").alignment = alignment_right
        ws_detail.cell(row=r_d, column=10, value=str(inc.get("columna", "—"))).alignment = alignment_left
        ws_detail.cell(row=r_d, column=11, value=str(inc.get("valor_actual", "—"))).alignment = alignment_center
        ws_detail.cell(row=r_d, column=12, value=str(inc.get("category_code", "—"))).alignment = alignment_center
        ws_detail.cell(row=r_d, column=13, value=str(inc.get("message", "—"))).alignment = alignment_left

        for col_idx in range(2, 14):
            cell_d = ws_detail.cell(row=r_d, column=col_idx)
            cell_d.border = border_thin
            if r_d % 2 == 0 and col_idx != 4:
                cell_d.fill = fill_zebra

    end_detail_row = max(grid_heading_row + 1, grid_heading_row + len(anomalies))
    ws_detail.auto_filter.ref = f"B{grid_heading_row}:L{end_detail_row}"

    # =========================================================================
    # --- HOJAS 4+: 🚨 PESTAÑAS INDIVIDUALES POR REGLA DE ERROR ---
    # =========================================================================
    for cat_code, mapping_data in active_sheets_mapping.items():
        sh_name = mapping_data["tab_name"]
        rule_name = mapping_data["rule_name"]
        sev = mapping_data["severity"]
        err_records = mapping_data["records"]

        ws_err = wb.create_sheet(title=sh_name)
        ws_err.views.sheetView[0].showGridLines = True

        # Botón de Retorno al Catálogo
        c_back = ws_err.cell(row=2, column=2)
        c_back.value = '=HYPERLINK("#' + "'📋 Catálogo de Errores'" + '!B2", "⬅ Volver al Catálogo de Errores")'
        c_back.font = Font(name="Segoe UI", size=11, bold=True, color="1B365D", underline="single")

        ws_err.cell(row=4, column=2, value=f"REPORTE ESPECÍFICO: {rule_name.upper()}").font = font_title
        ws_err.cell(row=5, column=2, value=f"Código de Regla: {cat_code} | Severidad: {sev}").font = font_subtitle

        # Tarjeta KPI de Registros Afectados
        write_kpi_card(
            ws_err, 7, 2, "REGISTROS AFECTADOS", f"{len(err_records):,}",
            fill_accent_red if sev == "ALERTA" else (fill_accent_orange if sev == "ADVERTENCIA" else fill_accent_yellow),
            font_kpi_val_red if sev == "ALERTA" else (font_kpi_val_orange if sev == "ADVERTENCIA" else font_kpi_val_blue),
            s
        )

        # Tabla: Distribución por Campaña
        ws_err.cell(row=10, column=2, value="DISTRIBUCIÓN POR CAMPAÑA").font = font_section
        headers_err_camp = ["Campaña", "Muestras Afectadas (N)", "% del Total de esta Inconsistencia"]
        for idx, col in enumerate(headers_err_camp, start=2):
            cell = ws_err.cell(row=11, column=idx, value=col)
            cell.font = font_header
            cell.fill = fill_primary
            cell.alignment = alignment_center
            cell.border = border_thin

        r_err_c = 12
        for camp_key in all_campaigns:
            c_matches = sum(1 for m in err_records if str(m.get("campana")) == str(camp_key))
            if c_matches > 0:
                ws_err.cell(row=r_err_c, column=2, value=str(camp_key)).font = font_bold
                ws_err.cell(row=r_err_c, column=2).alignment = alignment_center

                ws_err.cell(row=r_err_c, column=3, value=c_matches).number_format = '#,##0'
                ws_err.cell(row=r_err_c, column=3).alignment = alignment_right

                ws_err.cell(row=r_err_c, column=4, value=c_matches / len(err_records)).number_format = '0.00%'
                ws_err.cell(row=r_err_c, column=4).alignment = alignment_right

                for col_idx in range(2, 5):
                    ws_err.cell(row=r_err_c, column=col_idx).border = border_thin
                r_err_c += 1

        # Tabla: Registros Individuales Afectados
        r_indiv = r_err_c + 2
        ws_err.cell(row=r_indiv, column=2, value="REGISTROS INDIVIDUALES AFECTADOS (LISTADO COMPLETO)").font = font_section

        headers_indiv = ["N°", "Fila Excel", "Taladro", "Muestra", "Campaña", "From (m)", "To (m)", "Columna de Falla", "Valor Actual", "Diagnóstico Detallado"]
        for idx, col in enumerate(headers_indiv, start=2):
            cell = ws_err.cell(row=r_indiv + 1, column=idx, value=col)
            cell.font = font_header
            cell.fill = fill_primary
            cell.alignment = alignment_center
            cell.border = border_thin

        r_indiv_cur = r_indiv + 2
        for idx_rec, rec in enumerate(err_records, start=1):
            ws_err.cell(row=r_indiv_cur, column=2, value=idx_rec).alignment = alignment_center
            ws_err.cell(row=r_indiv_cur, column=3, value=safe_int(rec.get("row_index"))).alignment = alignment_center
            ws_err.cell(row=r_indiv_cur, column=4, value=str(rec.get("taladro", "—"))).alignment = alignment_left
            ws_err.cell(row=r_indiv_cur, column=5, value=str(rec.get("muestra", "—"))).alignment = alignment_center
            ws_err.cell(row=r_indiv_cur, column=6, value=str(rec.get("campana", "—"))).alignment = alignment_center
            ws_err.cell(row=r_indiv_cur, column=7, value=rec.get("from_m") if rec.get("from_m") is not None else "").alignment = alignment_right
            ws_err.cell(row=r_indiv_cur, column=8, value=rec.get("to_m") if rec.get("to_m") is not None else "").alignment = alignment_right
            ws_err.cell(row=r_indiv_cur, column=9, value=str(rec.get("columna", "—"))).alignment = alignment_left
            ws_err.cell(row=r_indiv_cur, column=10, value=str(rec.get("valor_actual", "—"))).alignment = alignment_center
            ws_err.cell(row=r_indiv_cur, column=11, value=str(rec.get("message", "—"))).alignment = alignment_left

            for col_idx in range(2, 12):
                cell_i = ws_err.cell(row=r_indiv_cur, column=col_idx)
                cell_i.border = border_thin
                if r_indiv_cur % 2 == 0:
                    cell_i.fill = fill_zebra
            r_indiv_cur += 1

        ws_err.auto_filter.ref = f"B{r_indiv + 1}:K{r_indiv_cur - 1}"

    # Auto-ajuste de ancho de columnas para todas las pestañas
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or '')
                if not val_str.startswith("=") and len(val_str) < 60:
                    max_len = max(max_len, len(val_str))
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 13)

    # Guardar Libro Excel con tolerancia a bloqueo por apertura
    try:
        wb.save(output_path)
        print(f">> Reporte Excel 100% idéntico a Mapeo generado con éxito en: {output_path}")
    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        dir_name = os.path.dirname(output_path)
        alt_path = os.path.join(dir_name, f"Reporte_Auditoria_PLT_Regulares_{timestamp}.xlsx")
        wb.save(alt_path)
        print(f">> AVISO: '{output_path}' está abierto en Excel.")
        print(f">> Se guardó una copia con timestamp en: {alt_path}")

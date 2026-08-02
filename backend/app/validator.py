import os
import json
import math
import unicodedata
import re
import openpyxl
from typing import List, Dict, Any, Optional
from app.core.rules import WEATHERING_COMPATIBILITY, MASTER_ERROR_RULES
from app.core.report_config import CAMPOS_EXTRA_A_CAPTURAR
from app.calculator import (
    STRENGTH_RATINGS, WEATHERING_RATINGS_76, WEATHERING_RATINGS_89,
    ROUGHNESS_RATINGS_76, ROUGHNESS_RATINGS_89, FILLING_CLASSES,
    calculate_rqd_rating, calculate_spacing_rating_76, calculate_spacing_rating_89,
    calculate_aperture_rating_76, calculate_aperture_rating_89,
    calculate_filling_rating_76, calculate_filling_rating_89,
    calculate_water_rating, get_rock_class
)

# Normalización robusta para mapeo de cabeceras (remueve acentos, espacios y especiales)
def normalize_text(text: str) -> str:
    if text is None:
        return ""
    s = str(text).lower()
    s = s.replace("≥", "mayorigual").replace("≤", "menorigual")
    s = s.replace(">=", "mayorigual").replace("<=", "menorigual").replace(">", "mayor").replace("<", "menor")
    s = s.replace("∑", "sum").replace("σ", "sum").replace("ς", "sum").replace("Σ", "sum")
    normalized = unicodedata.normalize('NFKD', s).encode('ASCII', 'ignore').decode('utf-8')
    return re.sub(r'[^a-z0-9]', '', normalized).strip()

# Catálogos estándar de validación
VALID_STRUCTURES = {"JN", "F", "RF", "F-10", "SZ", "BED", "VN", "CON", "SE", "F+10", "-1"}
VALID_STRENGTHS = {"R0", "R1", "R2", "R3", "R4", "R5", "R6", "-1"}
VALID_RUGOSITY = {str(x) for x in range(1, 10)}.union({"-1"})
VALID_WEATHERING = {"UWF", "SWD", "MWM", "HWA", "CWC", "RS", "-1"}
VALID_RELLENO = {"ca", "sand", "ch", "cl", "gy", "RXF", "FBX", "GOU", "PAT", "SIO", "QZ", "SU", "OX", "ep", "cwf", "-1"}
VALID_AGUA = {"CDC", "DPH", "WTM", "DGE", "FGF"}
VALID_FORMA = {str(x) for x in range(1, 7)}.union({"-1"})

LGG_PATTERNS = {
    "corrida": ["corrida", "id", "numcorrida"],
    "taladro": ["taladro", "sondaje", "drillhole", "holeid", "taladroid"],
    "de": ["de", "desde", "dem", "desdem", "from", "depthfrom"],
    "a": ["a", "hasta", "am", "hastam", "to", "depthto"],
    "rec_m": ["longitudrecuperadam", "recuperacionm", "recm", "recupm", "recuperacion", "recuperada", "longitudrecuperada", "longitudrecuperdadam", "longitudrecuperdada"],
    "rqd_m": ["rqdm", "rqd", "rqdmfragmentos10cm", "frag10cmm", "sumfrags10cm", "rqdsumfrags10cmm", "fragmentos10cmm", "frags10cmm", "fragmayorigual10cmm", "sumfrag10cm", "fragsmayor10cmm", "fragmayor10cmm", "rqdsumfragsmayorigual10cmm", "sumfragsmayorigual10cmm"],
    "lrf_m": ["longitudrocafracturadam", "lrfm", "lrf", "longitudrocafracturada", "rocafracturadam", "rocafracturada"],
    "small_frag_m": ["sumfrags10cmquenoentranalrqd", "smallfragm", "smallfrag", "sumfrags10cmquenoentranalrqdm", "frags10cmquenoentranalrqdm", "fragmenor10cmm", "fragmentosmenores10cm", "fragmenores10cm", "sumfragsmenor10cmquenoentranalrqdm"],
    "frac_nat": ["ndefracnaturales", "nfracnatur", "nfracnaturales", "fracnat", "fracturasnaturales", "naturales"],
    "lito1": ["lito1", "lito12023", "litologia1", "litologia12023", "litologia"],
    "lito2": ["lito2", "lito22023", "litologia2", "litologia22023"],
    "lito3": ["lito3", "lito32023", "litologia3", "litologia32023"],
    "resistencia": ["resistencia", "resistestimadaisrm", "resistmaxestimadaisrm", "resistestimada", "resistmax", "resist", "dureza", "isrm", "durezamaterial", "resistmaxestimada"],
    "tipo_est1": ["tipodeestruct", "tipoestructura1", "tipoest1", "estructura1", "tipodeestruct1"],
    "tipo_est2": ["tipodeestruct2", "tipoestructura2", "tipoest2", "estructura2"],
    "frac_buz30": ["nfracnatbuz30", "nfracnaturalbuz30", "buz30", "naturalesbuz30", "nfracnatbuzmenor30", "nfracnaturalbuzmenor30", "buzmenor30"],
    "frac_buz60": ["nfracn30buz60", "nfracnatural30buz60", "buz3060", "buz30a60", "nfracn30menorigualbuzmenor60", "nfracn30menorigualbuz60", "nfracnatural30menorigualbuzmenor60", "nfracn30menorbuzmenor60", "nfracn30menorbuz60"],
    "frac_buz90": ["nfracnatbuz60", "nfracnaturalbuz60", "buz60", "nfracnatbuz90", "nfracnaturalbuz90", "buz90", "nfracnatbuzmayor60", "nfracnaturalbuzmayor60", "buzmayor60"],
    "abertura": ["aberturamm", "abertura", "abert"],
    "rugosidad": ["rugosidadisrm", "rugosidad", "rugos"],
    "jrc10": ["jrc10", "jrc", "jrc10rugosidad"],
    "intemperismo": ["gradointempisrm", "gradointemp", "intemperismo", "alteracion", "weathering"],
    "relleno1": ["tipoderelleno1", "relleno1", "tiporelleno1"],
    "relleno2": ["tipoderelleno2", "relleno2", "tiporelleno2"],
    "espesor": ["espesorrellenomm", "espesorrelleno", "espesor", "espesormm"],
    "agua_obs": ["presenciadeaguaisrm", "presenaguaisrm", "presenciaagua", "aguaobs", "agua"],
    "geologo": ["geotecnico", "geotécnico", "geologo", "geotecnic", "geot"],
    "comentarios": ["comentarios", "comentario", "observaciones", "observacion", "comments"],
    "campana": ["campana", "anio", "campan", "campaign", "year"],
    "frf": ["frf", "fracturaspormetro", "fracturasmetro"]
}

EST_PATTERNS = {
    "taladro": ["taladro", "sondaje", "drillhole", "holeid", "taladroid"],
    "de": ["de", "desde", "dem", "desdem", "from", "depthfrom"],
    "a": ["a", "hasta", "am", "hastam", "to", "depthto"],
    "profundidad": ["profundidad", "prof", "depth", "profundidadm"],
    "lito1": ["lito1", "lito12023", "litologia1", "litologia12023", "litologia"],
    "lito2": ["lito2", "lito22023", "litologia2", "litologia22023"],
    "lito3": ["lito3", "lito32023", "litologia3", "litologia32023"],
    "tipo_estructura": ["tipodeestructura", "estructura", "tipoest", "tipodeestructu"],
    "alfa": ["alpha", "alfa"],
    "beta": ["beta"],
    "dip": ["dip"],
    "azimuth": ["azimuth", "azimut"],
    "forma": ["forma"],
    "rugosidad": ["rugosidadisrm", "rugosidad", "rugos"],
    "jrc10": ["jrc10", "jrc", "jrc10rugosidad"],
    "abertura": ["aberturamm", "abertura", "abert"],
    "weathering": ["gradointempisrm", "gradointemp", "intemperismo", "alteracion", "weathering", "gradointemperismo"],
    "espesor": ["espesorrellenomm", "espesorrelleno", "espesor", "espesormm"],
    "relleno1": ["tipoderelleno1", "relleno1", "tiporelleno1"],
    "relleno2": ["tipoderelleno2", "relleno2", "tiporelleno2"],
    "dureza_pared": ["durezadepared", "dureza", "durezapared", "durezaestructural", "durezadelapareddeestructura", "durezadelapareddeestructu"],
    "agua": ["presenciadeaguaisrm", "presenaguaisrm", "presenciaagua", "aguaobs", "agua"],
    "geotecnico": ["geotecnico", "geotécnico", "geologo", "geotecnic", "geot"],
    "comentario": ["comentarios", "comentario", "observaciones", "observacion", "comments", "intervalocomentario"],
    "campana": ["campana", "anio", "campan", "campaign", "year"]
}

RMR_PATTERNS = {
    "sondaje": ["sondaje", "taladro", "drillhole", "holeid"],
    "corrida": ["corrida", "id", "numcorrida"],
    "lito1": ["litho1", "lito1", "litologia1"],
    "lito2": ["litho2", "lito2", "litologia2"],
    "lito3": ["litho3", "lito3", "litologia3"],
    "de": ["desde", "desdem", "de", "from"],
    "a": ["hasta", "hastam", "a", "to"],
    "long_corrida": ["longcorrida", "longcorridam", "longitudcorrida"],
    "rec_m": ["recm", "recupm", "recuperacionm", "longitudrecuperada"],
    "rec_pct": ["recpct", "recporcentaje"],
    "rqd_m": ["rqdm", "rqd"],
    "rqd_pct": ["rqdpct", "rqdporcentaje"],
    "lrf_m": ["longtramofracturadom", "lrfm", "lrf", "longitudrocafracturada"],
    "frf": ["frfzonastrituradas", "frf"],
    "frac_nat": ["fracturasnaturales", "nfracnaturales", "fracnat"],
    "total_frac": ["toraldefracturas", "totaldefracturas", "totalfracturas"],
    "ff_1m": ["ff1m", "ffm"],
    "espaciamiento_mm": ["espaciamientomm", "espaciamiento"],
    "resistencia": ["resistencia", "resistenciainput"],
    "tipo_estructura": ["tipodeestructura", "tipoest"],
    "abertura_mm": ["aberturamm", "abertura"],
    "rugosidad": ["rugosidad"],
    "relleno": ["relleno"],
    "clasificacion_relleno": ["clasificacionrelleno"],
    "intemperismo": ["intemperismo"],
    "jrc10": ["jrc10"],
    "espesor_relleno": ["espesorderelleno", "espesorrelleno"],
    "presencia_agua": ["presenciadeagua", "presenciaagua"],
    "rmr76": ["rmr76", "rmr76total"],
    "rmr89": ["rmr89", "rmr89total"],
    "campana": ["campana", "anio"]
}

COLLAR_PATTERNS = {
    "taladro": ["taladro", "sondaje", "drillhole", "holeid", "taladroid", "hole_id", "hole"],
    "eoh": ["eoh", "profundidad_final", "prof_final", "max_depth", "total_depth"]
}

SURVEY_PATTERNS = {
    "taladro": ["taladro", "sondaje", "drillhole", "holeid", "taladroid", "hole_id", "hole"],
    "depth": ["profundidad", "depth", "depthm", "profundidadm", "prof"]
}

FALLBACK_LGG_MAP = {
    "corrida": 1, "taladro": 2, "de": 3, "a": 4, "rec_m": 5, "rqd_m": 6, "lrf_m": 7, "small_frag_m": 8,
    "frac_nat": 9, "lito1": 10, "lito2": 11, "lito3": 12, "resistencia": 13, "tipo_est1": 14, "tipo_est2": 15,
    "frac_buz30": 16, "frac_buz60": 17, "frac_buz90": 18, "abertura": 19, "rugosidad": 20, "jrc10": 21,
    "intemperismo": 22, "relleno1": 23, "relleno2": 24, "espesor": 25, "agua_obs": 26, "geologo": 27,
    "comentarios": 28, "campana": 29, "frf": 30
}

FALLBACK_EST_MAP = {
    "taladro": 2, "de": 3, "a": 4, "profundidad": 5, "lito1": 6, "lito2": 7, "lito3": 8, "tipo_estructura": 9,
    "alfa": 10, "beta": 11, "dip": 12, "azimuth": 13, "forma": 14, "rugosidad": 15, "jrc10": 16, "abertura": 17,
    "weathering": 18, "espesor": 19, "relleno1": 20, "relleno2": 21, "dureza_pared": 22, "agua": 23,
    "geotecnico": 24, "comentario": 25, "campana": 26
}

FALLBACK_RMR_MAP = {
    "sondaje": 1, "fecha": 2, "logueador": 3, "corrida": 4, "lito1": 5, "lito2": 6, "lito3": 7,
    "de": 8, "a": 9, "long_corrida": 10, "rec_m": 11, "rec_pct": 12, "rqd_m": 13, "rqd_pct": 14,
    "lrf_m": 15, "frf": 16, "frac_nat": 17, "total_frac": 18, "ff_1m": 19, "espaciamiento_mm": 20,
    "resistencia": 21, "tipo_estructura": 22, "abertura_mm": 23, "rugosidad": 24, "relleno": 25,
    "clasificacion_relleno": 26, "intemperismo": 27, "jrc10": 28, "espesor_relleno": 29, "presencia_agua": 30,
    "r76_resistencia": 32, "r76_rqd": 33, "r76_espaciamiento": 34, "r76_abertura": 35,
    "r76_rugosidad": 36, "r76_relleno": 37, "r76_intemperismo": 38, "r76_persistencia": 39,
    "r76_juntas": 40, "r76_agua": 41, "rmr76": 42, "r76_calidad_roca": 43, "r76_litologia": 44,
    "r89_resistencia": 46, "r89_rqd": 47, "r89_espaciamiento": 48, "r89_abertura": 49,
    "r89_rugosidad": 50, "r89_relleno": 51, "r89_intemperismo": 52, "r89_persistencia": 53,
    "r89_juntas": 54, "r89_agua": 55, "rmr89": 56, "r89_calidad_roca": 57, "campana": 58
}

def find_header_row_and_mapping(sheet, keyword_maps) -> tuple:
    best_row = 6
    best_matches = 0
    best_mapping = {}

    cleaned_maps = {}
    for key, patterns in keyword_maps.items():
        cleaned_maps[key] = [normalize_text(p) for p in patterns if p]

    for r in range(1, 21):
        matches = 0
        mapping = {}
        for c in range(1, sheet.max_column + 1):
            val = sheet.cell(row=r, column=c).value
            if val is None:
                continue
            norm_val = normalize_text(str(val))
            if not norm_val:
                continue
            
            for key, patterns in cleaned_maps.items():
                if key in mapping:
                    continue
                if norm_val in patterns:
                    mapping[key] = c
                    matches += 1
                    break
        if matches > best_matches:
            best_matches = matches
            best_row = r
            best_mapping = mapping
            
    return best_row, best_mapping

def safe_float(val, default=0.0):
    if val is None: return default
    try: return float(val)
    except: return default

def safe_int(val, default=0):
    if val is None: return default
    try: return int(float(val))
    except: return default

def safe_str(val, default=""):
    if val is None: return default
    return str(val).strip()

class FaltantesCollector:
    def __init__(self):
        self._registros = []

    def registrar(self, modulo, fila_excel, celda_padre, celda_hija, columna, valor_raw, campania, geotecnico):
        if valor_raw is None:
            tipo = "VACIO"
        else:
            txt = str(valor_raw).strip()
            if txt == "":
                tipo = "VACIO"
            elif txt in ("-1", "-1.0", "-1,0"):
                tipo = "SIN_INFORMACION"
            else:
                return
        self._registros.append({
            "fila_excel": fila_excel,
            "celda_padre": celda_padre,
            "celda_hija": celda_hija,
            "columna": columna,
            "tipo_incidencia": tipo,
            "campania": str(campania) if campania not in (None, "") else "N/A",
            "geotecnico": geotecnico if geotecnico else "N/A",
            "sector_geotecnico": "N/A",
            "modulo": modulo,
        })

    def dump(self):
        return list(self._registros)

def capturar_faltantes_extra(collector, modulo, row_dict, claves, fila_excel, celda_padre, celda_hija, campania, geotecnico):
    if collector is None:
        return
    for clave in claves:
        if clave not in row_dict:
            continue
        collector.registrar(modulo, fila_excel, celda_padre, celda_hija, clave,
                            row_dict.get(clave), campania, geotecnico)

def sanitize_val(val, target_type):
    if val is None:
        return None
    val_str = str(val).strip()
    val_upper = val_str.upper()
    if val_str == "" or val_upper in ["-1", "-1.0", "N/A", "NAN", "NONE", "-", "—", "-1,0"]:
        return None
    if target_type == str:
        return val_str
    try:
        if target_type == int:
            return int(round(float(val)))
        return target_type(val)
    except:
        return None

def get_canonical_value(val, valid_set):
    if val is None:
        return None
    s = str(val).strip()
    s_lower = s.lower()
    for code in valid_set:
        if code.lower() == s_lower:
            return code
    return None

def get_row_dict(ws, row_idx, mapping):
    row_data = {}
    is_empty = True
    for key, col_idx in mapping.items():
        cell_val = ws.cell(row=row_idx, column=col_idx).value
        if cell_val is not None and str(cell_val).strip() != "":
            is_empty = False
        row_data[key] = cell_val
    return row_data, is_empty

def validate_row_qaqc(data: Dict[str, Any]) -> List[Dict[str, str]]:
    """Ejecuta todos los controles de calidad geomecánicos QA/QC en una corrida de LGG en caliente."""
    alerts = []
    try:
        de = float(data.get("de", 0.0))
        a = float(data.get("a", 0.0))
        perf = round(a - de, 2)
        
        rec_val = float(data.get("rec_m", 0.0))
        rec_m = 0.0 if rec_val < 0 else rec_val
        rqd_val = float(data.get("rqd_m", 0.0))
        rqd_m = 0.0 if rqd_val < 0 else rqd_val
        lrf_val = float(data.get("lrf_m", 0.0))
        lrf_m = 0.0 if lrf_val < 0 else lrf_val
        small_val = float(data.get("small_frag_m", 0.0))
        small_frag_m = 0.0 if small_val < 0 else small_val
        
        fn_val = int(data.get("frac_nat", 0))
        frac_nat = 0 if fn_val < 0 else fn_val
        b30_val = int(data.get("frac_buz30", 0))
        buz30 = 0 if b30_val < 0 else b30_val
        buz60_val = int(data.get("frac_buz60", 0))
        buz60 = 0 if buz60_val < 0 else buz60_val
        buz90_val = int(data.get("frac_buz90", 0))
        buz90 = 0 if buz90_val < 0 else buz90_val
        
        resistencia = data.get("resistencia", "R4")
        weathering = data.get("intemperismo", "UWF")
        aperture = float(data.get("abertura", 0.0))
        thickness = float(data.get("espesor", 0.0))
        tipo_est1 = data.get("tipo_est1", "JN")
        tipo_est2 = data.get("tipo_est2", "")

        if perf <= 0:
            alerts.append({"type": "CRITICAL", "field": "a", "message": f"Profundidad final ({a}m) debe ser mayor a la inicial ({de}m)."})
        elif perf > 1.6:
            alerts.append({"type": "CRITICAL", "field": "a", "message": f"La longitud de corrida ({perf}m) excede el límite máximo de perforación de 1.6m."})
        
        if rec_m > perf:
            alerts.append({"type": "CRITICAL", "field": "rec_m", "message": f"Longitud recuperada ({rec_m}m) es físicamente mayor que la perforada ({perf}m)."})
        
        if rqd_m > rec_m:
            alerts.append({"type": "CRITICAL", "field": "rqd_m", "message": f"El metraje de RQD ({rqd_m}m) no puede ser mayor que la longitud recuperada ({rec_m}m)."})
        
        if lrf_m > rec_m:
            alerts.append({"type": "CRITICAL", "field": "lrf_m", "message": f"La longitud de roca fracturada LRF ({lrf_m}m) no puede ser mayor que la longitud recuperada ({rec_m}m)."})
            
        sum_frags = round(rqd_m + lrf_m + small_frag_m, 2)
        if sum_frags > perf:
            alerts.append({"type": "CRITICAL", "field": "rqd_m", "message": f"La suma de fragmentos físicos ({sum_frags}m) supera el avance total de la corrida ({perf}m)."})
            
        sum_bins = buz30 + buz60 + buz90
        if sum_bins != frac_nat:
            alerts.append({"type": "WARNING", "field": "frac_nat", "message": f"La suma de fracturas naturales clasificadas por buzamiento ({sum_bins}) no coincide con el conteo general ({frac_nat})."})
            
        exceptions = {"F", "RF", "VN", "SZ", "F+10", "BED"}
        if thickness > aperture and (tipo_est1 not in exceptions and tipo_est2 not in exceptions):
            alerts.append({"type": "CRITICAL", "field": "espesor", "message": f"El espesor de relleno ({thickness}mm) no puede ser mayor que la abertura de junta ({aperture}mm) excepto en estructuras F, RF, VN, SZ, F+10 o BED."})

        if thickness > 0 and aperture <= 0:
            alerts.append({"type": "WARNING", "field": "abertura", "message": f"Se ha registrado un espesor de relleno de {thickness}mm, pero la abertura de junta es 0mm."})
        elif thickness == 0 and aperture > 0:
            alerts.append({"type": "WARNING", "field": "espesor", "message": f"La abertura de junta es {aperture}mm, pero no se ha registrado espesor de relleno."})
            
        valid_weatherings = WEATHERING_COMPATIBILITY.get(resistencia)
        if valid_weatherings and weathering not in valid_weatherings:
            alerts.append({"type": "WARNING", "field": "intemperismo", "message": f"Incompatibilidad geológica: Roca con resistencia {resistencia} no puede registrar intemperismo {weathering}. Permitidos: {', '.join(valid_weatherings)}."})
    except Exception as e:
        alerts.append({"type": "CRITICAL", "field": "global", "message": f"Error al procesar reglas de consistencia: {str(e)}"})
    return alerts

from collections import defaultdict

def find_rmr_sheet_name(sheetnames: list) -> Optional[str]:
    for s in sheetnames:
        norm = normalize_text(s)
        if norm in ['validacionrmr', 'validacion_rmr', 'rmr', 'bdrmr', 'logueormr', 'hoja1']:
            return s
    for s in sheetnames:
        norm = normalize_text(s)
        if 'rmr' in norm:
            return s
    return None

def validate_rmr_sheet_data(
    ws_rmr,
    lgg_runs: list,
    resumen_celdas: dict,
    incidencias: list,
    r_counters: dict,
    custom_map: dict = None,
    collector: FaltantesCollector = None
) -> int:
    """
    Realiza el escaneo y auditoría geomecánica masiva de la hoja 'Validación_RMR' (o similar)
    verificando la congruencia de los 48 campos contra LGG y las reglas de Reglas.md.
    Clasifica datos ausentes en:
    - 'VACIO': Celda sin contenido (None o '').
    - 'SIN_INFORMACION': Celda con valor '-1' (Sin dato registrado).
    """
    h_idx, rmr_map = find_header_row_and_mapping(ws_rmr, RMR_PATTERNS)
    for k, v in FALLBACK_RMR_MAP.items():
        if k not in rmr_map:
            rmr_map[k] = v
    if custom_map:
        for k, v in custom_map.items():
            if v is not None and isinstance(v, int) and v >= 0:
                rmr_map[k] = v + 1

    rmr_extra = list(CAMPOS_EXTRA_A_CAPTURAR["Validación RMR"])

    print(f"[*] Iniciando escaneo de Validación_RMR. Fila inicial: {h_idx + 1}, Fila máxima: {ws_rmr.max_row}", flush=True)

    lgg_by_taladro = defaultdict(list)
    for run in lgg_runs:
        t_name = run.get("taladro")
        if t_name:
            lgg_by_taladro[t_name].append(run)

    rmr_by_taladro = defaultdict(list)
    empty_streak = 0
    current_taladro = None
    total_rmr_filas = 0

    for r in range(h_idx + 1, ws_rmr.max_row + 1):
        if r % 500 == 0:
            print(f"  ... [RMR] Fila {r} / {ws_rmr.max_row}", flush=True)
            
        row_dict, is_empty = get_row_dict(ws_rmr, r, rmr_map)

        t_val = row_dict.get("sondaje") or row_dict.get("taladro")
        if t_val is not None and str(t_val).strip() != "":
            current_taladro = safe_str(t_val)
        else:
            row_dict["sondaje"] = current_taladro

        if is_empty:
            empty_streak += 1
            if empty_streak >= 20:
                print(f"[*] [RMR] Freno de emergencia en fila {r}.", flush=True)
                break
            continue

        if not current_taladro:
            continue

        empty_streak = 0
        total_rmr_filas += 1
        taladro = current_taladro
        corrida_num = safe_int(row_dict.get("corrida", 0))
        celda_padre = taladro
        celda_hija = f"{taladro}-RMR{corrida_num if corrida_num > 0 else r}"

        if isinstance(resumen_celdas, dict):
            if celda_padre not in resumen_celdas:
                resumen_celdas[celda_padre] = {
                    "total_hijas": 0, "vacios": 0, "sin_informacion": 0,
                    "advertencias": 0, "alertas": 0, "estado_celda": "OK",
                    "dist_celda": 0.0, "campania": "N/A"
                }
            resumen_celdas[celda_padre]["total_hijas"] += 1

        row_has_errors = False

        def reg_err_rmr(col, val, tipo, msg):
            nonlocal row_has_errors
            incidencias.append({
                "fila_excel": r, "celda_padre": celda_padre, "celda_hija": celda_hija,
                "columna": col, "valor_actual": val, "tipo_incidencia": tipo, "mensaje": msg,
                "campania": safe_str(row_dict.get("campana")) or "N/A", "geotecnico": "N/A", "sector_geotecnico": "N/A",
                "modulo": "Validación RMR"
            })
            if tipo == "VACIO":
                r_counters["total_vacios"] += 1
                if isinstance(resumen_celdas, dict) and celda_padre in resumen_celdas:
                    resumen_celdas[celda_padre]["vacios"] += 1
            elif tipo == "SIN_INFORMACION":
                r_counters["total_sin_informacion"] += 1
                if isinstance(resumen_celdas, dict) and celda_padre in resumen_celdas:
                    if "sin_informacion" not in resumen_celdas[celda_padre]:
                        resumen_celdas[celda_padre]["sin_informacion"] = 0
                    resumen_celdas[celda_padre]["sin_informacion"] += 1
            elif tipo == "ADVERTENCIA":
                r_counters["total_advertencias"] += 1
                if isinstance(resumen_celdas, dict) and celda_padre in resumen_celdas:
                    resumen_celdas[celda_padre]["advertencias"] += 1
            elif tipo == "ALERTA":
                r_counters["total_alertas"] += 1
                if isinstance(resumen_celdas, dict) and celda_padre in resumen_celdas:
                    resumen_celdas[celda_padre]["alertas"] += 1
                row_has_errors = True

        # 1. CLASIFICACIÓN DE CAMPOS VACÍOS VS SIN INFORMACIÓN (-1)
        mandatory_rmr_fields = [
            ("sondaje", "Sondaje"), ("corrida", "Corrida"), ("logueador", "Logueador"),
            ("de", "Desde (m)"), ("a", "Hasta (m)"),
            ("long_corrida", "Long. Corrida (m)"), ("lito1", "Litho 1"), ("rec_m", "Rec (m)"),
            ("rec_pct", "Rec (%)"), ("rqd_m", "RQD (m)"), ("rqd_pct", "RQD (%)"),
            ("lrf_m", "Long. Tramo fracturado (m)"), ("frf", "FRF (zonas trituradas)"),
            ("frac_nat", "Fracturas naturales"), ("total_frac", "Total de Fracturas"),
            ("ff_1m", "FF/1m"), ("espaciamiento_mm", "Espaciamiento (mm)"), ("resistencia", "Resistencia"),
            ("tipo_estructura", "Tipo de Estructura"), ("abertura_mm", "Abertura (mm)"),
            ("rugosidad", "Rugosidad"), ("relleno", "Relleno"), ("clasificacion_relleno", "Clasificación Relleno"),
            ("intemperismo", "Intemperismo"), ("jrc10", "JRC10"), ("espesor_relleno", "Espesor de relleno")
        ]

        for field_key, field_lbl in mandatory_rmr_fields:
            val_raw = row_dict.get(field_key)
            if val_raw is None or str(val_raw).strip() == "":
                reg_err_rmr(field_key, None, "VACIO", f"El campo obligatorio '{field_lbl}' se encuentra vacío.")
            elif str(val_raw).strip() in ["-1", "-1.0", "-1,0"]:
                reg_err_rmr(field_key, val_raw, "SIN_INFORMACION", f"El campo obligatorio '{field_lbl}' no contiene información (-1).")

        capturar_faltantes_extra(collector, "Validación RMR", row_dict, rmr_extra, r,
                                 celda_padre, celda_hija, row_dict.get("campana"), None)

        # 2. MATCHING CON CORRIDA LGG
        rmr_by_taladro[taladro].append(row_dict)
        lgg_runs_for_t = lgg_by_taladro.get(taladro, [])

        matching_lgg_run = None
        if corrida_num > 0:
            for l_run in lgg_runs_for_t:
                if l_run.get("corrida") == corrida_num:
                    matching_lgg_run = l_run
                    break

        de = sanitize_val(row_dict.get("de"), float)
        a = sanitize_val(row_dict.get("a"), float)

        if matching_lgg_run is None and de is not None and a is not None:
            for l_run in lgg_runs_for_t:
                if abs(l_run.get("de", 0.0) - de) < 0.05 and abs(l_run.get("a", 0.0) - a) < 0.05:
                    matching_lgg_run = l_run
                    break

        if matching_lgg_run is None:
            reg_err_rmr("corrida", corrida_num, "ALERTA", f"Corrida en Validación RMR (Corrida #{corrida_num}) no coincide con ninguna corrida registrada en LGG para el taladro '{taladro}'.")
        else:
            # 3. REGLAS CRÍTICAS DE COINCIDENCIA E INSUMOS
            lgg_de = matching_lgg_run.get("de", 0.0)
            lgg_a = matching_lgg_run.get("a", 0.0)
            if de is not None and a is not None:
                if abs(de - lgg_de) > 0.001 or abs(a - lgg_a) > 0.001:
                    reg_err_rmr("de", f"{de}-{a}", "ALERTA", f"Intervalo Desde/Hasta ({de}m - {a}m) en RMR no coincide exactamente con el intervalo de LGG ({lgg_de}m - {lgg_a}m).")

            long_corrida = sanitize_val(row_dict.get("long_corrida"), float)
            if de is not None and a is not None and long_corrida is not None:
                expected_long = round(a - de, 2)
                if long_corrida <= 0:
                    reg_err_rmr("long_corrida", long_corrida, "ALERTA", f"La Longitud de Corrida ({long_corrida}m) debe ser mayor a 0.")
                if abs(long_corrida - expected_long) > 0.1:
                    reg_err_rmr("long_corrida", long_corrida, "ALERTA", f"La Longitud de Corrida ({long_corrida}m) no coincide con (Hasta - Desde = {expected_long}m) dentro de la tolerancia de 0.1m.")

            rmr_l1 = safe_str(row_dict.get("lito1"))
            rmr_l2 = safe_str(row_dict.get("lito2"))
            rmr_l3 = safe_str(row_dict.get("lito3"))
            lgg_l1 = safe_str(matching_lgg_run.get("lito1"))
            lgg_l2 = safe_str(matching_lgg_run.get("lito2"))
            lgg_l3 = safe_str(matching_lgg_run.get("lito3"))
            if rmr_l1 != lgg_l1 or rmr_l2 != lgg_l2 or rmr_l3 != lgg_l3:
                reg_err_rmr("lito1", f"{rmr_l1}/{rmr_l2}/{rmr_l3}", "ALERTA", f"Combinación litológica en RMR ({rmr_l1}, {rmr_l2}, {rmr_l3}) no coincide con LGG ({lgg_l1}, {lgg_l2}, {lgg_l3}).")

            rec_m = sanitize_val(row_dict.get("rec_m"), float)
            lgg_rec = matching_lgg_run.get("rec_m", 0.0)
            if rec_m is not None and abs(rec_m - lgg_rec) > 0.001:
                reg_err_rmr("rec_m", rec_m, "ALERTA", f"Rec (m) en RMR ({rec_m}m) no coincide con la Recuperación de LGG ({lgg_rec}m).")
            rec_pct = sanitize_val(row_dict.get("rec_pct"), float)
            if rec_m is not None and long_corrida is not None and long_corrida > 0 and rec_pct is not None:
                expected_rec_pct = round((rec_m / long_corrida) * 100)
                if abs(rec_pct - expected_rec_pct) > 1.0:
                    reg_err_rmr("rec_pct", rec_pct, "ALERTA", f"Rec (%) en RMR ({rec_pct}%) no coincide con la fórmula Rec(m)/Long.Corrida(m). Datos evaluados -> Rec({rec_m}m) / Long.Corrida({long_corrida}m) * 100 = {expected_rec_pct}%.")

            rqd_m = sanitize_val(row_dict.get("rqd_m"), float)
            lgg_rqd = matching_lgg_run.get("rqd_m", 0.0)
            if rqd_m is not None and abs(rqd_m - lgg_rqd) > 0.001:
                reg_err_rmr("rqd_m", rqd_m, "ALERTA", f"RQD (m) en RMR ({rqd_m}m) no coincide con el RQD de LGG ({lgg_rqd}m).")
            rqd_pct = sanitize_val(row_dict.get("rqd_pct"), float)
            if rqd_m is not None and long_corrida is not None and long_corrida > 0 and rqd_pct is not None:
                expected_rqd_pct = round((rqd_m / long_corrida) * 100)
                if abs(rqd_pct - expected_rqd_pct) > 1.0:
                    reg_err_rmr("rqd_pct", rqd_pct, "ALERTA", f"RQD (%) en RMR ({rqd_pct}%) no coincide con la fórmula RQD(m)/Long.Corrida(m). Datos evaluados -> RQD({rqd_m}m) / Long.Corrida({long_corrida}m) * 100 = {expected_rqd_pct}%.")

            lrf_m = sanitize_val(row_dict.get("lrf_m"), float)
            lgg_lrf = matching_lgg_run.get("lrf_m", 0.0)
            if lrf_m is not None and abs(lrf_m - lgg_lrf) > 0.001:
                reg_err_rmr("lrf_m", lrf_m, "ALERTA", f"Longitud de Tramo Fracturado ({lrf_m}m) no coincide con LRF de LGG ({lgg_lrf}m).")

            frf = sanitize_val(row_dict.get("frf"), int)
            lgg_frf = matching_lgg_run.get("frf")
            if frf is not None and lgg_frf is not None and frf != lgg_frf:
                reg_err_rmr("frf", frf, "ALERTA", f"FRF en RMR ({frf}) no coincide con FRF de LGG ({lgg_frf}).")

            frac_nat = sanitize_val(row_dict.get("frac_nat"), int)
            lgg_fn = matching_lgg_run.get("frac_nat")
            if frac_nat is not None and lgg_fn is not None and frac_nat != lgg_fn:
                reg_err_rmr("frac_nat", frac_nat, "ALERTA", f"Fracturas Naturales en RMR ({frac_nat}) no coincide con Frac Nat de LGG ({lgg_fn}).")

            total_frac = sanitize_val(row_dict.get("total_frac"), float)
            if frf is not None and frac_nat is not None and total_frac is not None:
                expected_total_frac = round(frf + frac_nat)
                if abs(total_frac - expected_total_frac) > 0.01:
                    reg_err_rmr("total_frac", total_frac, "ALERTA", f"Total de Fracturas en RMR ({total_frac}) no coincide con FRF + FracNat. Datos evaluados -> FRF({frf}) + FracNat({frac_nat}) = {expected_total_frac}.")

            ff_1m = sanitize_val(row_dict.get("ff_1m"), float)
            if total_frac is not None and long_corrida is not None and long_corrida > 0 and ff_1m is not None:
                expected_ff_1m = round(total_frac / long_corrida)
                if abs(ff_1m - expected_ff_1m) > 1.0:
                    reg_err_rmr("ff_1m", ff_1m, "ALERTA", f"FF/1m en RMR ({ff_1m}) no coincide con TotalFracturas / Long.Corrida. Datos evaluados -> TotalFracturas({total_frac}) / Long.Corrida({long_corrida}m) = {expected_ff_1m}.")

            espaciamiento = sanitize_val(row_dict.get("espaciamiento_mm"), float)
            if total_frac is not None and long_corrida is not None and long_corrida > 0 and espaciamiento is not None:
                if round(total_frac) == 0:
                    expected_esp = round(long_corrida * 1000)
                else:
                    expected_esp = round(long_corrida * 1000 / total_frac)
                if abs(espaciamiento - expected_esp) > 2.0:
                    reg_err_rmr("espaciamiento_mm", espaciamiento, "ALERTA", f"Espaciamiento ({espaciamiento}mm) no coincide con la fórmula calculada. Datos evaluados -> Long.Corrida({long_corrida}m) * 1000 / TotalFracturas({total_frac}) = {expected_esp}mm.")

            # OMITIDO A SOLICITUD: No se registran incidencias por dependencias vacías/cálculos negativos en RMR

            resistencia = safe_str(row_dict.get("resistencia"))
            lgg_res = safe_str(matching_lgg_run.get("resistencia"))
            if resistencia and lgg_res and resistencia.upper() != lgg_res.upper():
                reg_err_rmr("resistencia", resistencia, "ALERTA", f"Resistencia en RMR ({resistencia}) no coincide con LGG ({lgg_res}).")

            tipo_est = safe_str(row_dict.get("tipo_estructura"))
            lgg_tipo_est = safe_str(matching_lgg_run.get("tipo_est1"))
            if tipo_est and lgg_tipo_est and tipo_est.upper() != lgg_tipo_est.upper():
                reg_err_rmr("tipo_estructura", tipo_est, "ALERTA", f"Tipo de Estructura en RMR ({tipo_est}) no coincide con LGG ({lgg_tipo_est}).")

            abertura = sanitize_val(row_dict.get("abertura_mm"), float)
            lgg_ab = matching_lgg_run.get("abertura", 0.0)
            if abertura is not None and abs(abertura - lgg_ab) > 0.001:
                reg_err_rmr("abertura_mm", abertura, "ALERTA", f"Abertura en RMR ({abertura}mm) no coincide con LGG ({lgg_ab}mm).")

            rugosidad = safe_str(row_dict.get("rugosidad"))
            lgg_rug = safe_str(matching_lgg_run.get("rugosidad"))
            if rugosidad and lgg_rug and rugosidad != lgg_rug:
                reg_err_rmr("rugosidad", rugosidad, "ALERTA", f"Rugosidad en RMR ({rugosidad}) no coincide con LGG ({lgg_rug}).")

            relleno = safe_str(row_dict.get("relleno"))
            lgg_rel = safe_str(matching_lgg_run.get("relleno1"))
            if relleno and lgg_rel and relleno.lower() != lgg_rel.lower():
                reg_err_rmr("relleno", relleno, "ALERTA", f"Relleno en RMR ({relleno}) no coincide con LGG ({lgg_rel}).")

            clasif_relleno = sanitize_val(row_dict.get("clasificacion_relleno"), int)
            if relleno:
                expected_class = FILLING_CLASSES.get(relleno.strip().lower(), 1)
                if clasif_relleno is not None and clasif_relleno != expected_class:
                    reg_err_rmr("clasificacion_relleno", clasif_relleno, "ALERTA", f"Clasificación de Relleno ({clasif_relleno}) no coincide con el código '{relleno}' (Clase esperada: {expected_class}).")

            intemperismo = safe_str(row_dict.get("intemperismo"))
            lgg_int = safe_str(matching_lgg_run.get("intemperismo"))
            if intemperismo and lgg_int and intemperismo.upper() != lgg_int.upper():
                reg_err_rmr("intemperismo", intemperismo, "ALERTA", f"Intemperismo en RMR ({intemperismo}) no coincide con LGG ({lgg_int}).")

            jrc10 = sanitize_val(row_dict.get("jrc10"), int)
            lgg_jrc = matching_lgg_run.get("jrc10")
            if jrc10 is not None and lgg_jrc is not None and jrc10 != lgg_jrc:
                reg_err_rmr("jrc10", jrc10, "ALERTA", f"JRC10 en RMR ({jrc10}) no coincide con LGG ({lgg_jrc}).")

            espesor_rel = sanitize_val(row_dict.get("espesor_relleno"), float)
            lgg_esp = matching_lgg_run.get("espesor", 0.0)
            if espesor_rel is not None and abs(espesor_rel - lgg_esp) > 0.001:
                reg_err_rmr("espesor_relleno", espesor_rel, "ALERTA", f"Espesor de relleno en RMR ({espesor_rel}mm) no coincide con LGG ({lgg_esp}mm).")

            # REGLA PRESENCIA DE AGUA SEGÚN PROFUNDIDAD HASTA (m)
            if a is not None:
                expected_water_code = "CDC" if a < 92.0 else ("DPH" if a < 97.0 else "WTM")
                pres_agua_code = safe_str(row_dict.get("presencia_agua")).upper().strip()
                if pres_agua_code and pres_agua_code != "-1" and pres_agua_code != expected_water_code:
                    reg_err_rmr("presencia_agua", pres_agua_code, "ALERTA", f"Presencia de Agua en RMR ('{pres_agua_code}') no coincide con la tabla de profundidad para Hasta = {a}m (Código esperado: '{expected_water_code}').")

            # 4. REGLAS RATINGS RMR'76 (SUMA PURA DE SUB-RATINGS REGISTRADOS)
            rmr76_excel = sanitize_val(row_dict.get("rmr76"), float)
            r76_res = sanitize_val(row_dict.get("r76_resistencia"), float) or 0.0
            r76_rqd = sanitize_val(row_dict.get("r76_rqd"), float) or 0.0
            r76_esp = sanitize_val(row_dict.get("r76_espaciamiento"), float) or 0.0
            r76_juntas = sanitize_val(row_dict.get("r76_juntas"), float) or 0.0
            r76_agua = sanitize_val(row_dict.get("r76_agua"), float) or 0.0
            expected_rmr76 = r76_res + r76_rqd + r76_esp + r76_juntas + r76_agua

            if rmr76_excel is not None:
                if rmr76_excel < 0 or rmr76_excel > 100:
                    reg_err_rmr("rmr76", rmr76_excel, "ALERTA", f"Puntaje RMR'76 ({rmr76_excel}) fuera del rango permitido de 0 a 100.")
                if abs(rmr76_excel - expected_rmr76) > 0.5:
                    reg_err_rmr("rmr76", rmr76_excel, "ALERTA", f"Descuadre en RMR'76: Excel registra {rmr76_excel}, pero la suma de sub-ratings registrados es {expected_rmr76}. Desglose -> Resistencia({r76_res}) + RQD({r76_rqd}) + Espaciamiento({r76_esp}) + Condición de Juntas({r76_juntas}) + Presencia de Agua({r76_agua}) = {expected_rmr76}.")

            # 5. REGLAS RATINGS RMR'89 (SUMA PURA DE SUB-RATINGS REGISTRADOS)
            rmr89_excel = sanitize_val(row_dict.get("rmr89"), float)
            r89_res = sanitize_val(row_dict.get("r89_resistencia"), float) or 0.0
            r89_rqd = sanitize_val(row_dict.get("r89_rqd"), float) or 0.0
            r89_esp = sanitize_val(row_dict.get("r89_espaciamiento"), float) or 0.0
            r89_juntas = sanitize_val(row_dict.get("r89_juntas"), float) or 0.0
            r89_agua = sanitize_val(row_dict.get("r89_agua"), float) or 0.0
            expected_rmr89 = r89_res + r89_rqd + r89_esp + r89_juntas + r89_agua

            if rmr89_excel is not None:
                if rmr89_excel < 0 or rmr89_excel > 100:
                    reg_err_rmr("rmr89", rmr89_excel, "ALERTA", f"Puntaje RMR'89 ({rmr89_excel}) fuera del rango permitido de 0 a 100.")
                if abs(rmr89_excel - expected_rmr89) > 0.5:
                    reg_err_rmr("rmr89", rmr89_excel, "ALERTA", f"Descuadre en RMR'89: Excel registra {rmr89_excel}, pero la suma de sub-ratings registrados es {expected_rmr89}. Desglose -> Resistencia({r89_res}) + RQD({r89_rqd}) + Espaciamiento({r89_esp}) + Condición de Juntas({r89_juntas}) + Presencia de Agua({r89_agua}) = {expected_rmr89}.")

        if not row_has_errors:
            r_counters["total_ok"] += 1

    # REGLA CRÍTICA: CANTIDAD TOTAL DE CORRIDAS RMR VS LGG
    for t_name, lgg_runs_for_t in lgg_by_taladro.items():
        rmr_runs_for_t = rmr_by_taladro.get(t_name, [])
        if len(lgg_runs_for_t) > 0 and len(rmr_runs_for_t) != len(lgg_runs_for_t):
            msg = f"Inconsistencia total de corridas: El taladro '{t_name}' registra {len(rmr_runs_for_t)} corridas en Validación RMR vs {len(lgg_runs_for_t)} corridas en LGG."
            incidencias.append({
                "fila_excel": 0, "celda_padre": t_name, "celda_hija": t_name,
                "columna": "Total Corridas", "valor_actual": len(rmr_runs_for_t),
                "tipo_incidencia": "ALERTA", "mensaje": msg,
                "campania": "N/A", "geotecnico": "N/A", "sector_geotecnico": "N/A",
                "modulo": "Validación RMR"
            })
            if t_name in resumen_celdas:
                resumen_celdas[t_name]["alertas"] += 1
                r_counters["total_alertas"] += 1

    return total_rmr_filas

def validate_logueo_bulk_sheets(file_path: str, lgg_sheet: str, est_sheet: str, output_json_path: str):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    try:
        _validate_logueo_bulk_sheets_core(wb, lgg_sheet, est_sheet, output_json_path)
    finally:
        try: wb.close()
        except: pass

def _validate_logueo_bulk_sheets_core(wb, lgg_sheet: str, est_sheet: str, output_json_path: str):
    if lgg_sheet not in wb.sheetnames:
        raise ValueError(f"La hoja de LGG '{lgg_sheet}' no existe en el archivo.")
    if est_sheet not in wb.sheetnames:
        raise ValueError(f"La hoja de Estructural '{est_sheet}' no existe en el archivo.")
        
    ws_lgg = wb[lgg_sheet]
    ws_est = wb[est_sheet]

    lgg_header, lgg_map = find_header_row_and_mapping(ws_lgg, LGG_PATTERNS)
    lgg_claves_detectadas = set(lgg_map.keys())
    for k, v in FALLBACK_LGG_MAP.items():
        if k not in lgg_map:
            lgg_map[k] = v

    est_header, est_map = find_header_row_and_mapping(ws_est, EST_PATTERNS)
    est_claves_detectadas = set(est_map.keys())
    for k, v in FALLBACK_EST_MAP.items():
        if k not in est_map:
            est_map[k] = v

    lgg_extra = [k for k in CAMPOS_EXTRA_A_CAPTURAR["LGG"] if k in lgg_claves_detectadas]
    est_extra = [k for k in CAMPOS_EXTRA_A_CAPTURAR["Estructural"] if k in est_claves_detectadas]
    collector = FaltantesCollector()

    incidencias = []
    lgg_runs = []
    
    total_lgg_filas = 0
    total_est_filas = 0
    
    total_vacios = 0
    total_advertencias = 0
    total_alertas = 0
    total_ok = 0

    resumen_celdas = {}
    filas_por_campana = {}
    filas_por_geotecnico = {}
    last_a_by_taladro = {}

    print(f"[*] Iniciando escaneo de LGG. Fila inicial: {lgg_header + 1}, Fila máxima detectada: {ws_lgg.max_row}", flush=True)
    empty_streak_lgg = 0
    current_taladro_lgg = None
    
    for r in range(lgg_header + 1, ws_lgg.max_row + 1):
        if r % 50 == 0:
            print(f"  ... [LGG] Procesando fila {r} / {ws_lgg.max_row}", flush=True)
            
        row_dict, is_empty = get_row_dict(ws_lgg, r, lgg_map)
        
        t_val = row_dict.get("taladro")
        if t_val is not None and str(t_val).strip() != "":
            current_taladro_lgg = safe_str(t_val)
        else:
            row_dict["taladro"] = current_taladro_lgg
            
        if is_empty:
            empty_streak_lgg += 1
            if empty_streak_lgg >= 20:
                print(f"[*] [LGG] Se detectaron 20 filas vacías consecutivas. Terminando lectura de LGG en la fila {r}.", flush=True)
                break
            continue
            
        if not current_taladro_lgg:
            continue
            
        empty_streak_lgg = 0
        total_lgg_filas += 1
        
        taladro = current_taladro_lgg
        corrida_num = safe_int(row_dict.get("corrida", 0))
        camp = sanitize_val(row_dict.get("campana"), int)
        geo = sanitize_val(row_dict.get("geologo"), str)
        sector = "N/A"
        
        celda_padre = taladro
        celda_hija = f"{taladro}-C{corrida_num}"

        if camp: filas_por_campana[str(camp)] = filas_por_campana.get(str(camp), 0) + 1
        if geo: filas_por_geotecnico[geo] = filas_por_geotecnico.get(geo, 0) + 1

        if celda_padre not in resumen_celdas:
            resumen_celdas[celda_padre] = {
                "total_hijas": 0, "vacios": 0, "advertencias": 0, "alertas": 0,
                "estado_celda": "OK", "dist_celda": 0.0, "campania": str(camp) if camp else "N/A"
            }

        resumen_celdas[celda_padre]["total_hijas"] += 1
        row_has_errors = False

        def registrar_lgg_error(col, val, tipo, msg):
            nonlocal total_vacios, total_advertencias, total_alertas, row_has_errors
            incidencias.append({
                "fila_excel": r, "celda_padre": celda_padre, "celda_hija": celda_hija,
                "columna": col, "valor_actual": val, "tipo_incidencia": tipo, "mensaje": msg,
                "campania": str(camp) if camp else "N/A", "geotecnico": geo if geo else "N/A", "sector_geotecnico": sector,
                "modulo": "LGG"
            })
            if tipo == "VACIO":
                total_vacios += 1
                resumen_celdas[celda_padre]["vacios"] += 1
            elif tipo == "ADVERTENCIA":
                total_advertencias += 1
                resumen_celdas[celda_padre]["advertencias"] += 1
            elif tipo == "ALERTA":
                total_alertas += 1
                resumen_celdas[celda_padre]["alertas"] += 1
                row_has_errors = True

        mandatory_lgg = [
            "corrida", "de", "a", "rec_m", "rqd_m", "lrf_m", "small_frag_m", "frac_nat", "lito1", 
            "resistencia", "tipo_est1", "frac_buz30", "frac_buz60", "frac_buz90", 
            "abertura", "rugosidad", "jrc10", "intemperismo", "relleno1", "espesor", 
            "agua_obs", "campana", "geologo"
        ]
        for key in mandatory_lgg:
            v_san = sanitize_val(row_dict.get(key), str)
            if v_san is None:
                registrar_lgg_error(key, None, "VACIO", f"El campo obligatorio '{key}' se encuentra vacío o es -1.")

        capturar_faltantes_extra(collector, "LGG", row_dict, lgg_extra, r, celda_padre, celda_hija, camp, geo)

        de = sanitize_val(row_dict.get("de"), float)
        a = sanitize_val(row_dict.get("a"), float)

        rec_m = sanitize_val(row_dict.get("rec_m"), float)
        rqd_m = sanitize_val(row_dict.get("rqd_m"), float)
        lrf_m = sanitize_val(row_dict.get("lrf_m"), float)
        small_frag_m = sanitize_val(row_dict.get("small_frag_m"), float)

        frac_nat = sanitize_val(row_dict.get("frac_nat"), int)
        b30 = sanitize_val(row_dict.get("frac_buz30"), int)
        b60 = sanitize_val(row_dict.get("frac_buz60"), int)
        b90 = sanitize_val(row_dict.get("frac_buz90"), int)

        abertura = sanitize_val(row_dict.get("abertura"), float)
        espesor = sanitize_val(row_dict.get("espesor"), float)

        # Validaciones de valores negativos (se ejecutan solo si el valor existe)
        raw_de = row_dict.get("de")
        if de is not None and de < 0:
            registrar_lgg_error("de", raw_de, "ALERTA", f"El valor de 'de:' ({de}m) no puede ser negativo.")
        raw_a = row_dict.get("a")
        if a is not None and a < 0:
            registrar_lgg_error("a", raw_a, "ALERTA", f"El valor de 'a:' ({a}m) no puede ser negativo.")
        raw_rec = row_dict.get("rec_m")
        if rec_m is not None and rec_m < 0:
            registrar_lgg_error("rec_m", raw_rec, "ALERTA", f"La longitud recuperada ({rec_m}m) no puede ser negativa.")
        raw_rqd = row_dict.get("rqd_m")
        if rqd_m is not None and rqd_m < 0:
            registrar_lgg_error("rqd_m", raw_rqd, "ALERTA", f"El metraje RQD ({rqd_m}m) no puede ser negativo.")
        raw_lrf = row_dict.get("lrf_m")
        if lrf_m is not None and lrf_m < 0:
            registrar_lgg_error("lrf_m", raw_lrf, "ALERTA", f"La longitud de roca fracturada LRF ({lrf_m}m) no puede ser negativa.")
        raw_small = row_dict.get("small_frag_m")
        if small_frag_m is not None and small_frag_m < 0:
            registrar_lgg_error("small_frag_m", raw_small, "ALERTA", f"El metraje de fragmentos <10cm ({small_frag_m}m) no puede ser negativo.")

        raw_frac_nat = row_dict.get("frac_nat")
        # OMITIDO A SOLICITUD: No se registran incidencias por valores negativos de frac_nat por dependencias vacías
        raw_b30 = row_dict.get("frac_buz30")
        if b30 is not None and b30 < 0:
            registrar_lgg_error("frac_buz30", raw_b30, "ALERTA", f"El número de fracturas en Buz<30° ({b30}) no puede ser negativo.")
        raw_b60 = row_dict.get("frac_buz60")
        if b60 is not None and b60 < 0:
            registrar_lgg_error("frac_buz60", raw_b60, "ALERTA", f"El número de fracturas en 30°-60° ({b60}) no puede ser negativo.")
        raw_b90 = row_dict.get("frac_buz90")
        if b90 is not None and b90 < 0:
            registrar_lgg_error("frac_buz90", raw_b90, "ALERTA", f"El número de fracturas en Buz>60° ({b90}) no puede ser negativo.")

        raw_abertura = row_dict.get("abertura")
        if abertura is not None and abertura < 0:
            registrar_lgg_error("abertura", raw_abertura, "ALERTA", f"La abertura de junta ({abertura}mm) no puede ser negativa.")
        raw_espesor = row_dict.get("espesor")
        if espesor is not None and espesor < 0:
            registrar_lgg_error("espesor", raw_espesor, "ALERTA", f"El espesor de relleno ({espesor}mm) no puede ser negativo.")
        raw_camp = row_dict.get("campana")
        if camp is not None and camp < 0:
            registrar_lgg_error("campana", raw_camp, "ALERTA", f"El año de campaña ({camp}) no puede ser negativo.")

        for key, val_raw in [("frac_nat", raw_frac_nat), ("frac_buz30", raw_b30), ("frac_buz60", raw_b60), ("frac_buz90", raw_b90)]:
            if val_raw is not None and str(val_raw).strip() not in ["-1", "-1.0", ""]:
                try:
                    f_val = float(val_raw)
                    if not f_val.is_integer():
                        msg_int = "El número de fracturas naturales debe ser un número entero." if key == "frac_nat" else f"El campo '{key}' ({val_raw}) debe ser un número entero."
                        registrar_lgg_error(key, val_raw, "ALERTA", msg_int)
                except ValueError:
                    pass

        if "frf" in lgg_map:
            frf_raw = row_dict.get("frf")
            frf_val = sanitize_val(frf_raw, int)
            if frf_val is not None and str(frf_raw).strip() not in ["-1", "-1.0", ""]:
                if frf_val >= 0:
                    try:
                        f_frf = float(frf_raw)
                        if not f_frf.is_integer():
                            registrar_lgg_error("frf", frf_raw, "ALERTA", "El valor de FRF debe ser un número entero.")
                    except ValueError:
                        pass
                    if lrf_m is not None:
                        calc_frf = math.floor(round(lrf_m * 100) / 5) + 1 if lrf_m > 0 else 0
                        if frf_val != calc_frf:
                            registrar_lgg_error("frf", frf_raw, "ALERTA", f"El valor de FRF ({frf_val}) no coincide con el calculado por la fórmula: FRF = PISO( REDOND(LRF * 100) / 5 ) + 1 (si LRF > 0, sino 0). Calculado: {calc_frf} basado en LRF ({lrf_m}m).")

        if a is not None:
            resumen_celdas[celda_padre]["dist_celda"] = max(resumen_celdas[celda_padre]["dist_celda"], a)

        if de is not None and celda_padre in last_a_by_taladro:
            prev_a = last_a_by_taladro[celda_padre]
            if abs(de - prev_a) > 0.001:
                gap = round(abs(de - prev_a), 4)
                registrar_lgg_error("de", de, "ALERTA", f"Ruptura de continuidad espacial detectada. Datos evaluados -> Profundidad de inicio 'De': {de}m, Profundidad final anterior 'A': {prev_a}m (Brecha calculada: {gap}m).")
        if a is not None:
            last_a_by_taladro[celda_padre] = a

        if de is not None and a is not None:
            perf = round(a - de, 2)
            if perf <= 0:
                registrar_lgg_error("a", a, "ALERTA", f"Longitud de corrida perforada debe ser positiva. Datos evaluados -> De: {de}m, A: {a}m, Avance calculado: {perf}m.")
            
            if rec_m is not None and round(rec_m, 4) > round(perf, 4):
                registrar_lgg_error("rec_m", rec_m, "ALERTA", f"La longitud recuperada es mayor que el avance perforado. Datos evaluados -> Recuperada: {rec_m}m, Avance de corrida: {perf}m (De: {de}m, A: {a}m).")
                
            if rqd_m is not None and rec_m is not None and round(rqd_m, 4) > round(rec_m, 4):
                registrar_lgg_error("rqd_m", rqd_m, "ALERTA", f"Metraje RQD es mayor que la longitud recuperada. Datos evaluados -> RQD: {rqd_m}m, Recuperada: {rec_m}m, Avance de corrida: {perf}m (De: {de}m, A: {a}m).")

            if lrf_m is not None and rec_m is not None and round(lrf_m, 4) > round(rec_m, 4):
                registrar_lgg_error("lrf_m", lrf_m, "ALERTA", f"La longitud de roca fracturada LRF es mayor que la longitud recuperada. Datos evaluados -> LRF: {lrf_m}m, Recuperada: {rec_m}m, Avance de corrida: {perf}m (De: {de}m, A: {a}m).")

            if rqd_m is not None and lrf_m is not None and small_frag_m is not None:
                sum_frags = round(rqd_m + lrf_m + small_frag_m, 2)
                if round(sum_frags, 4) > round(perf, 4):
                    registrar_lgg_error("rqd_m", rqd_m, "ALERTA", f"La suma de fragmentos físicos supera el avance perforado. Datos evaluados -> Suma de fragmentos: {sum_frags}m (RQD: {rqd_m}m + LRF: {lrf_m}m + <10cm: {small_frag_m}m), Avance de corrida: {perf}m (De: {de}m, A: {a}m), Longitud Recuperada: {rec_m}m.")

        if b30 is not None and b60 is not None and b90 is not None and frac_nat is not None:
            sum_bins = b30 + b60 + b90
            if sum_bins != frac_nat:
                registrar_lgg_error("frac_nat", frac_nat, "ADVERTENCIA", f"La sumatoria de fracturas por buzamiento no coincide con el conteo general. Datos evaluados -> Conteo General (Frac_Nat): {frac_nat}, Suma por buzamiento: {sum_bins} (Buz <30°: {b30} + 30°-60°: {b60} + >60°: {b90}).")

        raw_frac_nat = row_dict.get("frac_nat")
        if frac_nat is not None and frac_nat < 0:
            registrar_lgg_error("frac_nat", raw_frac_nat, "ALERTA", f"El número de fracturas naturales ({frac_nat}) no puede ser negativo.")
        raw_b30 = row_dict.get("frac_buz30")
        if b30 is not None and b30 < 0:
            registrar_lgg_error("frac_buz30", raw_b30, "ALERTA", f"El número de fracturas en Buz<30° ({b30}) no puede ser negativo.")
        raw_b60 = row_dict.get("frac_buz60")
        if b60 is not None and b60 < 0:
            registrar_lgg_error("frac_buz60", raw_b60, "ALERTA", f"El número de fracturas en 30°-60° ({b60}) no puede ser negativo.")
        raw_b90 = row_dict.get("frac_buz90")
        if b90 is not None and b90 < 0:
            registrar_lgg_error("frac_buz90", raw_b90, "ALERTA", f"El número de fracturas en Buz>60° ({b90}) no puede ser negativo.")

        abertura = safe_float(sanitize_val(row_dict.get("abertura"), float))
        espesor = safe_float(sanitize_val(row_dict.get("espesor"), float))

        raw_abertura = row_dict.get("abertura")
        if abertura is not None and abertura < 0:
            registrar_lgg_error("abertura", raw_abertura, "ALERTA", f"La abertura de junta ({abertura}mm) no puede ser negativa.")
        raw_espesor = row_dict.get("espesor")
        if espesor is not None and espesor < 0:
            registrar_lgg_error("espesor", raw_espesor, "ALERTA", f"El espesor de relleno ({espesor}mm) no puede ser negativo.")
        raw_camp = row_dict.get("campana")
        if camp is not None and camp < 0:
            registrar_lgg_error("campana", raw_camp, "ALERTA", f"El año de campaña ({camp}) no puede ser negativo.")

        for key, val_raw in [("frac_nat", raw_frac_nat), ("frac_buz30", raw_b30), ("frac_buz60", raw_b60), ("frac_buz90", raw_b90)]:
            if val_raw is not None and val_raw != -1:
                try:
                    f_val = float(val_raw)
                    if not f_val.is_integer():
                        registrar_lgg_error(key, val_raw, "ALERTA", f"El campo '{key}' ({val_raw}) debe ser un número entero.")
                except ValueError:
                    pass

        if "frf" in lgg_map:
            frf_raw = row_dict.get("frf")
            frf_val = sanitize_val(frf_raw, int)
            if frf_val is not None and frf_val != -1:
                if frf_val < 0:
                    registrar_lgg_error("frf", frf_raw, "ALERTA", f"El valor de FRF ({frf_val}) no puede ser negativo.")
                try:
                    f_frf = float(frf_raw)
                    if not f_frf.is_integer():
                        registrar_lgg_error("frf", frf_raw, "ALERTA", f"El valor de FRF ({frf_raw}) debe ser un número entero.")
                except ValueError:
                    pass
                calc_frf = math.floor(round(lrf_m * 100) / 5) if lrf_m > 0 else 0
                if frf_val != calc_frf:
                    registrar_lgg_error("frf", frf_raw, "ALERTA", f"El valor de FRF ({frf_val}) no coincide con el calculado por la fórmula ({calc_frf}) basado en LRF ({lrf_m}m).")

        resumen_celdas[celda_padre]["dist_celda"] = max(resumen_celdas[celda_padre]["dist_celda"], a)

        if celda_padre in last_a_by_taladro:
            prev_a = last_a_by_taladro[celda_padre]
            if abs(de - prev_a) > 0.001:
                gap = round(abs(de - prev_a), 4)
                registrar_lgg_error("de", de, "ALERTA", f"Ruptura de continuidad espacial detectada. Datos evaluados -> Profundidad de inicio 'De': {de}m, Profundidad final anterior 'A': {prev_a}m (Brecha calculada: {gap}m).")
        last_a_by_taladro[celda_padre] = a

        if perf <= 0:
            registrar_lgg_error("a", a, "ALERTA", f"Longitud de corrida perforada debe ser positiva. Datos evaluados -> De: {de}m, A: {a}m, Avance calculado: {perf}m.")
        elif perf > 1.6:
            registrar_lgg_error("a", a, "ALERTA", f"Longitud de corrida perforada excede el límite crítico de 1.6m. Datos evaluados -> De: {de}m, A: {a}m, Avance calculado: {perf}m.")
        
        if rec_m is not None and rec_m > perf:
            registrar_lgg_error("rec_m", rec_m, "ALERTA", f"La longitud recuperada es mayor que el avance perforado. Datos evaluados -> Recuperada: {rec_m}m, Avance de corrida: {perf}m (De: {de}m, A: {a}m).")
            
        if rqd_m is not None and rec_m is not None and rqd_m > rec_m:
            registrar_lgg_error("rqd_m", rqd_m, "ALERTA", f"Metraje RQD es mayor que la longitud recuperada. Datos evaluados -> RQD: {rqd_m}m, Recuperada: {rec_m}m, Avance de corrida: {perf}m (De: {de}m, A: {a}m).")

        if lrf_m is not None and rec_m is not None and lrf_m > rec_m:
            registrar_lgg_error("lrf_m", lrf_m, "ALERTA", f"La longitud de roca fracturada LRF es mayor que la longitud recuperada. Datos evaluados -> LRF: {lrf_m}m, Recuperada: {rec_m}m, Avance de corrida: {perf}m (De: {de}m, A: {a}m).")

        if rqd_m is not None and lrf_m is not None and small_frag_m is not None:
            sum_frags = round(rqd_m + lrf_m + small_frag_m, 2)
            if sum_frags > perf:
                registrar_lgg_error("rqd_m", rqd_m, "ALERTA", f"La suma de fragmentos físicos supera el avance perforado. Datos evaluados -> Suma de fragmentos: {sum_frags}m (RQD: {rqd_m}m + LRF: {lrf_m}m + <10cm: {small_frag_m}m), Avance de corrida: {perf}m (De: {de}m, A: {a}m), Longitud Recuperada: {rec_m}m.")

        if b30 is not None and b60 is not None and b90 is not None and frac_nat is not None:
            sum_bins = b30 + b60 + b90
            if sum_bins != frac_nat:
                registrar_lgg_error("frac_nat", frac_nat, "ADVERTENCIA", f"La sumatoria de fracturas por buzamiento no coincide con el conteo general. Datos evaluados -> Conteo General (Frac_Nat): {frac_nat}, Suma por buzamiento: {sum_bins} (Buz <30°: {b30} + 30°-60°: {b60} + >60°: {b90}).")

        tipo_est1 = safe_str(sanitize_val(row_dict.get("tipo_est1"), str))
        tipo_est2 = safe_str(sanitize_val(row_dict.get("tipo_est2"), str))
        
        raw_resistencia = row_dict.get("resistencia")
        resistencia_can = get_canonical_value(raw_resistencia, VALID_STRENGTHS)
        if raw_resistencia is not None and not resistencia_can:
            registrar_lgg_error("resistencia", raw_resistencia, "ALERTA", f"Código de Resistencia ISRM no válido. Permitidos: {', '.join(VALID_STRENGTHS)}")
            resistencia = "R4"
        else:
            resistencia = resistencia_can or "R4"

        raw_intemperismo = row_dict.get("intemperismo")
        intemperismo_can = get_canonical_value(raw_intemperismo, VALID_WEATHERING)
        if raw_intemperismo is not None and not intemperismo_can:
            registrar_lgg_error("intemperismo", raw_intemperismo, "ALERTA", f"Código de Meteorización no válido. Permitidos: {', '.join(VALID_WEATHERING)}")
            weathering = "UWF"
        else:
            weathering = intemperismo_can or "UWF"

        raw_relleno1 = row_dict.get("relleno1")
        relleno1_can = get_canonical_value(raw_relleno1, VALID_RELLENO)
        if raw_relleno1 is not None and not relleno1_can:
            registrar_lgg_error("relleno1", raw_relleno1, "ALERTA", f"Código de Tipo de Relleno no válido. Permitidos: {', '.join(VALID_RELLENO)}")
            relleno1 = None
        else:
            relleno1 = relleno1_can

        raw_agua_obs = row_dict.get("agua_obs")
        agua_obs_can = get_canonical_value(raw_agua_obs, VALID_AGUA)
        if raw_agua_obs is not None and not agua_obs_can:
            registrar_lgg_error("agua_obs", raw_agua_obs, "ALERTA", f"Código de Presencia de Agua no válido. Permitidos: {', '.join(VALID_AGUA)}")

        jrc10 = safe_int(sanitize_val(row_dict.get("jrc10"), int), -1)
        if jrc10 != -1 and (jrc10 < 0 or jrc10 > 20):
            registrar_lgg_error("jrc10", jrc10, "ALERTA", "El valor de JRC10 es inválido. No se permiten valores mayores a 20.")
            
        raw_rugosidad = row_dict.get("rugosidad")
        rugosidad_can = get_canonical_value(raw_rugosidad, VALID_RUGOSITY)
        if raw_rugosidad is not None and not rugosidad_can:
            registrar_lgg_error("rugosidad", raw_rugosidad, "ALERTA", f"Código de Rugosidad no válido. Permitidos: {', '.join(VALID_RUGOSITY)}")

        tipo_est1_can = get_canonical_value(tipo_est1, VALID_STRUCTURES)
        if tipo_est1 and not tipo_est1_can:
            if tipo_est1.upper() == "J":
                tipo_est1 = "JN"
                registrar_lgg_error("tipo_est1", "J", "ADVERTENCIA", "Código de estructura 'J' reconocido como 'JN' (Junta).")
            else:
                registrar_lgg_error("tipo_est1", tipo_est1, "ALERTA", f"Código de estructura 1 no válido. Permitidos: {', '.join(VALID_STRUCTURES)}")
                tipo_est1 = "JN"
        else:
            tipo_est1 = tipo_est1_can or "JN"

        tipo_est2_can = get_canonical_value(tipo_est2, VALID_STRUCTURES)
        if tipo_est2 and not tipo_est2_can:
            registrar_lgg_error("tipo_est2", tipo_est2, "ALERTA", f"Código de estructura 2 no válido. Permitidos: {', '.join(VALID_STRUCTURES)}")
            tipo_est2 = ""
        else:
            tipo_est2 = tipo_est2_can or ""

        exceptions = {"F", "RF", "VN", "SZ", "F+10", "BED"}
        if espesor > abertura and (tipo_est1 not in exceptions and tipo_est2 not in exceptions):
            registrar_lgg_error("espesor", espesor, "ALERTA", f"El espesor de relleno no puede ser mayor que la abertura de junta excepto en estructuras F, RF, VN, SZ, F+10 o BED. Datos evaluados -> Espesor: {espesor}mm (Tipo Relleno: '{relleno1}'), Abertura de Junta: {abertura}mm, Estructuras: '{tipo_est1}' / '{tipo_est2}'.")

        if espesor > 0 and abertura <= 0:
            registrar_lgg_error("abertura", abertura, "ADVERTENCIA", f"Se declaró espesor de relleno de junta pero la abertura es 0mm. Datos evaluados -> Espesor: {espesor}mm (Tipo Relleno: '{relleno1}'), Abertura de Junta: {abertura}mm.")
        elif espesor == 0 and abertura > 0 and relleno1 not in [None, "-1"]:
            registrar_lgg_error("espesor", espesor, "ADVERTENCIA", f"La abertura de junta es mayor a 0mm pero no se ha registrado espesor de relleno. Datos evaluados -> Abertura de Junta: {abertura}mm, Espesor: {espesor}mm (Tipo Relleno: '{relleno1}').")

        lgg_runs.append({
            "taladro": taladro, "de": de, "a": a, "corrida": corrida_num,
            "resistencia": resistencia,
            "lito1": safe_str(row_dict.get("lito1")),
            "lito2": safe_str(row_dict.get("lito2")),
            "lito3": safe_str(row_dict.get("lito3")),
            "rec_m": rec_m,
            "rqd_m": rqd_m,
            "lrf_m": lrf_m,
            "small_frag_m": small_frag_m,
            "frac_nat": frac_nat,
            "frf": sanitize_val(row_dict.get("frf"), int),
            "abertura": abertura,
            "rugosidad": raw_rugosidad if raw_rugosidad is not None else row_dict.get("rugosidad"),
            "jrc10": jrc10,
            "intemperismo": weathering if 'weathering' in locals() else row_dict.get("intemperismo"),
            "relleno1": relleno1 if 'relleno1' in locals() else row_dict.get("relleno1"),
            "espesor": espesor,
            "tipo_est1": tipo_est1 if 'tipo_est1' in locals() else row_dict.get("tipo_est1")
        })

        if not row_has_errors:
            total_ok += 1

    # 2. VALIDAR HOJA ESTRUCTURAL
    print(f"[*] Iniciando escaneo de Estructural. Fila inicial: {est_header + 1}, Fila máxima detectada: {ws_est.max_row}", flush=True)
    empty_streak_est = 0
    current_taladro_est = None
    
    for r in range(est_header + 1, ws_est.max_row + 1):
        if r % 50 == 0:
            print(f"  ... [EST] Procesando fila {r} / {ws_est.max_row}", flush=True)
            
        row_dict, is_empty = get_row_dict(ws_est, r, est_map)
        
        t_val = row_dict.get("taladro")
        if t_val is not None and str(t_val).strip() != "":
            current_taladro_est = safe_str(t_val)
        else:
            row_dict["taladro"] = current_taladro_est
            
        if is_empty:
            empty_streak_est += 1
            if empty_streak_est >= 20:
                print(f"[*] [EST] Se detectaron 20 filas vacías consecutivas. Terminando lectura de Estructural en la fila {r}.", flush=True)
                break
            continue

        if not current_taladro_est:
            continue

        empty_streak_est = 0
        total_est_filas += 1
        taladro = current_taladro_est
        
        # --- PARSEO Y SANITIZACIÓN INMEDIATA (ESTRUCTURAL) ---
        depth = sanitize_val(row_dict.get("profundidad"), float)
        camp = sanitize_val(row_dict.get("campana"), int)
        geo = sanitize_val(row_dict.get("geotecnico"), str)
        
        abertura = sanitize_val(row_dict.get("abertura"), float)
        espesor = sanitize_val(row_dict.get("espesor"), float)
        
        dip = sanitize_val(row_dict.get("dip"), float)
        azimuth = sanitize_val(row_dict.get("azimuth"), float)
        
        est_de = sanitize_val(row_dict.get("de"), float)
        est_a = sanitize_val(row_dict.get("a"), float)
        
        alfa = sanitize_val(row_dict.get("alfa"), float)
        beta = sanitize_val(row_dict.get("beta"), float)
        jrc10 = sanitize_val(row_dict.get("jrc10"), int)
        sector = "N/A"

        celda_padre = taladro
        celda_hija = f"{taladro}-E{r}"

        if camp: filas_por_campana[str(camp)] = filas_por_campana.get(str(camp), 0) + 1
        if geo: filas_por_geotecnico[geo] = filas_por_geotecnico.get(geo, 0) + 1

        if celda_padre not in resumen_celdas:
            resumen_celdas[celda_padre] = {
                "total_hijas": 0, "vacios": 0, "advertencias": 0, "alertas": 0,
                "estado_celda": "OK", "dist_celda": 0.0, "campania": str(camp) if camp else "N/A"
            }

        resumen_celdas[celda_padre]["total_hijas"] += 1
        row_has_errors = False

        def registrar_est_error(col, val, tipo, msg):
            nonlocal total_vacios, total_advertencias, total_alertas, row_has_errors
            incidencias.append({
                "fila_excel": r, "celda_padre": celda_padre, "celda_hija": celda_hija,
                "columna": col, "valor_actual": val, "tipo_incidencia": tipo, "mensaje": msg,
                "campania": str(camp) if camp else "N/A", "geotecnico": geo if geo else "N/A", "sector_geotecnico": sector,
                "modulo": "Estructural"
            })
            if tipo == "VACIO":
                total_vacios += 1
                resumen_celdas[celda_padre]["vacios"] += 1
            elif tipo == "ADVERTENCIA":
                total_advertencias += 1
                resumen_celdas[celda_padre]["advertencias"] += 1
            elif tipo == "ALERTA":
                total_alertas += 1
                resumen_celdas[celda_padre]["alertas"] += 1
                row_has_errors = True

        mandatory_est = [
            "profundidad", "alfa", "beta", "forma", "rugosidad", "jrc10", "abertura", 
            "weathering", "espesor", "relleno1", "dureza_pared", "agua", "geotecnico", "campana",
            "dip", "azimuth"
        ]
        for key in mandatory_est:
            v_san = sanitize_val(row_dict.get(key), str)
            if v_san is None:
                registrar_est_error(key, None, "VACIO", f"El campo obligatorio '{key}' se encuentra vacío o es -1.")

        capturar_faltantes_extra(collector, "Estructural", row_dict, est_extra, r, celda_padre, celda_hija, camp, geo)

        raw_depth = row_dict.get("profundidad")
        if depth is not None and depth < 0:
            registrar_est_error("profundidad", raw_depth, "ALERTA", f"Profundidad ({depth}m) no puede ser negativa.")
        raw_abertura = row_dict.get("abertura")
        if abertura is not None and abertura < 0:
            registrar_est_error("abertura", raw_abertura, "ALERTA", f"La abertura ({abertura}mm) no puede ser negativa.")
        raw_espesor = row_dict.get("espesor")
        if espesor is not None and espesor < 0:
            registrar_est_error("espesor", raw_espesor, "ALERTA", f"El espesor de relleno ({espesor}mm) no puede ser negativo.")
        raw_camp = row_dict.get("campana")
        if camp is not None and camp < 0:
            registrar_est_error("campana", raw_camp, "ALERTA", f"El año de campaña ({camp}) no puede ser negativo.")

        raw_dip = row_dict.get("dip")
        if dip is not None:
            if dip < 0.0 or dip > 90.0:
                registrar_est_error("dip", raw_dip, "ALERTA", f"El ángulo Dip es inválido. Datos evaluados -> Dip: {dip}°. Debe estar entre 0° y 90°.")

        raw_azimuth = row_dict.get("azimuth")
        if azimuth is not None:
            if azimuth < 0.0 or azimuth > 360.0:
                registrar_est_error("azimuth", raw_azimuth, "ALERTA", f"El ángulo Azimut es inválido. Datos evaluados -> Azimut: {azimuth}°. Debe estar entre 0° y 360°.")

        matching_run = None
        if depth is not None:
            for run in lgg_runs:
                if run["taladro"] == taladro and run["de"] <= depth <= run["a"]:
                    matching_run = run
                    break

            if matching_run is None:
                registrar_est_error("profundidad", depth, "ALERTA", f"Profundidad huérfana de junta no corresponde a ningún tramo de corrida en LGG. Datos evaluados -> Profundidad de Junta: {depth}m, Taladro: '{taladro}'.")

        if est_de is not None and est_a is not None:
            has_exact_match = False
            for run in lgg_runs:
                if run["taladro"] == taladro and abs(run["de"] - est_de) < 0.001 and abs(run["a"] - est_a) < 0.001:
                    has_exact_match = True
                    break
            if not has_exact_match:
                registrar_est_error("de", raw_de, "ALERTA", f"La corrida asociada (de/a) no existe de forma exacta en las corridas de LGG para el taladro. Datos evaluados -> Tramo Estructural de corrida: {est_de}m - {est_a}m, Taladro: '{taladro}'.")
            if depth is not None and (depth < est_de or depth > est_a):
                registrar_est_error("profundidad", depth, "ALERTA", f"La profundidad se encuentra fuera del tramo de corrida especificado. Datos evaluados -> Profundidad de Junta: {depth}m, Tramo especificado: {est_de}m - {est_a}m.")

        if alfa is not None:
            if alfa < 0.0 or alfa > 90.0:
                registrar_est_error("alfa", alfa, "ALERTA", f"El ángulo Alfa es inválido. Datos evaluados -> Alfa: {alfa}°. Debe estar entre 0° y 90° o ser -1.")
            elif not float(alfa).is_integer():
                registrar_est_error("alfa", alfa, "ADVERTENCIA", f"El ángulo Alfa debería ser un número entero. Datos evaluados -> Alfa: {alfa}°.")

        if beta is not None:
            if beta < 0.0 or beta > 360.0:
                registrar_est_error("beta", beta, "ALERTA", f"El ángulo Beta es inválido. Datos evaluados -> Beta: {beta}°. Debe estar entre 0° y 360° o ser -1.")
            elif not float(beta).is_integer():
                registrar_est_error("beta", beta, "ADVERTENCIA", f"El ángulo Beta debería ser un número entero. Datos evaluados -> Beta: {beta}°.")

        if jrc10 is not None:
            if jrc10 > 20:
                registrar_est_error("jrc10", jrc10, "ALERTA", f"El valor de JRC10 es inválido. No se permiten valores mayores a 20. Datos evaluados -> JRC10: {jrc10}.")
            elif jrc10 < 0:
                registrar_est_error("jrc10", jrc10, "ALERTA", f"El valor de JRC10 no puede ser negativo. Datos evaluados -> JRC10: {jrc10}.")

        raw_forma = sanitize_val(row_dict.get("forma"), str)
        if raw_forma is not None:
            forma_can = get_canonical_value(raw_forma, VALID_FORMA)
            if not forma_can:
                registrar_est_error("forma", raw_forma, "ALERTA", f"Forma de junta no válida. Permitidos: Plano (1) a Irregular (6). Datos evaluados -> Forma: '{raw_forma}'.")

        tipo_est = safe_str(sanitize_val(row_dict.get("tipo_estructura"), str))
        if espesor is not None and abertura is not None:
            if espesor > abertura and tipo_est not in exceptions:
                registrar_est_error("espesor", espesor, "ALERTA", f"El espesor de relleno no puede ser mayor que la abertura de junta excepto en estructuras F, RF, VN, SZ, F+10 o BED. Datos evaluados -> Espesor: {espesor}mm (Tipo Relleno: '{relleno1}'), Abertura de Junta: {abertura}mm, Estructura: '{tipo_est}'.")

            raw_relleno1 = row_dict.get("relleno1")
            relleno1_can = get_canonical_value(raw_relleno1, VALID_RELLENO)
            if raw_relleno1 is not None and not relleno1_can:
                registrar_est_error("relleno1", raw_relleno1, "ALERTA", f"Código de Tipo de Relleno no válido. Permitidos: {', '.join(VALID_RELLENO)}")
                relleno1 = None
            else:
                relleno1 = relleno1_can

            if espesor > 0 and (not relleno1 or relleno1 in ["-1", "cwf"]):
                registrar_est_error(
                    "relleno1", 
                    relleno1, 
                    "ADVERTENCIA", 
                    f"Se declaró espesor de relleno pero el tipo de relleno está sin definir o es CWF. Datos evaluados -> Espesor: {espesor}mm, Tipo Relleno: '{relleno1}'."
                )

        if matching_run:
            raw_dureza = sanitize_val(row_dict.get("dureza_pared"), str)
            dureza_pared = get_canonical_value(raw_dureza, VALID_STRENGTHS) if raw_dureza is not None else None
            
            # Solo comparar si ambas durezas son válidas, no nulas y distintas de "-1"
            if dureza_pared and dureza_pared != "-1":
                res_matriz = matching_run["resistencia"]
                r_levels = {"R0": 0, "R1": 1, "R2": 2, "R3": 3, "R4": 4, "R5": 5, "R6": 6}
                if res_matriz and res_matriz != "-1" and dureza_pared in r_levels and res_matriz in r_levels:
                    if r_levels[dureza_pared] > r_levels[res_matriz]:
                        registrar_est_error(
                            "dureza_pared", 
                            raw_dureza, 
                            "ADVERTENCIA", 
                            f"Incompatibilidad geológica (Dureza de pared de junta supera la resistencia maxima estimada de la corrida). Datos evaluados -> Dureza de Pared de Junta en Estructural: {dureza_pared}, Resistencia Maxima Estimada en LGG: {res_matriz}."
                        )

        if not row_has_errors:
            total_ok += 1

    # --- 3. PROCESAR HOJA RMR (SI EXISTE) ---
    rmr_sheet_name = find_rmr_sheet_name(wb.sheetnames)
    total_rmr_filas = 0
    total_sin_informacion = 0
    if rmr_sheet_name:
        ws_rmr = wb[rmr_sheet_name]
        r_counters = {"total_ok": 0, "total_vacios": 0, "total_sin_informacion": 0, "total_advertencias": 0, "total_alertas": 0}
        total_rmr_filas = validate_rmr_sheet_data(ws_rmr, lgg_runs, resumen_celdas, incidencias, r_counters, collector=collector)
        total_vacios += r_counters["total_vacios"]
        total_sin_informacion += r_counters["total_sin_informacion"]
        total_advertencias += r_counters["total_advertencias"]
        total_alertas += r_counters["total_alertas"]
        total_ok += r_counters["total_ok"]

    total_celdas_ok = 0
    for celda, data in resumen_celdas.items():
        if data["alertas"] > 0:
            data["estado_celda"] = "ALERTA"
        elif data["vacios"] > 0 or data.get("sin_informacion", 0) > 0 or data["advertencias"] > 0:
            data["estado_celda"] = "ADVERTENCIA"
        else:
            data["estado_celda"] = "OK"
            total_celdas_ok += 1

    total_filas = total_lgg_filas + total_est_filas + total_rmr_filas
    total_campos = total_filas * 20

    output_json = {
        "total_filas_procesadas": total_filas,
        "total_celdas_evaluadas": total_campos,
        "metricas_globales": {
            "total_celdas_padre": len(resumen_celdas),
            "total_celdas_hija_procesadas": total_filas,
            "total_ok": total_ok,
            "total_vacios": total_vacios,
            "total_sin_informacion": total_sin_informacion,
            "total_advertencias": total_advertencias,
            "total_alertas": total_alertas,
            "total_celdas_ok": total_celdas_ok
        },
        "distribucion_filas_campana": filas_por_campana,
        "distribucion_geotecnico": filas_por_geotecnico,
        "incidencias": incidencias,
        "resumen_por_celda_padre": resumen_celdas,
        "faltantes_no_obligatorios": collector.dump()
    }

    tmp_path = output_json_path + ".tmp"
    with open(tmp_path, 'w', encoding='utf-8') as f:
        json.dump(output_json, f, ensure_ascii=False)
    
    os.replace(tmp_path, output_json_path)


def validate_revision_bulk_v2(file_paths: dict, config: dict, output_json_path: str):
    """
    Versión 2.0 que confía estrictamente en el JSON Configuration del Frontend,
    soportando 3 archivos (LGG/EST, Collar y Survey) y verificando el cruce cuádruple EOH,
    con todas las validaciones geomecánicas completas de V1.
    """
    import time
    
    wb_main = openpyxl.load_workbook(file_paths["lgg_est"], data_only=True)
    wb_col = openpyxl.load_workbook(file_paths["collar"], data_only=True) if file_paths.get("collar") else None
    wb_sur = openpyxl.load_workbook(file_paths["survey"], data_only=True) if file_paths.get("survey") else None

    incidencias = []
    lgg_runs = []
    total_lgg_filas = 0
    total_est_filas = 0
    
    total_vacios = 0
    total_sin_informacion = 0
    total_advertencias = 0
    total_alertas = 0
    total_ok = 0

    resumen_celdas = {} 
    filas_por_campana = {}
    filas_por_geotecnico = {}
    last_a_by_taladro = {}

    collector = FaltantesCollector()

    max_lgg = {}
    max_est = {}
    eoh_collar = {}
    max_survey = {}

    # --- 1. PROCESAR LGG ---
    conf_lgg = config.get("lgg")
    if conf_lgg:
        ws_lgg = wb_main[conf_lgg["sheet"]]
        h_idx, l_map = find_header_row_and_mapping(ws_lgg, LGG_PATTERNS)
        lgg_claves_detectadas = set(l_map.keys())
        for k, v in FALLBACK_LGG_MAP.items():
            if k not in l_map: l_map[k] = v

        lgg_extra = [k for k in CAMPOS_EXTRA_A_CAPTURAR["LGG"] if k in lgg_claves_detectadas]
            
        print(f"[*] Iniciando escaneo de LGG V2. Fila inicial: {h_idx + 1}, Fila máxima: {ws_lgg.max_row}", flush=True)
        
        empty_streak = 0
        current_taladro = None
        
        for r in range(h_idx + 1, ws_lgg.max_row + 1):
            if r % 500 == 0: print(f"  ... [LGG V2] Fila {r} / {ws_lgg.max_row}", flush=True)
            row_dict, is_empty = get_row_dict(ws_lgg, r, l_map)
            
            t_val = row_dict.get("taladro")
            if t_val is not None and str(t_val).strip() != "":
                current_taladro = safe_str(t_val)
            else:
                row_dict["taladro"] = current_taladro
                
            if is_empty:
                empty_streak += 1
                if empty_streak >= 20:
                    print(f"[*] [LGG V2] Freno de emergencia en fila {r}.", flush=True)
                    break
                continue
            
            if not current_taladro: continue
            
            empty_streak = 0
            total_lgg_filas += 1
            taladro = current_taladro
            
            corrida_num = safe_int(row_dict.get("corrida", 0))
            camp = sanitize_val(row_dict.get("campana"), int)
            geo = sanitize_val(row_dict.get("geologo"), str)
            celda_padre = taladro
            celda_hija = f"{taladro}-C{corrida_num}"

            if camp: filas_por_campana[str(camp)] = filas_por_campana.get(str(camp), 0) + 1
            if geo: filas_por_geotecnico[geo] = filas_por_geotecnico.get(geo, 0) + 1

            if celda_padre not in resumen_celdas:
                resumen_celdas[celda_padre] = {"total_hijas": 0, "vacios": 0, "advertencias": 0, "alertas": 0, "estado_celda": "OK", "dist_celda": 0.0, "campania": str(camp) if camp else "N/A"}

            resumen_celdas[celda_padre]["total_hijas"] += 1
            row_has_errors = False

            def reg_err(col, val, tipo, msg, mod="LGG", hija=celda_hija):
                nonlocal total_vacios, total_sin_informacion, total_advertencias, total_alertas, row_has_errors
                incidencias.append({"fila_excel": r, "celda_padre": celda_padre, "celda_hija": hija, "columna": col, "valor_actual": val, "tipo_incidencia": tipo, "mensaje": msg, "campania": str(camp) if camp else "N/A", "geotecnico": geo if geo else "N/A", "sector_geotecnico": "N/A", "modulo": mod})
                if tipo == "VACIO":
                    total_vacios += 1
                    resumen_celdas[celda_padre]["vacios"] += 1
                elif tipo == "SIN_INFORMACION":
                    total_sin_informacion += 1
                    if "sin_informacion" not in resumen_celdas[celda_padre]:
                        resumen_celdas[celda_padre]["sin_informacion"] = 0
                    resumen_celdas[celda_padre]["sin_informacion"] += 1
                elif tipo == "ADVERTENCIA":
                    total_advertencias += 1
                    resumen_celdas[celda_padre]["advertencias"] += 1
                elif tipo == "ALERTA":
                    total_alertas += 1
                    resumen_celdas[celda_padre]["alertas"] += 1
                    row_has_errors = True

            # --- 1. PARSEO Y SANITIZACIÓN INMEDIATA ---
            de = sanitize_val(row_dict.get("de"), float)
            a = sanitize_val(row_dict.get("a"), float)

            rec_m = sanitize_val(row_dict.get("rec_m"), float)
            rqd_m = sanitize_val(row_dict.get("rqd_m"), float)
            lrf_m = sanitize_val(row_dict.get("lrf_m"), float)
            small_frag_m = sanitize_val(row_dict.get("small_frag_m"), float)

            frac_nat = sanitize_val(row_dict.get("frac_nat"), int)
            b30 = sanitize_val(row_dict.get("frac_buz30"), int)
            b60 = sanitize_val(row_dict.get("frac_buz60"), int)
            b90 = sanitize_val(row_dict.get("frac_buz90"), int)

            abertura = sanitize_val(row_dict.get("abertura"), float)
            espesor = sanitize_val(row_dict.get("espesor"), float)

            raw_resistencia = row_dict.get("resistencia")
            resistencia_can = get_canonical_value(raw_resistencia, VALID_STRENGTHS)
            resistencia = resistencia_can or "R4"

            raw_intemperismo = row_dict.get("intemperismo")
            intemperismo_can = get_canonical_value(raw_intemperismo, VALID_WEATHERING)
            weathering = intemperismo_can or "UWF"

            raw_relleno1 = row_dict.get("relleno1")
            relleno1_can = get_canonical_value(raw_relleno1, VALID_RELLENO)
            relleno1 = relleno1_can

            raw_agua_obs = row_dict.get("agua_obs")
            agua_obs_can = get_canonical_value(raw_agua_obs, VALID_AGUA)

            jrc10 = safe_int(sanitize_val(row_dict.get("jrc10"), int))

            raw_rugosidad = row_dict.get("rugosidad")
            rugosidad_can = get_canonical_value(raw_rugosidad, VALID_RUGOSITY)

            tipo_est1_raw = safe_str(sanitize_val(row_dict.get("tipo_est1"), str))
            tipo_est2_raw = safe_str(sanitize_val(row_dict.get("tipo_est2"), str))

            # --- 2. VALIDAR CAMPOS OBLIGATORIOS VACÍOS ---
            mandatory_lgg = ["corrida", "de", "a", "rec_m", "rqd_m", "lrf_m", "small_frag_m", "frac_nat", "lito1", "resistencia", "tipo_est1", "frac_buz30", "frac_buz60", "frac_buz90", "abertura", "rugosidad", "jrc10", "intemperismo", "relleno1", "espesor", "agua_obs", "campana", "geologo"]
            for key in mandatory_lgg:
                if key not in l_map: continue
                val_raw = row_dict.get(key)
                if val_raw is None or str(val_raw).strip() == "":
                    reg_err(key, None, "VACIO", f"El campo obligatorio '{key}' se encuentra vacío.")
                elif str(val_raw).strip() in ["-1", "-1.0", "-1,0"]:
                    reg_err(key, val_raw, "SIN_INFORMACION", f"El campo obligatorio '{key}' no contiene información (-1).")

            capturar_faltantes_extra(collector, "LGG", row_dict, lgg_extra, r, celda_padre, celda_hija, camp, geo)

            # --- 3. VALIDACIÓN DE CATÁLOGOS ---
            tipo_est1_can = get_canonical_value(tipo_est1_raw, VALID_STRUCTURES)
            if tipo_est1_raw and not tipo_est1_can:
                if tipo_est1_raw.upper() == "J":
                    tipo_est1 = "JN"
                    reg_err("tipo_est1", "J", "ADVERTENCIA", "Código de estructura 'J' reconocido como 'JN' (Junta).")
                else:
                    reg_err("tipo_est1", tipo_est1_raw, "ALERTA", f"Código de estructura 1 no válido. Permitidos: {', '.join(VALID_STRUCTURES)}")
                    tipo_est1 = "JN"
            else:
                tipo_est1 = tipo_est1_can or "JN"

            tipo_est2_can = get_canonical_value(tipo_est2_raw, VALID_STRUCTURES)
            if tipo_est2_raw and not tipo_est2_can:
                reg_err("tipo_est2", tipo_est2_raw, "ALERTA", f"Código de estructura 2 no válido. Permitidos: {', '.join(VALID_STRUCTURES)}")
                tipo_est2 = ""
            else:
                tipo_est2 = tipo_est2_can or ""

            # --- 4. REGLAS DE CONSISTENCIA GEOMECÁNICA ---
            if a is not None:
                max_lgg[taladro] = max(max_lgg.get(taladro, 0.0), a)
                resumen_celdas[celda_padre]["dist_celda"] = max(resumen_celdas[celda_padre]["dist_celda"], a)

            raw_de = row_dict.get("de")
            if de is not None and de < -0.0001:
                reg_err("de", raw_de, "ALERTA", f"El valor de 'de:' ({de}m) no puede ser negativo.")
            raw_a = row_dict.get("a")
            if a is not None and a < -0.0001:
                reg_err("a", raw_a, "ALERTA", f"El valor de 'a:' ({a}m) no puede ser negativo.")
            raw_rec = row_dict.get("rec_m")
            if rec_m is not None and rec_m < -0.0001:
                reg_err("rec_m", raw_rec, "ALERTA", f"La longitud recuperada ({rec_m}m) no puede ser negativa.")
            raw_rqd = row_dict.get("rqd_m")
            if rqd_m is not None and rqd_m < -0.0001:
                reg_err("rqd_m", raw_rqd, "ALERTA", f"El metraje RQD ({rqd_m}m) no puede ser negativo.")
            raw_lrf = row_dict.get("lrf_m")
            if lrf_m is not None and lrf_m < -0.0001:
                reg_err("lrf_m", raw_lrf, "ALERTA", f"La longitud de roca fracturada LRF ({lrf_m}m) no puede ser negativa.")
            raw_small = row_dict.get("small_frag_m")
            if small_frag_m is not None and small_frag_m < -0.0001:
                reg_err("small_frag_m", raw_small, "ALERTA", f"El metraje de fragmentos <10cm ({small_frag_m}m) no puede ser negativo.")

            raw_frac_nat = row_dict.get("frac_nat")
            if frac_nat is not None and frac_nat < -0.0001:
                reg_err("frac_nat", raw_frac_nat, "ALERTA", f"El número de fracturas naturales ({frac_nat}) no puede ser negativo.")
            raw_b30 = row_dict.get("frac_buz30")
            if b30 is not None and b30 < -0.0001:
                reg_err("frac_buz30", raw_b30, "ALERTA", f"El número de fracturas en Buz<30° ({b30}) no puede ser negativo.")
            raw_b60 = row_dict.get("frac_buz60")
            if b60 is not None and b60 < -0.0001:
                reg_err("frac_buz60", raw_b60, "ALERTA", f"El número de fracturas en 30°-60° ({b60}) no puede ser negativo.")
            raw_b90 = row_dict.get("frac_buz90")
            if b90 is not None and b90 < -0.0001:
                reg_err("frac_buz90", raw_b90, "ALERTA", f"El número de fracturas en Buz>60° ({b90}) no puede ser negativo.")

            raw_abertura = row_dict.get("abertura")
            if abertura is not None and abertura < -0.0001:
                reg_err("abertura", raw_abertura, "ALERTA", f"La abertura de junta ({abertura}mm) no puede ser negativa.")
            raw_espesor = row_dict.get("espesor")
            if espesor is not None and espesor < -0.0001:
                reg_err("espesor", raw_espesor, "ALERTA", f"El espesor de relleno ({espesor}mm) no puede ser negativo.")
            raw_camp = row_dict.get("campana")
            if camp is not None and camp < -0.0001:
                reg_err("campana", raw_camp, "ALERTA", f"El año de campaña ({camp}) no puede ser negativo.")

            for key, val_raw in [("frac_nat", raw_frac_nat), ("frac_buz30", raw_b30), ("frac_buz60", raw_b60), ("frac_buz90", raw_b90)]:
                if val_raw is not None and val_raw != -1:
                    try:
                        f_val = float(val_raw)
                        if not f_val.is_integer():
                            reg_err(key, val_raw, "ALERTA", f"El campo '{key}' ({val_raw}) debe ser un número entero.")
                    except ValueError:
                        pass

            if "frf" in l_map:
                frf_raw = row_dict.get("frf")
                frf_val = sanitize_val(frf_raw, int)
                if frf_val is not None and frf_val != -1:
                    if frf_val < 0:
                        reg_err("frf", frf_raw, "ALERTA", f"El valor de FRF no puede ser negativo. Datos evaluados -> FRF: {frf_val}.")
                    try:
                        f_frf = float(frf_raw)
                        if not f_frf.is_integer():
                            reg_err("frf", frf_raw, "ALERTA", f"El valor de FRF debe ser un número entero. Datos evaluados -> FRF: {frf_raw}.")
                    except ValueError:
                        pass
                    if lrf_m is not None:
                        calc_frf = math.floor(round(lrf_m * 100) / 5) + 1 if lrf_m > 0 else 0
                        if frf_val != calc_frf:
                            reg_err("frf", frf_raw, "ALERTA", f"El valor de FRF ({frf_val}) no coincide con el calculado por la fórmula: FRF = PISO( REDOND(LRF * 100) / 5 ) + 1 (si LRF > 0, sino 0). Calculado: {calc_frf} basado en LRF ({lrf_m}m).")

            if de is not None and celda_padre in last_a_by_taladro:
                prev_a = last_a_by_taladro[celda_padre]
                if abs(de - prev_a) > 0.001:
                    gap = round(abs(de - prev_a), 4)
                    reg_err("de", de, "ALERTA", f"Ruptura de continuidad espacial detectada. Datos evaluados -> Profundidad de inicio 'De': {de}m, Profundidad final anterior 'A': {prev_a}m (Brecha calculada: {gap}m).")
            if a is not None:
                last_a_by_taladro[celda_padre] = a

            if de is not None and a is not None:
                perf = round(a - de, 2)
                if perf <= 0: reg_err("a", a, "ALERTA", f"Longitud de corrida perforada debe ser positiva. Datos evaluados -> De: {de}m, A: {a}m, Avance calculado: {perf}m.")
                
                if rec_m is not None and round(rec_m, 4) > round(perf, 4): reg_err("rec_m", rec_m, "ALERTA", f"La longitud recuperada es mayor que el avance perforado. Datos evaluados -> Recuperada: {rec_m}m, Avance de corrida: {perf}m (De: {de}m, A: {a}m).")
                if rqd_m is not None and rec_m is not None and round(rqd_m, 4) > round(rec_m, 4): reg_err("rqd_m", rqd_m, "ALERTA", f"Metraje RQD es mayor que la longitud recuperada. Datos evaluados -> RQD: {rqd_m}m, Recuperada: {rec_m}m, Avance de corrida: {perf}m (De: {de}m, A: {a}m).")
                if lrf_m is not None and rec_m is not None and round(lrf_m, 4) > round(rec_m, 4): reg_err("lrf_m", lrf_m, "ALERTA", f"La longitud de roca fracturada LRF es mayor que la longitud recuperada. Datos evaluados -> LRF: {lrf_m}m, Recuperada: {rec_m}m, Avance de corrida: {perf}m (De: {de}m, A: {a}m).")

                if rqd_m is not None and lrf_m is not None and small_frag_m is not None:
                    sum_frags = round(rqd_m + lrf_m + small_frag_m, 2)
                    if round(sum_frags, 4) > round(perf, 4): reg_err("rqd_m", rqd_m, "ALERTA", f"La suma de fragmentos físicos supera el avance perforado. Datos evaluados -> Suma de fragmentos: {sum_frags}m (RQD: {rqd_m}m + LRF: {lrf_m}m + <10cm: {small_frag_m}m), Avance de corrida: {perf}m (De: {de}m, A: {a}m), Longitud Recuperada: {rec_m}m.")

            if b30 is not None and b60 is not None and b90 is not None and frac_nat is not None:
                sum_bins = b30 + b60 + b90
                if sum_bins != frac_nat:
                    reg_err("frac_nat", frac_nat, "ADVERTENCIA", f"La sumatoria de fracturas por buzamiento no coincide con el conteo general. Datos evaluados -> Conteo General (Frac_Nat): {frac_nat}, Suma por buzamiento: {sum_bins} (Buz <30°: {b30} + 30°-60°: {b60} + >60°: {b90}).")

            exceptions = {"F", "RF", "VN", "SZ", "F+10", "BED"}
            if espesor is not None and abertura is not None:
                if espesor > abertura and (tipo_est1 not in exceptions and tipo_est2 not in exceptions):
                    reg_err("espesor", espesor, "ALERTA", f"El espesor de relleno no puede ser mayor que la abertura de junta. Datos evaluados -> Espesor: {espesor}mm (Tipo Relleno: '{relleno1}'), Abertura de Junta: {abertura}mm, Estructuras: '{tipo_est1}' / '{tipo_est2}'.")

                if espesor > 0 and abertura <= 0:
                    reg_err("abertura", abertura, "ADVERTENCIA", f"Se declaró espesor de relleno de junta pero la abertura es 0mm. Datos evaluados -> Espesor: {espesor}mm (Tipo Relleno: '{relleno1}'), Abertura de Junta: {abertura}mm.")
                elif espesor == 0 and abertura > 0 and relleno1 not in [None, "-1"]: 
                    reg_err("espesor", espesor, "ADVERTENCIA", f"La abertura de junta es mayor a 0mm pero no se ha registrado espesor de relleno. Datos evaluados -> Abertura de Junta: {abertura}mm, Espesor: {espesor}mm (Tipo Relleno: '{relleno1}').")

            if raw_resistencia is not None and not resistencia_can:
                pass
            if raw_intemperismo is not None and not intemperismo_can:
                pass
            if raw_relleno1 is not None and not relleno1_can:
                pass
            if raw_agua_obs is not None and not agua_obs_can:
                pass
            if raw_rugosidad is not None and not rugosidad_can:
                reg_err("rugosidad", raw_rugosidad, "ALERTA", f"Código de Rugosidad no válido. Permitidos: {', '.join(VALID_RUGOSITY)}")

            frf_val = sanitize_val(row_dict.get("frf"), int) if "frf" in l_map else None

            lgg_runs.append({
                "taladro": taladro, "de": de, "a": a, "corrida": corrida_num,
                "resistencia": resistencia,
                "lito1": safe_str(row_dict.get("lito1")),
                "lito2": safe_str(row_dict.get("lito2")),
                "lito3": safe_str(row_dict.get("lito3")),
                "rec_m": rec_m,
                "rqd_m": rqd_m,
                "lrf_m": lrf_m,
                "small_frag_m": small_frag_m,
                "frac_nat": frac_nat,
                "frf": frf_val,
                "abertura": abertura,
                "rugosidad": raw_rugosidad if raw_rugosidad is not None else row_dict.get("rugosidad"),
                "jrc10": jrc10,
                "intemperismo": weathering if 'weathering' in locals() else row_dict.get("intemperismo"),
                "relleno1": relleno1 if 'relleno1' in locals() else row_dict.get("relleno1"),
                "espesor": espesor,
                "tipo_est1": tipo_est1 if 'tipo_est1' in locals() else row_dict.get("tipo_est1")
            })
            if not row_has_errors: total_ok += 1

    # --- 2. PROCESAR ESTRUCTURAL ---
    conf_est = config.get("est")
    if conf_est:
        ws_est = wb_main[conf_est["sheet"]]
        h_idx, e_map = find_header_row_and_mapping(ws_est, EST_PATTERNS)
        est_claves_detectadas = set(e_map.keys())
        for k, v in FALLBACK_EST_MAP.items():
            if k not in e_map: e_map[k] = v

        est_extra = [k for k in CAMPOS_EXTRA_A_CAPTURAR["Estructural"] if k in est_claves_detectadas]
            
        print(f"[*] Iniciando escaneo de EST V2. Fila inicial: {h_idx + 1}, Fila máxima: {ws_est.max_row}", flush=True)
        
        empty_streak = 0
        current_taladro_est = None
        
        for r in range(h_idx + 1, ws_est.max_row + 1):
            if r % 500 == 0: print(f"  ... [EST V2] Fila {r} / {ws_est.max_row}", flush=True)
            row_dict, is_empty = get_row_dict(ws_est, r, e_map)
            
            t_val = row_dict.get("taladro")
            if t_val is not None and str(t_val).strip() != "":
                current_taladro_est = safe_str(t_val)
            else:
                row_dict["taladro"] = current_taladro_est
                
            if is_empty:
                empty_streak += 1
                if empty_streak >= 20:
                    print(f"[*] [EST V2] Freno de emergencia en fila {r}.", flush=True)
                    break
                continue
            
            if not current_taladro_est: continue
            
            empty_streak = 0
            total_est_filas += 1
            taladro = current_taladro_est
            celda_padre = taladro
            celda_hija = f"{taladro}-E{r}"

            if celda_padre not in resumen_celdas:
                resumen_celdas[celda_padre] = {"total_hijas": 0, "vacios": 0, "advertencias": 0, "alertas": 0, "estado_celda": "OK", "dist_celda": 0.0, "campania": "N/A"}
            
            resumen_celdas[celda_padre]["total_hijas"] += 1
            row_has_errors = False

            def reg_err_est(col, val, tipo, msg):
                nonlocal total_vacios, total_sin_informacion, total_advertencias, total_alertas, row_has_errors
                incidencias.append({"fila_excel": r, "celda_padre": celda_padre, "celda_hija": celda_hija, "columna": col, "valor_actual": val, "tipo_incidencia": tipo, "mensaje": msg, "campania": str(camp) if camp else "N/A", "geotecnico": geo if geo else "N/A", "sector_geotecnico": "N/A", "modulo": "Estructural"})
                if tipo == "VACIO":
                    total_vacios += 1
                    resumen_celdas[celda_padre]["vacios"] += 1
                elif tipo == "SIN_INFORMACION":
                    total_sin_informacion += 1
                    if "sin_informacion" not in resumen_celdas[celda_padre]:
                        resumen_celdas[celda_padre]["sin_informacion"] = 0
                    resumen_celdas[celda_padre]["sin_informacion"] += 1
                elif tipo == "ADVERTENCIA":
                    total_advertencias += 1
                    resumen_celdas[celda_padre]["advertencias"] += 1
                elif tipo == "ALERTA":
                    total_alertas += 1
                    resumen_celdas[celda_padre]["alertas"] += 1
                    row_has_errors = True

            # --- 1. PARSEO Y SANITIZACIÓN INMEDIATA (ESTRUCTURAL) ---
            depth = sanitize_val(row_dict.get("profundidad"), float)
            camp = sanitize_val(row_dict.get("campana"), int)
            geo = sanitize_val(row_dict.get("geotecnico"), str)

            abertura = sanitize_val(row_dict.get("abertura"), float)
            espesor = sanitize_val(row_dict.get("espesor"), float)

            dip = sanitize_val(row_dict.get("dip"), float)
            azimuth = sanitize_val(row_dict.get("azimuth"), float)

            est_de = sanitize_val(row_dict.get("de"), float)
            est_a = sanitize_val(row_dict.get("a"), float)

            alfa = sanitize_val(row_dict.get("alfa"), float)
            beta = sanitize_val(row_dict.get("beta"), float)
            jrc10 = sanitize_val(row_dict.get("jrc10"), int)
            
            raw_forma = sanitize_val(row_dict.get("forma"), str)
            forma_can = get_canonical_value(raw_forma, VALID_FORMA) if raw_forma is not None else None
            
            tipo_est = safe_str(sanitize_val(row_dict.get("tipo_estructura"), str))
            
            raw_relleno1 = row_dict.get("relleno1")
            relleno1_can = get_canonical_value(raw_relleno1, VALID_RELLENO)
            relleno1 = relleno1_can

            raw_dureza = sanitize_val(row_dict.get("dureza_pared"), str)
            dureza_can = get_canonical_value(raw_dureza, VALID_STRENGTHS) if raw_dureza is not None else None
            dureza_pared = dureza_can

            # --- 2. VALIDACIONES MAESTRAS (EST) ---
            if depth is not None:
                max_est[taladro] = max(max_est.get(taladro, 0.0), depth)
                
                lgg_max_for_t = max_lgg.get(taladro, 0.0)
                if lgg_max_for_t > 0 and depth > lgg_max_for_t:
                    reg_err_est("profundidad", depth, "ALERTA", f"La profundidad en logueo estructural excede el límite final registrado en LGG. Datos evaluados -> Profundidad Estructural: {depth}m, Profundidad Máxima LGG: {lgg_max_for_t}m.")

            if camp:
                resumen_celdas[celda_padre]["campania"] = str(camp)

            mandatory_est = ["profundidad", "alfa", "beta", "forma", "rugosidad", "jrc10", "abertura", "weathering", "espesor", "relleno1", "dureza_pared", "agua", "geotecnico", "campana", "dip", "azimuth"]
            for key in mandatory_est:
                if key not in e_map: continue
                val_raw = row_dict.get(key)
                if val_raw is None or str(val_raw).strip() == "":
                    reg_err_est(key, None, "VACIO", f"El campo obligatorio '{key}' se encuentra vacío.")
                elif str(val_raw).strip() in ["-1", "-1.0", "-1,0"]:
                    reg_err_est(key, val_raw, "SIN_INFORMACION", f"El campo obligatorio '{key}' no contiene información (-1).")

            capturar_faltantes_extra(collector, "Estructural", row_dict, est_extra, r, celda_padre, celda_hija, camp, geo)

            raw_depth = row_dict.get("profundidad")
            if depth is not None and depth < -0.0001:
                reg_err_est("profundidad", raw_depth, "ALERTA", f"Profundidad ({depth}m) no puede ser negativa.")
            raw_abertura = row_dict.get("abertura")
            if abertura is not None and abertura < -0.0001:
                reg_err_est("abertura", raw_abertura, "ALERTA", f"La abertura ({abertura}mm) no puede ser negativa.")
            raw_espesor = row_dict.get("espesor")
            if espesor is not None and espesor < -0.0001:
                reg_err_est("espesor", raw_espesor, "ALERTA", f"El espesor de relleno ({espesor}mm) no puede ser negativo.")
            raw_camp = row_dict.get("campana")
            if camp is not None and camp < -0.0001:
                reg_err_est("campana", raw_camp, "ALERTA", f"El año de campaña ({camp}) no puede ser negativo.")

            raw_dip = row_dict.get("dip")
            if dip is not None:
                if dip < 0.0 or dip > 90.0:
                    reg_err_est("dip", raw_dip, "ALERTA", f"El ángulo Dip es inválido. Datos evaluados -> Dip: {dip}°. Debe estar entre 0° y 90°.")

            raw_azimuth = row_dict.get("azimuth")
            if azimuth is not None:
                if azimuth < 0.0 or azimuth > 360.0:
                    reg_err_est("azimuth", raw_azimuth, "ALERTA", f"El ángulo Azimut es inválido. Datos evaluados -> Azimut: {azimuth}°. Debe estar entre 0° y 360°.")

            matching_run = None
            if depth is not None:
                for run in lgg_runs:
                    if run["taladro"] == taladro and run["de"] <= depth <= run["a"]:
                        matching_run = run
                        break

                if matching_run is None:
                    reg_err_est("profundidad", depth, "ALERTA", f"Profundidad huérfana de junta no corresponde a ningún tramo de corrida en LGG. Datos evaluados -> Profundidad de Junta: {depth}m, Taladro: '{taladro}'.")

            raw_de = row_dict.get("de")
            raw_a = row_dict.get("a")
            if est_de is not None and est_a is not None:
                has_exact_match = False
                for run in lgg_runs:
                    if run["taladro"] == taladro and abs(run["de"] - est_de) < 0.001 and abs(run["a"] - est_a) < 0.001:
                        has_exact_match = True
                        break
                if not has_exact_match:
                    reg_err_est("de", raw_de, "ALERTA", f"La corrida asociada (de/a) no existe de forma exacta en las corridas de LGG para el taladro. Datos evaluados -> Tramo Estructural de corrida: {est_de}m - {est_a}m, Taladro: '{taladro}'.")
                if depth is not None and (depth < est_de or depth > est_a):
                    reg_err_est("profundidad", depth, "ALERTA", f"La profundidad se encuentra fuera del tramo de corrida especificado. Datos evaluados -> Profundidad de Junta: {depth}m, Tramo especificado: {est_de}m - {est_a}m.")

            if alfa is not None:
                if alfa < 0.0 or alfa > 90.0:
                    reg_err_est("alfa", alfa, "ALERTA", f"El ángulo Alfa es inválido. Datos evaluados -> Alfa: {alfa}°. Debe estar entre 0° y 90° o ser -1.")
                elif not float(alfa).is_integer():
                    reg_err_est("alfa", alfa, "ADVERTENCIA", f"El ángulo Alfa debería ser un número entero. Datos evaluados -> Alfa: {alfa}°.")

            if beta is not None:
                if beta < 0.0 or beta > 360.0:
                    reg_err_est("beta", beta, "ALERTA", f"El ángulo Beta es inválido. Datos evaluados -> Beta: {beta}°. Debe estar entre 0° y 360° o ser -1.")
                elif not float(beta).is_integer():
                    reg_err_est("beta", beta, "ADVERTENCIA", f"El ángulo Beta debería ser un número entero. Datos evaluados -> Beta: {beta}°.")

            if jrc10 is not None:
                if jrc10 > 20:
                    reg_err_est("jrc10", jrc10, "ALERTA", f"El valor de JRC10 es inválido. No se permiten valores mayores a 20. Datos evaluados -> JRC10: {jrc10}.")
                elif jrc10 < 0:
                    reg_err_est("jrc10", jrc10, "ALERTA", f"El valor de JRC10 no puede ser negativo. Datos evaluados -> JRC10: {jrc10}.")

            if raw_forma is not None and not forma_can:
                reg_err_est("forma", raw_forma, "ALERTA", f"Forma de junta no válida. Permitidos: Plano (1) a Irregular (6). Datos evaluados -> Forma: '{raw_forma}'.")

            exceptions = {"F", "RF", "VN", "SZ", "F+10", "BED"}
            if espesor is not None and abertura is not None:
                if espesor > abertura and tipo_est not in exceptions:
                    reg_err_est("espesor", espesor, "ALERTA", f"El espesor de relleno no puede ser mayor que la abertura de junta excepto en estructuras F, RF, VN, SZ, F+10 o BED. Datos evaluados -> Espesor: {espesor}mm (Tipo Relleno: '{relleno1}'), Abertura de Junta: {abertura}mm, Estructura: '{tipo_est}'.")

                if espesor > 0 and (not relleno1 or relleno1 in ["-1", "cwf"]):
                    reg_err_est(
                        "relleno1", 
                        relleno1, 
                        "ADVERTENCIA",  # <-- Parámetro posicional 'tipo' restaurado con éxito
                        f"Se declaró espesor de relleno pero el tipo de relleno está sin definir o es CWF. Datos evaluados -> Espesor: {espesor}mm, Tipo Relleno: '{relleno1}'."
                    )
                elif relleno1 and relleno1 not in ["-1", "cwf"] and abertura <= 0:
                    reg_err_est("relleno1", relleno1, "ADVERTENCIA", f"El tipo de relleno está definido pero la abertura de junta es 0mm. Datos evaluados -> Tipo Relleno: '{relleno1}', Abertura de Junta: {abertura}mm, Espesor: {espesor}mm.")

            if matching_run:
                if raw_dureza is not None and not dureza_can:
                    reg_err_est("dureza_pared", raw_dureza, "ALERTA", f"Código de Resistencia ISRM no válido. Permitidos: {', '.join(VALID_STRENGTHS)}")
                
                # Solo comparar compatibilidad si ambas durezas son válidas y distintas de "-1"
                if dureza_pared and dureza_pared != "-1":
                    res_matriz = matching_run["resistencia"]
                    r_levels = {"R0": 0, "R1": 1, "R2": 2, "R3": 3, "R4": 4, "R5": 5, "R6": 6}
                    if res_matriz and res_matriz != "-1" and dureza_pared in r_levels and res_matriz in r_levels:
                        if r_levels[dureza_pared] > r_levels[res_matriz]:
                            reg_err_est(
                                "dureza_pared", 
                                raw_dureza, 
                                "ADVERTENCIA", 
                                f"Incompatibilidad geológica (Dureza de pared de junta supera la resistencia maxima estimada de la corrida). Datos evaluados -> Dureza de Pared de Junta en Estructural: {dureza_pared}, Resistencia Maxima Estimada en LGG: {res_matriz}."
                            )

            if not row_has_errors: total_ok += 1

    # --- 2.5. PROCESAR HOJA RMR (SI EXISTE) ---
    conf_rmr = config.get("rmr")
    rmr_sheet_name = conf_rmr["sheet"] if conf_rmr else find_rmr_sheet_name(wb_main.sheetnames)
    total_rmr_filas = 0
    total_sin_informacion = 0
    if rmr_sheet_name and rmr_sheet_name in wb_main.sheetnames:
        ws_rmr = wb_main[rmr_sheet_name]
        custom_rmr_map = conf_rmr.get("mappings") if conf_rmr else None
        r_counters = {"total_ok": 0, "total_vacios": 0, "total_sin_informacion": 0, "total_advertencias": 0, "total_alertas": 0}
        total_rmr_filas = validate_rmr_sheet_data(ws_rmr, lgg_runs, resumen_celdas, incidencias, r_counters, custom_map=custom_rmr_map, collector=collector)
        total_vacios += r_counters["total_vacios"]
        total_sin_informacion += r_counters["total_sin_informacion"]
        total_advertencias += r_counters["total_advertencias"]
        total_alertas += r_counters["total_alertas"]
        total_ok += r_counters["total_ok"]

    # --- 3. PROCESAR COLLAR PARA EOH ---
    conf_col = config.get("collar")
    if wb_col and conf_col:
        ws_col = wb_col[conf_col["sheet"]]
        h_idx, c_map = find_header_row_and_mapping(ws_col, COLLAR_PATTERNS)
        
        print(f"[*] Escaneando Collar. Fila {h_idx + 1} a {ws_col.max_row}", flush=True)
        for r in range(h_idx + 1, ws_col.max_row + 1):
            row_dict, _ = get_row_dict(ws_col, r, c_map)
            t_val = row_dict.get("taladro")
            e_val = row_dict.get("eoh")
            if not t_val: continue
            
            t_str = safe_str(t_val)
            e_float = safe_float(sanitize_val(e_val, float))
            eoh_collar[t_str] = e_float

    # --- 4. PROCESAR SURVEY PARA DEPTH ---
    conf_sur = config.get("survey")
    if wb_sur and conf_sur:
        ws_sur = wb_sur[conf_sur["sheet"]]
        h_idx, s_map = find_header_row_and_mapping(ws_sur, SURVEY_PATTERNS)
        
        print(f"[*] Escaneando Survey. Fila {h_idx + 1} a {ws_sur.max_row}", flush=True)
        for r in range(h_idx + 1, ws_sur.max_row + 1):
            row_dict, _ = get_row_dict(ws_sur, r, s_map)
            t_val = row_dict.get("taladro")
            d_val = row_dict.get("depth")
            if not t_val: continue
            
            t_str = safe_str(t_val)
            d_float = safe_float(sanitize_val(d_val, float))
            max_survey[t_str] = max(max_survey.get(t_str, 0.0), d_float)

    # --- 5. CRUCE CUÁDRUPLE (REGLA CRÍTICA DE PROFUNDIDAD FINAL) ---
    taladros_procesados = set(list(max_lgg.keys()) + list(max_est.keys()))
    for t in taladros_procesados:
        l_val = max_lgg.get(t, 0.0)
        e_val = max_est.get(t, 0.0)
        
        has_collar = wb_col and t in eoh_collar
        has_survey = wb_sur and t in max_survey
        
        c_val = eoh_collar.get(t, l_val) if has_collar else l_val
        s_val = max_survey.get(t, l_val) if has_survey else l_val
        
        has_conflict = False
        
        if l_val > 0 and e_val > 0 and abs(l_val - e_val) > 0.05:
            has_conflict = True
            
        if has_collar and abs(l_val - c_val) > 0.05:
            has_conflict = True
            
        if has_survey and abs(l_val - s_val) > 0.05:
            has_conflict = True
            
        if has_conflict:
            c_str = f"{c_val}m" if has_collar else "No Cargado"
            s_str = f"{s_val}m" if has_survey else "No Cargado"
            
            msg = f"Las profundidades finales del taladro no coinciden entre módulos (LGG, Estructural, Collar, Survey). Datos evaluados -> LGG Max: {l_val}m, Estructural Max: {e_val}m, Collar EOH: {c_str}, Survey Max: {s_str}."
            
            incidencias.append({
                "fila_excel": 0, "celda_padre": t, "celda_hija": t,
                "columna": "Profundidad Final EOH", "valor_actual": l_val, 
                "tipo_incidencia": "ALERTA", "mensaje": msg,
                "campania": "N/A", "geotecnico": "N/A", "sector_geotecnico": "N/A",
                "modulo": "Cruce General"
            })
            if t in resumen_celdas:
                resumen_celdas[t]["alertas"] += 1
                total_alertas += 1

    total_celdas_ok = 0
    for celda, data in resumen_celdas.items():
        if data["alertas"] > 0: data["estado_celda"] = "ALERTA"
        elif data["vacios"] > 0 or data.get("sin_informacion", 0) > 0 or data["advertencias"] > 0: data["estado_celda"] = "ADVERTENCIA"
        else:
            data["estado_celda"] = "OK"
            total_celdas_ok += 1

    total_filas = total_lgg_filas + total_est_filas + total_rmr_filas
    total_campos = total_filas * 20

    output_json = {
        "total_filas_procesadas": total_filas,
        "total_celdas_evaluadas": total_campos,
        "metricas_globales": {
            "total_celdas_padre": len(resumen_celdas),
            "total_celdas_hija_procesadas": total_filas,
            "total_ok": total_ok,
            "total_vacios": total_vacios,
            "total_sin_informacion": total_sin_informacion,
            "total_advertencias": total_advertencias,
            "total_alertas": total_alertas,
            "total_celdas_ok": total_celdas_ok
        },
        "distribucion_filas_campana": filas_por_campana,
        "distribucion_geotecnico": filas_por_geotecnico,
        "incidencias": incidencias,
        "resumen_por_celda_padre": resumen_celdas,
        "faltantes_no_obligatorios": collector.dump()
    }

    tmp_path = output_json_path + ".tmp"
    with open(tmp_path, 'w', encoding='utf-8') as f:
        json.dump(output_json, f, ensure_ascii=False)
    os.replace(tmp_path, output_json_path)

    if wb_col: wb_col.close()
    if wb_sur: wb_sur.close()
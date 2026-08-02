from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

font_title = Font(name="Segoe UI", size=16, bold=True, color="1B365D")
font_subtitle = Font(name="Segoe UI", size=10, italic=True, color="555555")
font_section = Font(name="Segoe UI", size=11, bold=True, color="1B365D")
font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
font_bold = Font(name="Segoe UI", size=10, bold=True, color="000000")
font_regular = Font(name="Segoe UI", size=10, color="000000")
font_italic = Font(name="Segoe UI", size=10, italic=True, color="7F8C8D")
font_kpi_lbl = Font(name="Segoe UI", size=9, bold=True, color="555555")
font_kpi_val_blue = Font(name="Segoe UI", size=18, bold=True, color="1B365D")
font_kpi_val_green = Font(name="Segoe UI", size=18, bold=True, color="375623")
font_kpi_val_red = Font(name="Segoe UI", size=18, bold=True, color="C00000")
font_kpi_val_orange = Font(name="Segoe UI", size=18, bold=True, color="C65911")
font_link = Font(name="Segoe UI", size=10, bold=True, color="1B365D", underline="single")

fill_primary = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
fill_accent_green = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
fill_accent_yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
fill_accent_orange = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
fill_accent_red = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
fill_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
fill_kpi_gray = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")

border_thin = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0'),
)
border_kpi = Border(
    left=Side(style='thin', color='B0C4DE'),
    right=Side(style='thin', color='B0C4DE'),
    top=Side(style='thin', color='B0C4DE'),
    bottom=Side(style='thin', color='B0C4DE'),
)

alignment_center = Alignment(horizontal="center", vertical="center")
alignment_left = Alignment(horizontal="left", vertical="center")
alignment_right = Alignment(horizontal="right", vertical="center")
alignment_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

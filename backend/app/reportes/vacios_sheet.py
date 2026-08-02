import math

from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.core.report_config import COLORES
from app.reportes.estilos import (
    alignment_center,
    alignment_left,
    alignment_right,
    alignment_wrap,
    border_kpi,
    border_thin,
    fill_accent_orange,
    fill_kpi_gray,
    fill_primary,
    fill_zebra,
    font_bold,
    font_header,
    font_italic,
    font_kpi_lbl,
    font_kpi_val_blue,
    font_kpi_val_orange,
    font_kpi_val_red,
    font_link,
    font_regular,
    font_section,
    font_subtitle,
    font_title,
)

HOJA_TITULO = "📊 Análisis de Datos Faltantes"
SEL_CELDA = "$J$5"
OPCION_MODULO = "Por módulo"
OPCION_GLOBAL = "% del total identificado"

NIVEL_FILLS = {
    "Alta": PatternFill(start_color=COLORES["alta_durazno"], end_color=COLORES["alta_durazno"], fill_type="solid"),
    "Media": PatternFill(start_color=COLORES["media_amarillo"], end_color=COLORES["media_amarillo"], fill_type="solid"),
    "Puntual": PatternFill(start_color=COLORES["puntual_verde"], end_color=COLORES["puntual_verde"], fill_type="solid"),
    "Sin acción": PatternFill(start_color=COLORES["sin_accion_gris"], end_color=COLORES["sin_accion_gris"], fill_type="solid"),
}
ROSA_FILL = PatternFill(start_color=COLORES["rosa_claro"], end_color=COLORES["rosa_claro"], fill_type="solid")
VERDE_FILL = PatternFill(start_color=COLORES["obligatorio_verde_claro"], end_color=COLORES["obligatorio_verde_claro"], fill_type="solid")
AMBAR_FILL = PatternFill(start_color=COLORES["no_obligatorio_ambar_claro"], end_color=COLORES["no_obligatorio_ambar_claro"], fill_type="solid")

ANCHO_COLUMNAS = {
    "A": 3, "B": 14, "C": 30, "D": 10, "E": 10, "F": 10, "G": 13, "H": 13,
    "I": 30, "J": 9, "K": 9, "L": 13, "M": 9, "N": 14, "O": 30, "P": 10,
    "Q": 10, "R": 10, "S": 13, "T": 13,
}

MODULOS_PANEL = ["LGG", "Estructural", "Validación RMR"]
METRICAS_PANEL = [
    ("Taladros afectados", 0),
    ("Faltantes totales", 0),
    ("Celdas vacías (N)", 0),
    ("Sin información -1 (N)", 0),
    ("Campos con faltantes (N)", 0),
    ("Campos en nivel ALTA (N)", 0),
    ("Top 1 campo · faltantes", 0),
    ("Top 1 · cantidad", 0),
    ("Top 2 campo · faltantes", 0),
    ("Top 2 · cantidad", 0),
    ("Top 3 campo · faltantes", 0),
    ("Top 3 · cantidad", 0),
]


def _put(ws, row, col, value=None, font=None, fill=None, align=None, fmt=None, border=True):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align
    if fmt:
        cell.number_format = fmt
    if border:
        cell.border = border_thin
    return cell


def _link(ws, row, col, destino, texto, align=None):
    cell = ws.cell(row=row, column=col,
                   value=f'=HYPERLINK("#\'{HOJA_TITULO}\'!{destino}", "{texto}")')
    cell.font = font_link
    cell.alignment = align or alignment_center
    return cell


def _pct_formula(pct_m, pct_g):
    return f'=IF({SEL_CELDA}="{OPCION_GLOBAL}",{pct_g / 100.0:.6f},{pct_m / 100.0:.6f})'


def _kpi_card(ws, r, c, label, value, fill, val_font):
    c1 = ws.cell(row=r, column=c, value=label)
    c1.font = font_kpi_lbl
    c1.alignment = alignment_center
    c2 = ws.cell(row=r + 1, column=c, value=value)
    c2.font = val_font
    c2.alignment = alignment_center
    for rr in (r, r + 1):
        for cc in (c, c + 1):
            cell = ws.cell(row=rr, column=cc)
            cell.fill = fill
            cell.border = border_kpi
    ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + 1)
    ws.merge_cells(start_row=r + 1, start_column=c, end_row=r + 1, end_column=c + 1)


def _cabecera(ws, r, c0, headers):
    for i, h in enumerate(headers):
        cell = ws.cell(row=r, column=c0 + i, value=h)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = alignment_center
        cell.border = border_thin
    return r + 1


def _titulo_desc(ws, r, c0, span, titulo, desc, link_destino=None):
    _put(ws, r, c0, titulo, font=font_section, align=alignment_left)
    ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c0 + span - 1)
    if link_destino:
        _link(ws, r, c0 + span, f"B{link_destino}", "⬆ Índice", align=alignment_center)
    _put(ws, r + 1, c0, desc, font=font_subtitle, align=alignment_wrap)
    ws.merge_cells(start_row=r + 1, start_column=c0, end_row=r + 1, end_column=c0 + span - 1)
    return r + 2


def _fila_total(ws, r, c0, span, valores, nota=None, fill=fill_kpi_gray):
    for i, v in enumerate(valores):
        cell = _put(ws, r, c0 + i, v, font=font_bold, fill=fill,
                    align=alignment_right if isinstance(v, (int, float)) else alignment_center)
        if isinstance(v, (int, float)):
            cell.number_format = "#,##0"
    if nota:
        ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c0 + span - 1)
        _put(ws, r, c0, nota, font=font_italic, align=alignment_wrap, fill=fill)
    return r + 1


def _nota(ws, r, c0, span, texto):
    cell = _put(ws, r, c0, texto, font=font_italic, align=alignment_wrap)
    ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c0 + span - 1)
    return r + 1


def _cf_contadores(ws, r0, r1, col):
    if r1 > r0:
        ws.conditional_formatting.add(
            f"{get_column_letter(col)}{r0}:{get_column_letter(col)}{r1}",
            DataBarRule(start_type="min", end_type="max", color="638EC6", showValue=True),
        )


def _cf_pct(ws, r0, r1, col):
    if r1 > r0:
        ws.conditional_formatting.add(
            f"{get_column_letter(col)}{r0}:{get_column_letter(col)}{r1}",
            ColorScaleRule(start_type="min", start_color="F8696B",
                           mid_type="percentile", mid_value=50, mid_color="FFEB84",
                           end_type="max", end_color="63BE7B"),
        )


def _celda_nivel(ws, r, c, nivel, zebra=None):
    cell = _put(ws, r, c, nivel, font=font_bold, align=alignment_center, fill=zebra)
    if nivel in NIVEL_FILLS:
        cell.fill = NIVEL_FILLS[nivel]
    return cell


def _h_campos(n):
    return n + 6


def _seccion_tabla_ab(ws, r0, c0, titulo, desc, filas, campo, toggle, base_total, base_nota,
                      col_header):
    r = _titulo_desc(ws, r0, c0, 5, titulo, desc)
    r = _cabecera(ws, r, c0, ["Módulo", "Campo", col_header, "% del Total", "Nivel de Atención"])
    for idx, f in enumerate(filas):
        zebra = fill_zebra if idx % 2 == 1 else None
        _put(ws, r, c0, f["modulo"], font=font_bold, fill=zebra, align=alignment_center)
        _put(ws, r, c0 + 1, f["etiqueta"], fill=zebra, align=alignment_left)
        _put(ws, r, c0 + 2, f[campo], fmt="#,##0", align=alignment_right, fill=zebra)
        if toggle:
            _put(ws, r, c0 + 3, _pct_formula(f["pct_m"], f["pct_g"]), fmt="0.00%",
                 align=alignment_right, fill=zebra)
        else:
            _put(ws, r, c0 + 3, f["pct_g"] / 100.0, fmt="0.00%", align=alignment_right, fill=zebra)
        _celda_nivel(ws, r, c0 + 4, f["nivel"], zebra)
        r += 1
    if not filas:
        cell = _put(ws, r, c0, "Sin datos en este tipo.", font=font_italic, align=alignment_left)
        ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c0 + 4)
        r += 1
    else:
        _cf_contadores(ws, r0 + 3, r - 1, c0 + 2)
        _cf_pct(ws, r0 + 3, r - 1, c0 + 3)
        total = sum(f[campo] for f in filas)
        _put(ws, r, c0, "TOTAL", font=font_bold, align=alignment_center, fill=fill_kpi_gray)
        _put(ws, r, c0 + 1, "—", font=font_regular, align=alignment_center, fill=fill_kpi_gray)
        _put(ws, r, c0 + 2, total, font=font_bold, fmt="#,##0", align=alignment_right, fill=fill_kpi_gray)
        _put(ws, r, c0 + 3, "—", font=font_regular, align=alignment_center, fill=fill_kpi_gray)
        _put(ws, r, c0 + 4, "—", font=font_regular, align=alignment_center, fill=fill_kpi_gray)
        r += 1
        r = _nota(ws, r, c0, 5, f"Base del %: {base_total} {base_nota}. El % de cada fila es "
                               f"sobre el total de su módulo (o global según el selector).")
    return r + 1


def _h_sub(n):
    return n + 6


def _seccion_tabla_c(ws, r0, c0, filas, gran_total):
    r = _titulo_desc(ws, r0, c0, 7, "TABLA C · COMBINADO (VACÍAS + -1)",
                     "Visión global de los datos faltantes por campo. Nivel de Atención "
                     "según el % combinado del módulo; es la tabla de referencia del "
                     "reporte.")
    r = _cabecera(ws, r, c0, ["Módulo", "Campo", "Vacías", "-1", "Total", "% del Total",
                              "Nivel de Atención"])
    for idx, f in enumerate(filas):
        zebra = fill_zebra if idx % 2 == 1 else None
        _put(ws, r, c0, f["modulo"], font=font_bold, fill=zebra, align=alignment_center)
        _put(ws, r, c0 + 1, f["etiqueta"], fill=zebra, align=alignment_left)
        _put(ws, r, c0 + 2, f["v"], fmt="#,##0", align=alignment_right, fill=zebra)
        _put(ws, r, c0 + 3, f["s"], fmt="#,##0", align=alignment_right, fill=zebra)
        _put(ws, r, c0 + 4, f["total"], font=font_bold, fmt="#,##0", align=alignment_right, fill=zebra)
        _put(ws, r, c0 + 5, _pct_formula(f["pct_m"], f["pct_g"]), fmt="0.00%",
             align=alignment_right, fill=zebra)
        _celda_nivel(ws, r, c0 + 6, f["nivel"], zebra)
        r += 1
    if not filas:
        cell = _put(ws, r, c0, "Sin datos en este tipo.", font=font_italic, align=alignment_left)
        ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c0 + 6)
        r += 1
    else:
        _cf_contadores(ws, r0 + 3, r - 1, c0 + 4)
        _cf_pct(ws, r0 + 3, r - 1, c0 + 5)
        _put(ws, r, c0, "TOTAL", font=font_bold, align=alignment_center, fill=fill_kpi_gray)
        _put(ws, r, c0 + 1, "—", font=font_regular, align=alignment_center, fill=fill_kpi_gray)
        _put(ws, r, c0 + 2, sum(f["v"] for f in filas), font=font_bold, fmt="#,##0",
             align=alignment_right, fill=fill_kpi_gray)
        _put(ws, r, c0 + 3, sum(f["s"] for f in filas), font=font_bold, fmt="#,##0",
             align=alignment_right, fill=fill_kpi_gray)
        _put(ws, r, c0 + 4, sum(f["total"] for f in filas), font=font_bold, fmt="#,##0",
             align=alignment_right, fill=fill_kpi_gray)
        _put(ws, r, c0 + 5, "—", font=font_regular, align=alignment_center, fill=fill_kpi_gray)
        _put(ws, r, c0 + 6, "—", font=font_regular, align=alignment_center, fill=fill_kpi_gray)
        r += 1
        r = _nota(ws, r, c0, 7, f"Base del %: {gran_total} faltantes identificados "
                               f"(vacías + -1) = 100%. El % de cada fila es sobre el total "
                               f"de su módulo (o global según el selector).")
    return r + 1


def _seccion_subratings(ws, r0, c0, filas):
    r = _titulo_desc(ws, r0, c0, 7, "SUB-RATINGS RMR'76 / RMR'89 (sección separada)",
                     "Campos derivados del cálculo RMR (resultado, no entrada). Se muestran "
                     "aparte para no mezclarlos con los campos principales.")
    r = _cabecera(ws, r, c0, ["Módulo", "Campo (Sub-ratings)", "Vacías", "-1", "Total",
                              "% del Total", "Nivel de Atención"])
    for idx, f in enumerate(filas):
        zebra = fill_zebra if idx % 2 == 1 else None
        _put(ws, r, c0, f["modulo"], font=font_italic, fill=zebra, align=alignment_center)
        _put(ws, r, c0 + 1, f["etiqueta"], font=font_italic, fill=zebra, align=alignment_left)
        _put(ws, r, c0 + 2, f["v"], fmt="#,##0", align=alignment_right, fill=zebra)
        _put(ws, r, c0 + 3, f["s"], fmt="#,##0", align=alignment_right, fill=zebra)
        _put(ws, r, c0 + 4, f["total"], font=font_bold, fmt="#,##0", align=alignment_right, fill=zebra)
        _put(ws, r, c0 + 5, f["pct_g"] / 100.0, fmt="0.00%", align=alignment_right, fill=zebra)
        _celda_nivel(ws, r, c0 + 6, f["nivel"], zebra)
        r += 1
    if not filas:
        cell = _put(ws, r, c0, "Sin sub-ratings con datos faltantes.", font=font_italic,
                    align=alignment_left)
        ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c0 + 6)
        r += 1
    else:
        _cf_contadores(ws, r0 + 3, r - 1, c0 + 4)
        _cf_pct(ws, r0 + 3, r - 1, c0 + 5)
        total = sum(f["total"] for f in filas)
        _put(ws, r, c0, "TOTAL", font=font_bold, align=alignment_center, fill=fill_kpi_gray)
        _put(ws, r, c0 + 1, "—", font=font_regular, align=alignment_center, fill=fill_kpi_gray)
        _put(ws, r, c0 + 2, sum(f["v"] for f in filas), font=font_bold, fmt="#,##0",
             align=alignment_right, fill=fill_kpi_gray)
        _put(ws, r, c0 + 3, sum(f["s"] for f in filas), font=font_bold, fmt="#,##0",
             align=alignment_right, fill=fill_kpi_gray)
        _put(ws, r, c0 + 4, total, font=font_bold, fmt="#,##0", align=alignment_right, fill=fill_kpi_gray)
        _put(ws, r, c0 + 5, "—", font=font_regular, align=alignment_center, fill=fill_kpi_gray)
        _put(ws, r, c0 + 6, "—", font=font_regular, align=alignment_center, fill=fill_kpi_gray)
        r += 1
        r = _nota(ws, r, c0, 7, f"Base del %: {total} faltantes en sub-ratings (solo "
                               f"campos derivados del cálculo RMR).")
    return r + 1


def _h_alta(n):
    return n + 6


def _seccion_alta(ws, r0, c0, filas, anclas, r_indice, gran_total):
    r = _titulo_desc(ws, r0, c0, 8, "CAMPOS EN NIVEL DE ATENCIÓN ALTA",
                     "Campos obligatorios con más del 10% de faltantes en su módulo. "
                     "Requieren acción prioritaria; usa los enlaces para ver el detalle anual.",
                     link_destino=r_indice)
    r = _cabecera(ws, r, c0, ["Módulo", "Campo", "Vacías", "-1", "Total", "% del Total",
                              "Taladros Afectados", "Navegación"])
    for idx, f in enumerate(filas):
        zebra = fill_zebra if idx % 2 == 1 else None
        _put(ws, r, c0, f["modulo"], font=font_bold, fill=zebra, align=alignment_center)
        _put(ws, r, c0 + 1, f["etiqueta"], fill=zebra, align=alignment_left)
        _put(ws, r, c0 + 2, f["v"], fmt="#,##0", align=alignment_right, fill=zebra)
        _put(ws, r, c0 + 3, f["s"], fmt="#,##0", align=alignment_right, fill=zebra)
        _put(ws, r, c0 + 4, f["total"], font=font_bold, fmt="#,##0", align=alignment_right, fill=zebra)
        _put(ws, r, c0 + 5, _pct_formula(f["pct_m"], f["pct_g"]), fmt="0.00%",
             align=alignment_right, fill=zebra)
        _put(ws, r, c0 + 6, f["celdas"], fmt="#,##0", align=alignment_right, fill=zebra)
        _link(ws, r, c0 + 7, f"B{anclas[idx]}", "🔍 Ver detalle", align=alignment_center)
        r += 1
    if not filas:
        cell = _put(ws, r, c0, "Sin campos en nivel ALTA.", font=font_italic, align=alignment_left)
        ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c0 + 7)
        r += 1
    else:
        _cf_contadores(ws, r0 + 3, r - 1, c0 + 4)
        _cf_pct(ws, r0 + 3, r - 1, c0 + 5)
        _put(ws, r, c0, "TOTAL", font=font_bold, align=alignment_center, fill=fill_kpi_gray)
        _put(ws, r, c0 + 1, "—", font=font_regular, align=alignment_center, fill=fill_kpi_gray)
        _put(ws, r, c0 + 2, sum(f["v"] for f in filas), font=font_bold, fmt="#,##0",
             align=alignment_right, fill=fill_kpi_gray)
        _put(ws, r, c0 + 3, sum(f["s"] for f in filas), font=font_bold, fmt="#,##0",
             align=alignment_right, fill=fill_kpi_gray)
        _put(ws, r, c0 + 4, sum(f["total"] for f in filas), font=font_bold, fmt="#,##0",
             align=alignment_right, fill=fill_kpi_gray)
        _put(ws, r, c0 + 5, "—", font=font_regular, align=alignment_center, fill=fill_kpi_gray)
        _put(ws, r, c0 + 6, sum(f["celdas"] for f in filas), font=font_bold, fmt="#,##0",
             align=alignment_right, fill=fill_kpi_gray)
        _put(ws, r, c0 + 7, "—", font=font_regular, align=alignment_center, fill=fill_kpi_gray)
        r += 1
        r = _nota(ws, r, c0, 8, f"TOTAL: {sum(f['total'] for f in filas)} faltantes en "
                               f"{len(filas)} campos ALTA · sobre {gran_total} identificados. "
                               f"Taladros afectados: suma por campo (un taladro puede "
                               f"repetirse en varios campos).")
    return r + 1


def _h_parrafos(np, ns):
    return max(np, 1) + max(ns, 1) + 5


def _escribir_parrafos(ws, r0, c0, principales, secundarios):
    r = _titulo_desc(ws, r0, c0, 18, "INTERPRETACIONES PRINCIPALES · CAMPOS OBLIGATORIOS",
                     "Párrafos automáticos basados en los datos identificados, priorizando "
                     "campos obligatorios. Los de campos no obligatorios se muestran al "
                     "final, separados y etiquetados.")
    for p in principales:
        lineas = max(1, math.ceil(len(p) / 110))
        cell = _put(ws, r, c0, p, font=font_regular, align=alignment_wrap)
        ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c0 + 17)
        for cc in range(c0, c0 + 18):
            ws.cell(row=r, column=cc).border = border_thin
        ws.row_dimensions[r].height = max(28, lineas * 14 + 10)
        r += 1
    if not principales:
        cell = _put(ws, r, c0, "Sin interpretaciones para los datos actuales.",
                    font=font_italic, align=alignment_wrap)
        ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c0 + 17)
        r += 1
    r = _titulo_desc(ws, r, c0, 18, "CAMPOS NO OBLIGATORIOS · DE MENOR PRIORIDAD",
                     "Mismos análisis restringidos a campos no obligatorios; su corrección "
                     "es secundaria frente a los obligatorios.")
    for p in secundarios:
        lineas = max(1, math.ceil(len(p) / 110))
        cell = _put(ws, r, c0, p, font=font_italic, align=alignment_wrap)
        ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c0 + 17)
        for cc in range(c0, c0 + 18):
            ws.cell(row=r, column=cc).border = border_thin
        ws.row_dimensions[r].height = max(28, lineas * 14 + 10)
        r += 1
    if not secundarios:
        cell = _put(ws, r, c0, "Sin hallazgos en campos no obligatorios.", font=font_italic,
                    align=alignment_wrap)
        ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c0 + 17)
        r += 1
    return r + 1


def _h_detalle(n):
    return 2 * n + 7


def _panel_detalle(ws, r0, c0, d, r_indice):
    _put(ws, r0, c0, f"MÓDULO {d['modulo']} · {d['etiqueta']}", font=font_section, align=alignment_left)
    ws.merge_cells(start_row=r0, start_column=c0, end_row=r0, end_column=c0 + 6)
    _link(ws, r0, c0 + 7, f"B{r_indice}", "⬆ Índice", align=alignment_center)

    r = _cabecera(ws, r0 + 1, c0, ["Año", "N° Taladros", "Total Filas Año", "Vacías", "-1",
                                   "Total", "% del Año"])
    for idx, fila in enumerate(d["anual"]):
        zebra = fill_zebra if idx % 2 == 1 else None
        _put(ws, r, c0, fila["anio"], font=font_bold, fill=zebra, align=alignment_center)
        _put(ws, r, c0 + 1, fila["n_taladros"], fmt="#,##0", align=alignment_right, fill=zebra)
        _put(ws, r, c0 + 2, fila["total_filas"], fmt="#,##0", align=alignment_right, fill=zebra)
        _put(ws, r, c0 + 3, fila["v"], fmt="#,##0", align=alignment_right, fill=zebra)
        _put(ws, r, c0 + 4, fila["s"], fmt="#,##0", align=alignment_right, fill=zebra)
        _put(ws, r, c0 + 5, fila["total"], font=font_bold, fmt="#,##0", align=alignment_right, fill=zebra)
        _put(ws, r, c0 + 6, fila["pct_anio"] / 100.0, fmt="0.00%", align=alignment_right, fill=zebra)
        r += 1
    _put(ws, r, c0, "TOTAL", font=font_bold, align=alignment_center, fill=fill_kpi_gray)
    _put(ws, r, c0 + 1, "—", font=font_regular, align=alignment_center, fill=fill_kpi_gray)
    _put(ws, r, c0 + 2, "—", font=font_regular, align=alignment_center, fill=fill_kpi_gray)
    _put(ws, r, c0 + 3, sum(f["v"] for f in d["anual"]), font=font_bold, fmt="#,##0",
         align=alignment_right, fill=fill_kpi_gray)
    _put(ws, r, c0 + 4, sum(f["s"] for f in d["anual"]), font=font_bold, fmt="#,##0",
         align=alignment_right, fill=fill_kpi_gray)
    _put(ws, r, c0 + 5, sum(f["total"] for f in d["anual"]), font=font_bold, fmt="#,##0",
         align=alignment_right, fill=fill_kpi_gray)
    _put(ws, r, c0 + 6, "—", font=font_regular, align=alignment_center, fill=fill_kpi_gray)
    _cf_contadores(ws, r0 + 2, r - 1, c0 + 5)
    _cf_pct(ws, r0 + 2, r - 1, c0 + 6)
    r += 2

    r2 = _cabecera(ws, r, c0, ["Año", "Taladros (Vacías)", "Taladros (-1)", "Taladros (Total)"])
    for idx, fila in enumerate(d["taladros"]):
        zebra = fill_zebra if idx % 2 == 1 else None
        _put(ws, r2, c0, fila["anio"], font=font_bold, fill=zebra, align=alignment_center)
        _put(ws, r2, c0 + 1, fila["t_v"], fmt="#,##0", align=alignment_right, fill=zebra)
        _put(ws, r2, c0 + 2, fila["t_s"], fmt="#,##0", align=alignment_right, fill=zebra)
        _put(ws, r2, c0 + 3, fila["t_t"], fmt="#,##0", align=alignment_right, fill=zebra)
        r2 += 1
    _put(ws, r2, c0, "TOTAL", font=font_bold, align=alignment_center, fill=fill_accent_orange)
    _put(ws, r2, c0 + 1, "—", font=font_regular, align=alignment_center, fill=fill_accent_orange)
    _put(ws, r2, c0 + 2, "—", font=font_regular, align=alignment_center, fill=fill_accent_orange)
    _put(ws, r2, c0 + 3, d["total_taladros"], font=font_bold, fmt="#,##0",
         align=alignment_right, fill=fill_accent_orange)
    _cf_contadores(ws, r + 1, r2 - 1, c0 + 3)
    return r2 + 2


def _h_pareto(n):
    return n + 6


def _fondo_categoria(row):
    if row["afecta_rmr"]:
        return ROSA_FILL
    if row["requerido"]:
        return VERDE_FILL
    return AMBAR_FILL


def _color_barra(row):
    if row["afecta_rmr"]:
        return COLORES["rmr_fucsia"]
    if row["requerido"]:
        return COLORES["obligatorio_verde"]
    return COLORES["no_rmr_ambar"]


def _leyenda(ws, r, c0):
    chips = [
        ("RMR", COLORES["rmr_fucsia"], "Afectan al RMR"),
        ("OBL", COLORES["obligatorio_verde"], "Obligatorios"),
        ("NO", COLORES["no_rmr_ambar"], "No obligatorios"),
    ]
    col = c0
    for texto_corto, color, texto in chips:
        cel = _put(ws, r, col, texto_corto, font=font_header, align=alignment_center)
        cel.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        _put(ws, r, col + 1, texto, font=font_subtitle, align=alignment_left, border=False)
        col += 2


def _seccion_pareto(ws, r0, c0, pareto, gran_total):
    r = _titulo_desc(ws, r0, c0, 7, "ANÁLISIS DE PARETO · ¿QUÉ CAMPOS CONCENTRAN LOS "
                                   "DATOS FALTANTES?",
                     "Campos ordenados de mayor a menor cantidad de faltantes, con % del "
                     "total y % acumulado. Rosa: afectan al RMR · Verde: obligatorios · "
                     "Ámbar: no obligatorios.")
    _leyenda(ws, r - 1, c0 + 9)
    r = _cabecera(ws, r, c0, ["Módulo", "CAMPO", "VACÍAS", "-1", "OCURRENCIAS",
                              "% DEL TOTAL", "% ACUM."])
    for row in pareto:
        if row.get("divisor"):
            cell = _put(ws, r, c0, "Sub-ratings del cálculo RMR'76 / RMR'89", font=font_italic,
                        align=alignment_left, fill=fill_kpi_gray)
            ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c0 + 6)
            for cc in range(c0, c0 + 7):
                ws.cell(row=r, column=cc).border = border_thin
            r += 1
            continue
        fondo = _fondo_categoria(row)
        fuente = font_italic if row["es_subrating"] else font_regular
        _put(ws, r, c0, row["modulo"], font=font_bold, fill=fondo, align=alignment_center)
        _put(ws, r, c0 + 1, row["etiqueta"], font=fuente, fill=fondo, align=alignment_left)
        _put(ws, r, c0 + 2, row["v"], fmt="#,##0", align=alignment_right, fill=fondo)
        _put(ws, r, c0 + 3, row["s"], fmt="#,##0", align=alignment_right, fill=fondo)
        _put(ws, r, c0 + 4, row["total"], font=font_bold, fmt="#,##0", align=alignment_right, fill=fondo)
        _put(ws, r, c0 + 5, row["pct"] / 100.0, fmt="0.00%", align=alignment_right, fill=fondo)
        _put(ws, r, c0 + 6, row["acum"] / 100.0, fmt="0.00%", align=alignment_right, fill=fondo)
        r += 1
    if not pareto:
        cell = _put(ws, r, c0, "Sin datos para el Pareto.", font=font_italic, align=alignment_left)
        ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c0 + 6)
        r += 1
    else:
        _cf_contadores(ws, r0 + 3, r - 1, c0 + 4)
        _cf_pct(ws, r0 + 3, r - 1, c0 + 5)
        _cf_pct(ws, r0 + 3, r - 1, c0 + 6)
        total_v = sum(row["v"] for row in pareto if not row.get("divisor"))
        total_s = sum(row["s"] for row in pareto if not row.get("divisor"))
        _put(ws, r, c0, "TOTAL", font=font_bold, align=alignment_center, fill=fill_kpi_gray)
        _put(ws, r, c0 + 1, "—", font=font_regular, align=alignment_center, fill=fill_kpi_gray)
        _put(ws, r, c0 + 2, total_v, font=font_bold, fmt="#,##0", align=alignment_right, fill=fill_kpi_gray)
        _put(ws, r, c0 + 3, total_s, font=font_bold, fmt="#,##0", align=alignment_right, fill=fill_kpi_gray)
        _put(ws, r, c0 + 4, gran_total, font=font_bold, fmt="#,##0", align=alignment_right, fill=fill_kpi_gray)
        _put(ws, r, c0 + 5, 1.0, fmt="0.00%", align=alignment_right, fill=fill_kpi_gray)
        _put(ws, r, c0 + 6, 1.0, fmt="0.00%", align=alignment_right, fill=fill_kpi_gray)
        r += 1
        r = _nota(ws, r, c0, 7, f"Base: {gran_total} faltantes identificados = 100% "
                               f"(vacías + -1 de todos los módulos).")
    return r + 1


def _grafico_pareto(ws, r0, c0, pareto):
    top = [row for row in pareto if not row.get("divisor")][:15]
    if not top:
        return
    r1 = r0 + 2
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "¿Qué campos concentran los datos faltantes?"
    chart.y_axis.title = "Ocurrencias"
    chart.x_axis.title = None
    data_ref = Reference(ws, min_col=c0 + 4, min_row=r1, max_row=r1 + len(top) - 1)
    cats_ref = Reference(ws, min_col=c0 + 1, min_row=r1, max_row=r1 + len(top) - 1)
    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cats_ref)
    chart.legend = None
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    for i, row in enumerate(top):
        dp = DataPoint(idx=i)
        dp.graphicalProperties.solidFill = _color_barra(row)
        chart.series[0].data_points.append(dp)
    line = LineChart()
    line.style = 12
    line.add_data(Reference(ws, min_col=c0 + 6, min_row=r1, max_row=r1 + len(top) - 1),
                  titles_from_data=False)
    line.set_categories(cats_ref)
    line.y_axis.axId = 200
    line.y_axis.title = "% Acumulado"
    line.y_axis.crosses = "max"
    chart.y_axis.axId = 100
    chart += line
    chart.width = 24
    chart.height = 13
    ws.add_chart(chart, f"{get_column_letter(c0 + 9)}{r0 + 2}")


def _h_modulos():
    return 17


def _seccion_modulos(ws, r0, c0, modulos, r_indice, gran_total):
    r = _titulo_desc(ws, r0, c0, 8, "PANEL INTERACTIVO POR MÓDULO",
                     "Selecciona un módulo en el desplegable para actualizar "
                     "automáticamente las métricas de esta sección.", link_destino=r_indice)
    fila_selector = r
    _put(ws, r, c0, "Módulo:", font=font_bold, align=alignment_right)
    selector = _put(ws, r, c0 + 1, MODULOS_PANEL[0], font=font_bold, align=alignment_center,
                    border=False)
    selector.fill = NIVEL_FILLS["Puntual"]
    selector.border = border_kpi
    dv = DataValidation(type="list", formula1=f'"{",".join(MODULOS_PANEL)}"', allow_blank=False)
    dv.showDropDown = False
    ws.add_data_validation(dv)
    dv.add(f"{get_column_letter(c0 + 1)}{fila_selector}")
    r += 1

    col_h = 24
    for i, m in enumerate(modulos):
        cel_h = ws.cell(row=3, column=col_h + i, value=m["modulo"])
        cel_h.font = font_bold
        vals = [
            m["taladros_afectados"], m["faltantes"], m["vacias"], m["sin_info"],
            m["n_campos"], m["n_alta"],
        ]
        for t in m["top3"]:
            vals.append(t["etiqueta"])
            vals.append(t["total"])
        for j, v in enumerate(vals):
            ws.cell(row=4 + j, column=col_h + i, value=v)
    for i in range(3):
        ws.column_dimensions[get_column_letter(col_h + i)].hidden = True

    r = _cabecera(ws, r, c0, ["Métrica", "Valor"])
    for i, (label, _) in enumerate(METRICAS_PANEL):
        _put(ws, r + i, c0, label, font=font_regular, align=alignment_left,
             fill=fill_zebra if i % 2 == 1 else None)
        formula = (f"=INDEX({get_column_letter(col_h)}$4:{get_column_letter(col_h + 2)}$15,"
                   f"{i + 1},MATCH({get_column_letter(c0 + 1)}${fila_selector},"
                   f"{get_column_letter(col_h)}$3:{get_column_letter(col_h + 2)}$3,0))")
        cel = _put(ws, r + i, c0 + 1, formula, font=font_bold, align=alignment_center,
                   fill=fill_zebra if i % 2 == 1 else None)
        if i in (1, 2, 3, 4, 5, 7, 9, 11):
            cel.number_format = "#,##0"
    r += len(METRICAS_PANEL)
    r = _nota(ws, r, c0, 8, f"Los datos se calculan sobre el total identificado de cada "
                           f"módulo ({gran_total} faltantes globales).")
    return r + 1


def crear_hoja_analisis_vacios(wb, result, index=3):
    ws = wb.create_sheet(title=HOJA_TITULO, index=index)
    ws.views.sheetView[0].showGridLines = True

    tabla_a = result.tablas.get("A", [])
    tabla_b = result.tablas.get("B", [])
    tabla_c = result.tablas.get("C", [])
    sub = result.subratings
    vista_alta = result.vista_alta
    detalles = result.detalles
    pareto = result.pareto
    parrafos = result.parrafos
    parrafos_sec = result.parrafos_secundarios
    modulos = result.modulos

    r_tablas = 13
    h_abc = max(_h_campos(len(tabla_a)), _h_campos(len(tabla_b)), _h_campos(len(tabla_c)))
    r_sub = r_tablas + h_abc + 1
    h_sub = _h_sub(len(sub))
    r_alta = r_sub + h_sub + 1
    h_alta = _h_alta(len(vista_alta))
    r_parrafos = r_alta + h_alta + 1
    h_parrafos = _h_parrafos(len(parrafos), len(parrafos_sec))
    r_detalles = r_parrafos + h_parrafos + 1
    bandas = [detalles[i:i + 2] for i in range(0, len(detalles), 2)]
    h_bandas = []
    for banda in bandas:
        h = _h_detalle(len(banda[0]["anual"]))
        if len(banda) > 1:
            h = max(h, _h_detalle(len(banda[1]["anual"])))
        h_bandas.append(h + 1)
    r_pareto = r_detalles + (sum(h_bandas) if bandas else 1)
    h_pareto = _h_pareto(len(pareto))
    r_panel = r_pareto + h_pareto + 1
    h_panel = _h_modulos()
    r_nota = r_panel + h_panel + 1
    anclas_detalle = []
    acc = r_detalles
    for i, banda in enumerate(bandas):
        anclas_detalle.extend([acc] * len(banda))
        acc += h_bandas[i]

    r_indice = 10
    enlaces_indice = [
        ("Tabla A", f"B{r_tablas}"),
        ("Tabla B", f"H{r_tablas}"),
        ("Tabla C", f"N{r_tablas}"),
        ("Campos ALTA", f"B{r_alta}"),
        ("Lectura", f"B{r_parrafos}"),
        ("Detalles", f"B{r_detalles}"),
        ("Pareto", f"B{r_pareto}"),
        ("Panel Módulos", f"B{r_panel}"),
    ]
    for i, (texto, destino) in enumerate(enlaces_indice):
        _link(ws, r_indice, 2 + i, destino, texto, align=alignment_center)

    _put(ws, 2, 2, "ANÁLISIS DE DATOS FALTANTES", font=font_title)
    _put(ws, 3, 2, "Consolidado de campos vacíos y sin información (-1) por módulo, con "
                   "nivel de atención, tendencias anuales y análisis de Pareto.",
         font=font_subtitle)

    _kpi_card(ws, 5, 2, "TALADROS AFECTADOS", result.total_taladros_afectados,
              fill_kpi_gray, font_kpi_val_blue)
    _kpi_card(ws, 5, 4, "FALTANTES TOTALES", result.total_faltantes,
              fill_kpi_gray, font_kpi_val_orange)
    _kpi_card(ws, 5, 6, "CAMPOS EN NIVEL ALTA", result.num_campos_alta,
              fill_kpi_gray, font_kpi_val_red)

    _put(ws, 5, 9, "BASE DEL %", font=font_kpi_lbl, align=alignment_center)
    sel = _put(ws, 5, 10, OPCION_MODULO, font=font_bold, align=alignment_center, border=False)
    sel.fill = NIVEL_FILLS["Puntual"]
    sel.border = border_kpi
    _put(ws, 6, 9, "Cambia el selector para comparar los porcentajes: por módulo o sobre "
                   "el total identificado (global).", font=font_subtitle, align=alignment_wrap,
         border=False)
    ws.merge_cells(start_row=6, start_column=9, end_row=8, end_column=13)
    dv = DataValidation(type="list", formula1=f'"{OPCION_MODULO},{OPCION_GLOBAL}"',
                        allow_blank=False)
    dv.showDropDown = False
    ws.add_data_validation(dv)
    dv.add("J5")

    _seccion_tabla_ab(ws, r_tablas, 2, "TABLA A · CAMPOS VACÍOS",
                      "Celdas sin contenido (posible descuido de registro). Nivel de "
                      "Atención según el % de celdas vacías del módulo.",
                      tabla_a, "v", True, result.total_vacias, "celdas vacías identificadas",
                      "Vacías (N)")
    _seccion_tabla_ab(ws, r_tablas, 8, "TABLA B · SIN INFORMACIÓN (-1)",
                      "Celdas con valor -1 (sin dato registrado a propósito). Nivel de "
                      "Atención según el % de -1 del módulo.",
                      tabla_b, "s", True, result.total_sin_info, "valores -1 identificados",
                      "Sin info (N)")
    _seccion_tabla_c(ws, r_tablas, 14, tabla_c, result.total_faltantes)

    _seccion_subratings(ws, r_sub, 2, sub)
    _seccion_alta(ws, r_alta, 2, vista_alta, anclas_detalle, r_indice, result.total_faltantes)
    _escribir_parrafos(ws, r_parrafos, 2, parrafos, parrafos_sec)

    r_banda = r_detalles
    for banda in bandas:
        r_ini = r_banda
        for j, d in enumerate(banda):
            c0 = 2 if j == 0 else 11
            _panel_detalle(ws, r_ini, c0, d, r_indice)
        r_banda += h_bandas[bandas.index(banda)]

    _seccion_pareto(ws, r_pareto, 2, pareto, result.total_faltantes)
    _grafico_pareto(ws, r_pareto, 2, pareto)

    _seccion_modulos(ws, r_panel, 2, modulos, r_indice, result.total_faltantes)

    nota_rmr = ("Los campos en ROSA son parámetros que alimentan el cálculo del RMR "
                "(Bieniawski): provienen del Logueo General (LGG) y de la orientación "
                "estructural (Dip/Azimut). En VERDE: campos obligatorios. En ÁMBAR: campos "
                "no obligatorios. Su ausencia impide obtener una clasificación "
                "geomecánica válida del macizo rocoso.")
    cel = _put(ws, r_nota, 2, nota_rmr, font=font_italic, align=alignment_wrap)
    ws.merge_cells(start_row=r_nota, start_column=2, end_row=r_nota, end_column=15)
    ws.row_dimensions[r_nota].height = 45

    for col, w in ANCHO_COLUMNAS.items():
        ws.column_dimensions[col].width = w
    return ws

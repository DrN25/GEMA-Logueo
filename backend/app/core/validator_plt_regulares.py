"""
validator_plt_regulares.py — Motor de Auditoría y Validación QA/QC para Ensayos PLT Regulares (DDH).
Ejecuta la validación fila a fila según el catálogo oficial de reglas (reglas_plt_regulares.md).
Incorpora:
  1. Control de duplicados (muestras duplicadas, tramos repetidos, tramos cruzados, mediciones repetitivas).
  2. Consistencia y auto-sugerencia de Factor K vs Litología basada en el catálogo maestro SSOT.
  3. Cruce geomecánico opcional con Logueo General (LGG) para corridas, profundidades, litología y dureza ISRM.
"""

from collections import defaultdict
from datetime import datetime, date
import io
import math
import unicodedata
import numpy as np
import pandas as pd
from typing import Dict, List, Any, Optional, Tuple, Union

from app.core.rules_plt_regulares import (
    CATEGORIES_REGISTRY_PLT_REGULARES,
    NOMINAL_DIAMETERS,
    VALID_ENSAYO_TYPES,
    VALID_NOMINATIONS,
    VALID_FRACTURE_TYPES,
    VALID_ROTURA_DIRECTIONS,
    VALID_GEOLOGIC_GROUPS,
    ISRM_SCALE,
    OFFICIAL_34_COLUMNS,
    MANDATORY_PLT_COLUMNS,
    resolve_expected_k_and_type,
)


def strip_accents(s: str) -> str:
    """Elimina tildes y diacríticos de un string."""
    if not s:
        return ""
    return "".join(c for c in unicodedata.normalize("NFD", str(s)) if unicodedata.category(c) != "Mn")


def clean_str(val: Any) -> str:
    """Limpia cadenas, elimina caracteres invisibles y espacios en blanco."""
    if val is None or pd.isna(val):
        return ""
    s = str(val).strip()
    if s.lower() in ("nan", "none", "null", "#value!", "#ref!", "#div/0!"):
        return ""
    return s


def is_valid_caja(val: Any) -> bool:
    """Valida si el valor de Nro Caja es un número entero positivo o un rango válido (ej. 34-35, 34 - 35, 34/35)."""
    if val is None:
        return False
    s = str(val).strip()
    if not s or s.lower() in ("nan", "none", "null", ""):
        return False
    try:
        f = float(s)
        return f > 0
    except ValueError:
        pass
    # Rangos como 34-35, 34 - 35, 34/35, 34 a 35, 34_35
    s_clean = s.replace(" a ", "-").replace(" A ", "-").replace("/", "-").replace("_", "-")
    parts = s_clean.split("-")
    if len(parts) == 2:
        try:
            v1 = float(parts[0].strip())
            v2 = float(parts[1].strip())
            return v1 > 0 and v2 >= v1
        except ValueError:
            pass
    return False


def to_float(val: Any) -> Optional[float]:
    """Convierte un valor a float si es numérico válido, o None."""
    if val is None or pd.isna(val):
        return None
    s = str(val).strip().replace(",", ".")
    if s in ("", "-", "N/A", "NA", "null", "None"):
        return None
    try:
        f = float(s)
        if math.isnan(f) or math.isinf(f):
            return None
        return f
    except (ValueError, TypeError):
        return None


def to_int(val: Any) -> Optional[int]:
    """Convierte un valor a entero seguro."""
    f = to_float(val)
    if f is None:
        return None
    return int(round(f))


def parse_date(val: Any) -> Optional[date]:
    """Parsea fechas desde seriales de Excel, objetos datetime o strings."""
    if val is None or pd.isna(val):
        return None
    if isinstance(val, (datetime, pd.Timestamp)):
        return val.date()
    if isinstance(val, date):
        return val

    # Si es un número serial de Excel (ej. 44819)
    try:
        f = float(val)
        if 35000 <= f <= 60000:
            return pd.to_datetime("1899-12-30") + pd.to_timedelta(f, unit="D")
    except (ValueError, TypeError):
        pass

    s = clean_str(val)
    if not s:
        return None

    # Formatos comunes
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d", "%d-%m-%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue

    try:
        dt = pd.to_datetime(s, errors="coerce")
        if pd.notna(dt):
            return dt.date()
    except Exception:
        pass

    return None


def get_isrm_grade(ucs: float) -> str:
    """Retorna la clase ISRM oficial (R0-R6 o Suelo) según el UCS calculado en MPa."""
    if ucs is None or math.isnan(ucs):
        return ""
    for limit, grade in ISRM_SCALE:
        if ucs < limit:
            return grade
    return "R6"


def extract_lgg_dataframe(lgg_source: Any) -> Optional[pd.DataFrame]:
    """
    Extrae y normaliza las corridas oficiales de LGG desde una ruta de archivo, bytes o DataFrame.
    Detecta la cabecera y columnas requeridas: taladro, de:, a:, litologías e ISRM.
    """
    if lgg_source is None:
        return None

    if isinstance(lgg_source, pd.DataFrame):
        return lgg_source

    try:
        # Intentar lectura ultrarrápida con Calamine
        try:
            from python_calamine import CalamineWorkbook
            if isinstance(lgg_source, (bytes, io.BytesIO)):
                data_bytes = lgg_source.getvalue() if isinstance(lgg_source, io.BytesIO) else lgg_source
                wb = CalamineWorkbook.from_filelike(io.BytesIO(data_bytes))
            else:
                wb = CalamineWorkbook.from_path(str(lgg_source))

            sheet_name = "BD LGG" if "BD LGG" in wb.sheet_names else wb.sheet_names[0]
            sheet = wb.get_sheet_by_name(sheet_name)
            raw_rows = sheet.to_python()
        except Exception:
            import openpyxl
            wb_file = io.BytesIO(lgg_source) if isinstance(lgg_source, bytes) else lgg_source
            wb_px = openpyxl.load_workbook(wb_file, data_only=True)
            sheet_name = "BD LGG" if "BD LGG" in wb_px.sheetnames else wb_px.sheetnames[0]
            ws = wb_px[sheet_name]
            raw_rows = [[cell.value for cell in row] for row in ws.iter_rows()]

        if not raw_rows or len(raw_rows) < 2:
            return None

        # Detectar fila de cabecera buscando celdas exactas 'taladro', 'de'/'desde' y 'a'/'hasta'
        header_idx = -1
        for i in range(min(15, len(raw_rows))):
            cells = [str(c or "").lower().strip().replace(":", "").replace(" ", "") for c in raw_rows[i]]
            if "taladro" in cells and ("de" in cells or "desde" in cells) and ("a" in cells or "hasta" in cells):
                header_idx = i
                break

        if header_idx == -1:
            header_idx = 0

        raw_headers = [str(c or "").replace("\n", " ").strip() for c in raw_rows[header_idx]]
        df_lgg = pd.DataFrame(raw_rows[header_idx + 1:], columns=raw_headers)

        col_taladro = None
        col_desde = None
        col_hasta = None
        col_l1 = None
        col_l2 = None
        col_l3 = None
        col_isrm = None

        for col in df_lgg.columns:
            c_norm = strip_accents(str(col).lower().replace(" ", "").replace("_", "").replace(":", ""))
            if not col_taladro and any(p in c_norm for p in ("taladro", "sondaje", "drillhole")):
                col_taladro = col
            elif not col_desde and c_norm in ("de", "desde", "dem", "desdem", "from"):
                col_desde = col
            elif not col_hasta and c_norm in ("a", "hasta", "am", "hastam", "to"):
                col_hasta = col
            elif not col_l1 and any(p in c_norm for p in ("lito1", "litologia1", "lito12023")):
                col_l1 = col
            elif not col_l2 and any(p in c_norm for p in ("lito2", "litologia2", "lito22023")):
                col_l2 = col
            elif not col_l3 and any(p in c_norm for p in ("lito3", "litologia3", "lito32023")):
                col_l3 = col
            elif not col_isrm and any(p in c_norm for p in ("resistestimada", "isrm", "resistencia", "resistmax")):
                col_isrm = col

        if not col_taladro or not col_desde or not col_hasta:
            return None

        res_df = pd.DataFrame()
        res_df["taladro"] = df_lgg[col_taladro].astype(str).str.strip().str.upper()
        res_df["desde_m"] = pd.to_numeric(df_lgg[col_desde].astype(str).str.replace(",", "."), errors="coerce")
        res_df["hasta_m"] = pd.to_numeric(df_lgg[col_hasta].astype(str).str.replace(",", "."), errors="coerce")
        res_df["lito1"] = df_lgg[col_l1].astype(str).str.strip().str.upper() if col_l1 else ""
        res_df["lito2"] = df_lgg[col_l2].astype(str).str.strip().str.upper() if col_l2 else ""
        res_df["lito3"] = df_lgg[col_l3].astype(str).str.strip().str.upper() if col_l3 else ""
        res_df["isrm"] = df_lgg[col_isrm].astype(str).str.strip().str.upper() if col_isrm else ""

        res_df = res_df[res_df["taladro"].notna() & (res_df["taladro"] != "") & (res_df["taladro"] != "NAN")].copy()
        res_df = res_df[res_df["desde_m"].notna() & res_df["hasta_m"].notna()].copy()
        return res_df

    except Exception as e:
        print(f"[!] Error al extraer DataFrame de LGG: {e}")
        return None

class Counter_custom(dict):
    def __missing__(self, key):
        return 0


class PltRegularesValidator:
    """
    Motor de Auditoría y Validación QA/QC para Ensayos PLT Regulares (DDH).
    Soporta modo autónomo (solo planilla PLT) o modo cruzado geomecánico (PLT + LGG).
    """

    def __init__(self, tolerance_is: float = 0.05, tolerance_f: float = 0.05, tolerance_ucs: float = 0.5):
        self.tol_is = tolerance_is
        self.tol_f = tolerance_f
        self.tol_ucs = tolerance_ucs

    def audit_dataframe(self, df: pd.DataFrame, df_lgg: Optional[pd.DataFrame] = None) -> Dict[str, Any]:
        """
        Audita un DataFrame completo de ensayos PLT regulares y genera el diagnóstico estructurado.
        Si se suministra `df_lgg`, ejecuta adicionalmente las validaciones cruzadas con Logueo General.
        """
        # Normalizar nombres de columnas
        col_map = {}
        for c in df.columns:
            clean_c = str(c).replace("\n", " ").strip()
            col_map[c] = clean_c
        df = df.rename(columns=col_map)

        total_rows = len(df)
        anomalies: List[Dict[str, Any]] = []
        drillhole_stats = defaultdict(lambda: {"total": 0, "alertas": 0, "advertencias": 0, "vacios": 0})
        campaign_stats = defaultdict(lambda: {"total": 0, "alertas": 0, "advertencias": 0, "vacios": 0})
        category_counter = Counter_custom()
        column_error_counter = Counter_custom()
        severity_counter = {"ALERTA": 0, "ADVERTENCIA": 0, "VACIO": 0}

        row_severity_flags = [False] * total_rows  # True si tiene ALERTA o VACIO
        today = datetime.now().date()

        # Helper para registrar anomalías
        def add_anomaly(excel_row: int, campana: str, taladro: str, muestra: str,
                        from_val: Any, to_val: Any, col_name: str, cat_code: str,
                        message: str, sev_override: Optional[str] = None):
            cat = CATEGORIES_REGISTRY_PLT_REGULARES.get(cat_code)
            sev = sev_override or (cat.severity if cat else "ALERTA")
            cat_name = cat.name if cat else cat_code

            anomalies.append({
                "row_index": excel_row,
                "campana": campana,
                "taladro": taladro,
                "muestra": muestra or f"Fila_{excel_row}",
                "from_m": to_float(from_val),
                "to_m": to_float(to_val),
                "columna": col_name,
                "category_code": cat_code,
                "category_name": cat_name,
                "severity": sev,
                "message": message
            })
            category_counter[cat_code] += 1
            column_error_counter[col_name] += 1
            severity_counter[sev] = severity_counter.get(sev, 0) + 1

            if sev == "ALERTA":
                drillhole_stats[taladro]["alertas"] += 1
                campaign_stats[campana]["alertas"] += 1
                row_severity_flags[excel_row - 2] = True
            elif sev == "VACIO":
                drillhole_stats[taladro]["vacios"] += 1
                campaign_stats[campana]["vacios"] += 1
                row_severity_flags[excel_row - 2] = True
            elif sev == "ADVERTENCIA":
                drillhole_stats[taladro]["advertencias"] += 1
                campaign_stats[campana]["advertencias"] += 1

        # =====================================================================
        # FASE 1: AUDITORÍA DE DUPLICADOS Y MEDICIONES REPETITIVAS A NIVEL DATASET
        # =====================================================================
        seen_samples: Dict[Tuple[str, str], int] = {}       # (taladro, muestra) -> first_row
        seen_intervals: Dict[Tuple[str, float, float], int] = {}  # (taladro, from, to) -> first_row

        # Indexar intervalos por taladro para chequear solapamientos/cruces
        intervals_by_dh = defaultdict(list)
        p_measurements_by_group = defaultdict(list)

        for idx in range(total_rows):
            r = df.iloc[idx]
            row_num = idx + 2
            t_clean = clean_str(r.get("Taladro")).upper()
            m_clean = clean_str(r.get("Nro Muestra")).upper()
            f_val = to_float(r.get("From"))
            t_val = to_float(r.get("To"))
            fecha_str = clean_str(r.get("Fecha"))
            p_val = to_float(r.get("P instr (kN)"))
            d_val = to_float(r.get("D (mm)"))
            c_key = str(to_int(r.get("Campaña")) or clean_str(r.get("Campaña")) or "S/C")
            t_key = t_clean or "S/T"

            # 1. Muestra duplicada en el taladro
            if t_clean and m_clean:
                sample_key = (t_clean, m_clean)
                if sample_key in seen_samples:
                    prev_row = seen_samples[sample_key]
                    add_anomaly(row_num, c_key, t_key, m_clean, f_val, t_val,
                                "Nro Muestra", "CAT_PLT_MUESTRA_DUPLICADA",
                                f"Muestra '{m_clean}' repetida en el taladro '{t_key}' (ya fue registrada en la fila {prev_row}).",
                                sev_override="ALERTA")
                else:
                    seen_samples[sample_key] = row_num

            # 2. Mismo tramo ensayado más de una vez
            if t_clean and f_val is not None and t_val is not None:
                int_key = (t_clean, round(f_val, 3), round(t_val, 3))
                if int_key in seen_intervals:
                    prev_row = seen_intervals[int_key]
                    add_anomaly(row_num, c_key, t_key, m_clean, f_val, t_val,
                                "From", "CAT_PLT_TRAMO_DUPLICADO",
                                f"El tramo [{f_val:.2f} - {t_val:.2f} m] se encuentra repetido en el taladro '{t_key}' (ya registrado en fila {prev_row}).",
                                sev_override="ALERTA")
                else:
                    seen_intervals[int_key] = row_num

                intervals_by_dh[t_clean].append({
                    "row_num": row_num, "campana": c_key, "taladro": t_key,
                    "muestra": m_clean, "from": f_val, "to": t_val
                })

            # 3. Agrupación para medición repetitiva
            if t_clean and fecha_str and p_val is not None and d_val is not None:
                p_group_key = (t_clean, fecha_str, round(p_val, 2), round(d_val, 1))
                p_measurements_by_group[p_group_key].append({
                    "row_num": row_num, "campana": c_key, "taladro": t_key,
                    "muestra": m_clean, "from": f_val, "to": t_val, "p": p_val, "d": d_val
                })

        # 3. Tramos que se cruzan o montan dentro de cada taladro
        for dh, items in intervals_by_dh.items():
            items_sorted = sorted(items, key=lambda x: x["from"])
            for i in range(1, len(items_sorted)):
                curr = items_sorted[i]
                prev = items_sorted[i - 1]
                # Si el From actual es menor al To anterior (con tolerancia de 5mm)
                if curr["from"] < (prev["to"] - 0.005) and (curr["from"] != prev["from"] or curr["to"] != prev["to"]):
                    add_anomaly(curr["row_num"], curr["campana"], curr["taladro"], curr["muestra"],
                                curr["from"], curr["to"], "From", "CAT_PLT_TRAMOS_CRUZADOS",
                                f"El tramo [{curr['from']:.2f} - {curr['to']:.2f} m] se cruza con la muestra anterior {prev['muestra']} [{prev['from']:.2f} - {prev['to']:.2f} m] de la fila {prev['row_num']}.",
                                sev_override="ALERTA")

        # 4. Carga de ensayo repetida continuamente (posible copia de datos)
        for (dh, f_str, p_k, d_k), p_list in p_measurements_by_group.items():
            if len(p_list) >= 4 and (p_k == round(p_k)):  # Si se repite 4+ veces un entero redondo
                for it in p_list:
                    add_anomaly(it["row_num"], it["campana"], it["taladro"], it["muestra"],
                                it["from"], it["to"], "P instr (kN)", "CAT_PLT_CARGA_REPETIDA",
                                f"Carga P={it['p']:.2f} kN repetida de forma idéntica en {len(p_list)} muestras seguidas el día {f_str} (revisar posible copia de datos).",
                                sev_override="ADVERTENCIA")

        # Indexar LGG por taladro si está activo el modo cruzado
        lgg_by_dh = {}
        has_lgg = False
        if df_lgg is not None and not df_lgg.empty:
            has_lgg = True
            for dh_name, grp in df_lgg.groupby("taladro"):
                lgg_by_dh[str(dh_name).strip().upper()] = grp

        # =====================================================================
        # FASE 2: AUDITORÍA FILA A FILA (34 COLUMNAS, FACTOR K Y CRUCE CON LGG)
        # =====================================================================
        for idx in range(total_rows):
            row = df.iloc[idx]
            excel_row_num = idx + 2

            # Extraer campos
            campana_raw = row.get("Campaña")
            fecha_raw = row.get("Fecha")
            taladro_raw = row.get("Taladro")
            muestra_raw = row.get("Nro Muestra")
            caja_raw = row.get("Nro Caja")

            c_desde_raw = row.get("Corrida Desde (m)")
            c_hasta_raw = row.get("Corrida Hasta (m)")
            from_raw = row.get("From")
            to_raw = row.get("To")
            verif_corrida_raw = row.get("Verif. corrida")
            long_corrida_raw = row.get("Long. de Corrida (m)")

            este_raw = row.get("Este (m)")
            norte_raw = row.get("Norte (m)")
            cota_raw = row.get("Elevación (msnm)")

            long_muestra_raw = row.get("Long. de Muestra (mm)")
            tipo_ensayo_raw = row.get("Tipo de Ensayo")
            nominacion_raw = row.get("Diametro de Taladro")
            d_raw = row.get("D (mm)")
            verif_long_raw = row.get("Verif. de longitud")

            lito1_raw = row.get("Litologia 1")
            lito2_raw = row.get("Litologia 2")
            lito3_raw = row.get("Litologia 3")
            tipo_lito_raw = row.get("Tipo litológico")

            p_raw = row.get("P instr (kN)")
            tipo_rotura_raw = row.get("Tipo de Rotura")
            dir_rotura_raw = row.get("Dirección de rotura")
            ejecutado_raw = row.get("Ejecutado por")

            is_raw = row.get("Is (Mpa)")
            fact_corr_raw = row.get("Fact. Corr")
            is50_raw = row.get("Is(50) (Mpa)")
            factor_k_raw = row.get("Factor K")
            ucs_raw = row.get("UCS")
            isrm_raw = row.get("ISRM Indice R")

            # Normalizar claves
            taladro_str = clean_str(taladro_raw)
            muestra_str = clean_str(muestra_raw)
            campana_int = to_int(campana_raw)
            campana_key = str(campana_int) if campana_int else (clean_str(campana_raw) or "S/C")
            taladro_key = taladro_str or "S/T"

            drillhole_stats[taladro_key]["total"] += 1
            campaign_stats[campana_key]["total"] += 1

            # Helper local vinculado a la fila
            def reg_err(col: str, cat: str, msg: str, sev: Optional[str] = None):
                add_anomaly(excel_row_num, campana_key, taladro_key, muestra_str,
                            from_raw, to_raw, col, cat, msg, sev_override=sev)

            # --- 1. Identificación y Fecha ---
            if not taladro_str:
                reg_err("Taladro", "CAT_CAMPO_OBLIGATORIO_VACIO", "El campo Taladro es obligatorio y se encuentra vacío.")
            if not muestra_str:
                reg_err("Nro Muestra", "CAT_CAMPO_OBLIGATORIO_VACIO", "El campo Nro Muestra es obligatorio y se encuentra vacío.")

            if caja_raw is None or clean_str(caja_raw) == "":
                reg_err("Nro Caja", "CAT_CAMPO_OBLIGATORIO_VACIO", "El campo Nro Caja es obligatorio y se encuentra vacío.")
            elif not is_valid_caja(caja_raw):
                reg_err("Nro Caja", "CAT_PLT_CAJA_INVALIDA", f"Nro de Caja '{caja_raw}' no es válido (debe ser un entero positivo >= 1 o rango ej. '34-35').")

            if campana_raw is None or clean_str(campana_raw) == "":
                reg_err("Campaña", "CAT_CAMPO_OBLIGATORIO_VACIO", "El campo Campaña es obligatorio y se encuentra vacío.")
            elif campana_int is None or not (2000 <= campana_int <= 2035):
                reg_err("Campaña", "CAT_PLT_CAMPANA_INVALIDA", f"Año de Campaña '{campana_raw}' no es válido (debe ser un año entre 2000 y 2035).")

            fecha_dt = parse_date(fecha_raw)
            if fecha_raw is None or clean_str(fecha_raw) == "":
                reg_err("Fecha", "CAT_CAMPO_OBLIGATORIO_VACIO", "El campo Fecha es obligatorio y se encuentra vacío.")
            elif fecha_dt is None:
                reg_err("Fecha", "CAT_PLT_FECHA_INVALIDA", f"Fecha de ensayo '{fecha_raw}' posee un formato inválido o no reconocido.")
            else:
                if fecha_dt > today:
                    reg_err("Fecha", "CAT_PLT_FECHA_FUTURA", f"Fecha de ensayo '{fecha_dt}' es posterior a la fecha actual ({today}).")
                if campana_int and fecha_dt.year != campana_int:
                    reg_err("Fecha", "CAT_PLT_CAMPANA_INVALIDA", f"Año de la fecha ({fecha_dt.year}) no coincide con la Campaña ({campana_int}).")

            # --- 2. Corridas y Tramos ---
            c_desde_f = to_float(c_desde_raw)
            c_hasta_f = to_float(c_hasta_raw)
            from_f = to_float(from_raw)
            to_f = to_float(to_raw)

            if c_desde_raw is None or clean_str(c_desde_raw) == "":
                reg_err("Corrida Desde (m)", "CAT_CAMPO_OBLIGATORIO_VACIO", "El campo Corrida Desde es obligatorio y se encuentra vacío.")
            elif c_desde_f is None or c_desde_f < 0:
                reg_err("Corrida Desde (m)", "CAT_PLT_CORRIDA_INCONGRUENTE", f"Corrida Desde ({c_desde_raw}) no puede ser un valor negativo.")

            if c_hasta_raw is None or clean_str(c_hasta_raw) == "":
                reg_err("Corrida Hasta (m)", "CAT_CAMPO_OBLIGATORIO_VACIO", "El campo Corrida Hasta es obligatorio y se encuentra vacío.")
            elif c_hasta_f is None or (c_desde_f is not None and c_hasta_f <= c_desde_f):
                reg_err("Corrida Hasta (m)", "CAT_PLT_CORRIDA_INCONGRUENTE", f"Corrida Hasta ({c_hasta_raw}) debe ser estrictamente mayor a Corrida Desde ({c_desde_raw}).")

            if from_raw is None or clean_str(from_raw) == "":
                reg_err("From", "CAT_CAMPO_OBLIGATORIO_VACIO", "El campo From es obligatorio y se encuentra vacío.")
            elif from_f is None or from_f < 0:
                reg_err("From", "CAT_PLT_TRAMO_MUESTRA_INCONGRUENTE", f"From ({from_raw}) no puede ser un valor negativo.")

            if to_raw is None or clean_str(to_raw) == "":
                reg_err("To", "CAT_CAMPO_OBLIGATORIO_VACIO", "El campo To es obligatorio y se encuentra vacío.")
            elif to_f is None or (from_f is not None and to_f <= from_f):
                reg_err("To", "CAT_PLT_TRAMO_MUESTRA_INCONGRUENTE", f"To ({to_raw}) debe ser estrictamente mayor a From ({from_raw}).")

            # Verificación geométrica de contención en corrida de PLT
            if c_desde_f is not None and c_hasta_f is not None and from_f is not None and to_f is not None:
                is_contained = (from_f >= c_desde_f - 0.01) and (to_f <= c_hasta_f + 0.01)
                if not is_contained:
                    reg_err("Verif. corrida", "CAT_PLT_MUESTRA_FUERA_CORRIDA",
                            f"La muestra [{from_f:.2f}, {to_f:.2f}] excede los límites de la corrida reportada [{c_desde_f:.2f}, {c_hasta_f:.2f}].")

                expected_verif = "OK" if is_contained else "ERROR"
                verif_c_str = clean_str(verif_corrida_raw).upper()
                if verif_corrida_raw is None or verif_c_str == "":
                    reg_err("Verif. corrida", "CAT_CAMPO_OBLIGATORIO_VACIO", "El campo Verif. corrida se encuentra vacío.")
                elif verif_c_str not in ("OK", "SI", "SÍ", "ERROR", "NO", "FALSE"):
                    reg_err("Verif. corrida", "CAT_PLT_VERIF_CORRIDA_INCONGRUENTE",
                            f"Valor '{verif_corrida_raw}' no coincide con el estado esperado '{expected_verif}'.")

                expected_long_c = c_hasta_f - c_desde_f
                long_c_f = to_float(long_corrida_raw)
                if long_corrida_raw is None or clean_str(long_corrida_raw) == "":
                    reg_err("Long. de Corrida (m)", "CAT_CAMPO_OBLIGATORIO_VACIO", "El campo Long. de Corrida se encuentra vacío.")
                elif long_c_f is None or abs(long_c_f - expected_long_c) > 0.02:
                    reg_err("Long. de Corrida (m)", "CAT_PLT_LONGITUD_CORRIDA_INCONGRUENTE",
                            f"Longitud reportada ({long_corrida_raw}) difiere de ({c_hasta_f:.2f} - {c_desde_f:.2f} = {expected_long_c:.2f} m).")

            # --- 3. Coordenadas ---
            este_f = to_float(este_raw)
            if este_raw is None or clean_str(este_raw) == "":
                reg_err("Este (m)", "CAT_CAMPO_OBLIGATORIO_VACIO", "Coordenada Este es obligatoria y se encuentra vacía.")
            elif este_f is None or este_f <= 0:
                reg_err("Este (m)", "CAT_PLT_COORD_ESTE_RANGO", f"Coordenada Este ({este_raw}) debe ser un valor positivo mayor a cero.")

            norte_f = to_float(norte_raw)
            if norte_raw is None or clean_str(norte_raw) == "":
                reg_err("Norte (m)", "CAT_CAMPO_OBLIGATORIO_VACIO", "Coordenada Norte es obligatoria y se encuentra vacía.")
            elif norte_f is None or norte_f <= 0:
                reg_err("Norte (m)", "CAT_PLT_COORD_NORTE_RANGO", f"Coordenada Norte ({norte_raw}) debe ser un valor positivo mayor a cero.")

            cota_f = to_float(cota_raw)
            if cota_raw is None or clean_str(cota_raw) == "":
                reg_err("Elevación (msnm)", "CAT_CAMPO_OBLIGATORIO_VACIO", "Cota topográfica Elevación es obligatoria y se encuentra vacía.")
            elif cota_f is None or cota_f <= 0:
                reg_err("Elevación (msnm)", "CAT_PLT_ELEVACION_RANGO", f"Elevación ({cota_raw}) debe ser un valor positivo mayor a cero.")

            # --- 4. Geometría del Testigo ---
            long_m_f = to_float(long_muestra_raw)
            if long_muestra_raw is None or clean_str(long_muestra_raw) == "":
                reg_err("Long. de Muestra (mm)", "CAT_CAMPO_OBLIGATORIO_VACIO", "Longitud de muestra es obligatoria y se encuentra vacía.")
            elif long_m_f is None or long_m_f <= 0:
                reg_err("Long. de Muestra (mm)", "CAT_PLT_LONGITUD_MUESTRA_RANGO", f"Longitud de muestra ({long_muestra_raw} mm) debe ser mayor a cero.")

            tipo_ensayo_str = clean_str(tipo_ensayo_raw).upper()
            if not tipo_ensayo_str:
                reg_err("Tipo de Ensayo", "CAT_CAMPO_OBLIGATORIO_VACIO", "Tipo de Ensayo es obligatorio y se encuentra vacío.")
            elif tipo_ensayo_str not in VALID_ENSAYO_TYPES:
                reg_err("Tipo de Ensayo", "CAT_PLT_TIPO_ENSAYO_INVALIDO", f"Tipo de ensayo '{tipo_ensayo_raw}' no es válido (debe ser D, A o B).")

            nom_str = clean_str(nominacion_raw).upper().replace(" ", "")
            if not nom_str:
                reg_err("Diametro de Taladro", "CAT_CAMPO_OBLIGATORIO_VACIO", "Diámetro de taladro es obligatorio y se encuentra vacío.")
            elif nom_str not in VALID_NOMINATIONS:
                reg_err("Diametro de Taladro", "CAT_PLT_NOMINACION_INVALIDA", f"Nominación de broca '{nominacion_raw}' no pertenece al catálogo (HQ, HQ3, NQ, PQ, BQ).")

            d_f = to_float(d_raw)
            if d_raw is None or clean_str(d_raw) == "":
                reg_err("D (mm)", "CAT_CAMPO_OBLIGATORIO_VACIO", "Diámetro D (mm) es obligatorio y se encuentra vacío.")
            elif d_f is None or d_f <= 0:
                reg_err("D (mm)", "CAT_PLT_DIAMETRO_RANGO", f"Diámetro D ({d_raw} mm) debe ser un valor positivo mayor a cero.")

            if long_m_f is not None and d_f is not None and d_f > 0:
                is_esbelto = long_m_f >= (0.5 * d_f)
                expected_verif_l = "OK" if is_esbelto else "CORTO"
                verif_l_str = clean_str(verif_long_raw).upper()
                if verif_long_raw is None or verif_l_str == "":
                    reg_err("Verif. de longitud", "CAT_CAMPO_OBLIGATORIO_VACIO", "El campo Verif. de longitud se encuentra vacío.")
                elif verif_l_str not in ("OK", "SI", "SÍ", "CORTO", "NO", "ERROR"):
                    reg_err("Verif. de longitud", "CAT_PLT_VERIF_LONGITUD_INCONGRUENTE",
                            f"Verificación de longitud '{verif_long_raw}' no coincide con el criterio ISRM '{expected_verif_l}'.")

            # --- 5. Litología y Factor K Canónico SSOT ---
            lito1_str = clean_str(lito1_raw)
            lito2_str = clean_str(lito2_raw)
            lito3_str = clean_str(lito3_raw)

            if not lito1_str:
                reg_err("Litologia 1", "CAT_CAMPO_OBLIGATORIO_VACIO", "Litología 1 es obligatoria y se encuentra vacía.")

            tipo_lito_str = strip_accents(clean_str(tipo_lito_raw).upper())
            if not tipo_lito_str:
                reg_err("Tipo litológico", "CAT_CAMPO_OBLIGATORIO_VACIO", "Tipo litológico es obligatorio y se encuentra vacío.")

            # Resolución canónica de Factor K y Grupo Geológico
            exp_grupo, exp_k = resolve_expected_k_and_type(lito1_str, lito2_str, lito3_str)

            if exp_grupo and tipo_lito_str:
                exp_grupo_clean = strip_accents(exp_grupo).upper()
                if not (exp_grupo_clean in tipo_lito_str or tipo_lito_str in exp_grupo_clean):
                    reg_err("Tipo litológico", "CAT_PLT_TIPO_LITOLOGICO_INCONGRUENTE",
                            f"Tipo litológico '{tipo_lito_raw}' no coincide con el grupo oficial '{exp_grupo}' para esta roca.",
                            sev="ADVERTENCIA")

            # --- 6. Ensayo Físico ---
            p_f = to_float(p_raw)
            if p_raw is None or clean_str(p_raw) == "":
                reg_err("P instr (kN)", "CAT_CAMPO_OBLIGATORIO_VACIO", "Fuerza P (kN) es obligatoria y se encuentra vacía.")
            elif p_f is None or p_f <= 0:
                reg_err("P instr (kN)", "CAT_PLT_FUERZA_P_RANGO", f"Fuerza P ({p_raw} kN) debe ser un valor positivo mayor a cero.")

            rotura_str = clean_str(tipo_rotura_raw).upper().replace(" ", "")
            if not rotura_str:
                reg_err("Tipo de Rotura", "CAT_CAMPO_OBLIGATORIO_VACIO", "Tipo de Rotura es obligatorio y se encuentra vacío.")
            elif rotura_str not in VALID_FRACTURE_TYPES:
                reg_err("Tipo de Rotura", "CAT_PLT_TIPO_ROTURA_INVALIDO", f"Tipo de rotura '{tipo_rotura_raw}' no es válido (debe ser M, E o C).")

            dir_str = clean_str(dir_rotura_raw).upper().replace(" ", "")
            if not dir_str:
                if rotura_str != "M":
                    reg_err("Dirección de rotura", "CAT_CAMPO_OBLIGATORIO_VACIO", "Dirección de rotura es obligatoria y se encuentra vacía.")
            elif dir_str not in VALID_ROTURA_DIRECTIONS:
                reg_err("Dirección de rotura", "CAT_PLT_DIRECCION_ROTURA_INVALIDA", f"Dirección de rotura '{dir_rotura_raw}' no es válida (debe ser Pa, Pe o NA).")

            if not clean_str(ejecutado_raw):
                reg_err("Ejecutado por", "CAT_CAMPO_OBLIGATORIO_VACIO", "Campo Ejecutado por es obligatorio y se encuentra vacío.")

            # --- 7. Cálculos Geomecánicos y Factor K ---
            is_f = to_float(is_raw)
            fact_f = to_float(fact_corr_raw)
            is50_f = to_float(is50_raw)
            k_f = to_float(factor_k_raw)
            ucs_f = to_float(ucs_raw)
            isrm_str = clean_str(isrm_raw).upper()

            # 1. Is (MPa) = (P * 1000) / D^2
            calc_is = None
            if is_raw is None or clean_str(is_raw) == "":
                reg_err("Is (Mpa)", "CAT_CAMPO_OBLIGATORIO_VACIO", "Índice Is (MPa) se encuentra vacío.")
            elif p_f is not None and d_f is not None and d_f > 0:
                calc_is = (p_f * 1000.0) / (d_f ** 2)
                if is_f is None or abs(is_f - calc_is) > self.tol_is:
                    reg_err("Is (Mpa)", "CAT_PLT_IS_INCONGRUENTE",
                            f"Is ({is_raw} MPa) difiere de la fórmula ({p_f}*1000 / {d_f}^2 = {calc_is:.2f} MPa).")

            # 2. Fact. Corr = (D / 50)^0.45
            calc_f = None
            if fact_corr_raw is None or clean_str(fact_corr_raw) == "":
                reg_err("Fact. Corr", "CAT_CAMPO_OBLIGATORIO_VACIO", "Factor de corrección se encuentra vacío.")
            elif d_f is not None and d_f > 0:
                calc_f = (d_f / 50.0) ** 0.45
                if fact_f is None or abs(fact_f - calc_f) > self.tol_f:
                    reg_err("Fact. Corr", "CAT_PLT_FACTOR_F_INCONGRUENTE",
                            f"Factor F ({fact_corr_raw}) difiere de ({d_f}/50)^0.45 = {calc_f:.3f}.")

            # 3. Is(50) = Is * Fact.Corr
            calc_is50 = None
            eff_is = is_f if is_f is not None else calc_is
            eff_f = fact_f if fact_f is not None else calc_f
            if is50_raw is None or clean_str(is50_raw) == "":
                reg_err("Is(50) (Mpa)", "CAT_CAMPO_OBLIGATORIO_VACIO", "Índice Is(50) (MPa) se encuentra vacío.")
            elif eff_is is not None and eff_f is not None:
                calc_is50 = eff_is * eff_f
                if is50_f is None or abs(is50_f - calc_is50) > self.tol_is:
                    reg_err("Is(50) (Mpa)", "CAT_PLT_IS50_INCONGRUENTE",
                            f"Is(50) ({is50_raw} MPa) difiere de ({eff_is:.2f} * {eff_f:.3f} = {calc_is50:.2f} MPa).")

            # 4. Factor K y Consistencia Litológica
            effective_k = k_f
            if factor_k_raw is None or clean_str(factor_k_raw) == "":
                reg_err("Factor K", "CAT_CAMPO_OBLIGATORIO_VACIO", "Factor K se encuentra vacío.")
                if exp_k is not None:
                    effective_k = exp_k  # Auto-adaptación para cálculo posterior
            elif k_f is None or not (5.0 <= k_f <= 30.0):
                reg_err("Factor K", "CAT_PLT_FACTOR_K_RANGO", f"Factor K ({factor_k_raw}) fuera de rango razonable [5.0, 30.0].")
            elif exp_k is not None and abs(k_f - exp_k) > 0.05:
                lito_comb = f"{lito1_str}/{lito2_str}/{lito3_str}".strip("/")
                reg_err("Factor K", "CAT_PLT_FACTOR_K_INCONGRUENTE",
                        f"Factor K ({k_f}) no coincide con la litología '{lito_comb}' (le corresponde K = {exp_k}).")

            # 5. UCS = Is(50) * K
            calc_ucs = None
            eff_is50 = is50_f if is50_f is not None else calc_is50
            if ucs_raw is None or clean_str(ucs_raw) == "":
                reg_err("UCS", "CAT_CAMPO_OBLIGATORIO_VACIO", "UCS se encuentra vacío.")
            elif eff_is50 is not None and effective_k is not None:
                calc_ucs = eff_is50 * effective_k
                if ucs_f is None or abs(ucs_f - calc_ucs) > self.tol_ucs:
                    reg_err("UCS", "CAT_PLT_UCS_INCONGRUENTE",
                            f"UCS ({ucs_raw} MPa) difiere de ({eff_is50:.2f} * {effective_k} = {calc_ucs:.2f} MPa).")

            # 6. ISRM Indice R
            expected_isrm = ""
            eff_ucs = ucs_f if ucs_f is not None else calc_ucs
            if not isrm_str:
                reg_err("ISRM Indice R", "CAT_CAMPO_OBLIGATORIO_VACIO", "ISRM Indice R se encuentra vacío.")
            elif eff_ucs is not None:
                expected_isrm = get_isrm_grade(eff_ucs).upper()
                if isrm_str != expected_isrm:
                    reg_err("ISRM Indice R", "CAT_PLT_RESISTENCIA_ISRM_INCONGRUENTE",
                            f"Índice ISRM '{isrm_raw}' no coincide con el rango de UCS={eff_ucs:.2f} MPa (esperado: '{expected_isrm}').")

            # --- 8. Cruce Geomecánico con Logueo General (LGG) (Opcional) ---
            if has_lgg and from_f is not None and to_f is not None:
                t_lookup = taladro_key.upper()
                if t_lookup not in lgg_by_dh:
                    reg_err("Taladro", "CAT_PLT_CORRIDA_NO_EXISTE_LGG",
                            f"El taladro '{t_lookup}' no existe en la base de datos de logueo general (LGG).",
                            sev="ALERTA")
                else:
                    dh_runs = lgg_by_dh[t_lookup]
                    # Buscar corrida de LGG que contenga el intervalo de la muestra
                    matched_runs = dh_runs[(dh_runs["desde_m"] <= from_f + 0.05) & (dh_runs["hasta_m"] >= to_f - 0.05)]

                    if matched_runs.empty:
                        # Evaluar si se sale de los límites de una corrida cercana
                        partial_runs = dh_runs[(dh_runs["desde_m"] <= to_f) & (dh_runs["hasta_m"] >= from_f)]
                        if not partial_runs.empty:
                            p_row = partial_runs.iloc[0]
                            reg_err("From", "CAT_PLT_MUESTRA_FUERA_DE_CORRIDA",
                                    f"El tramo de la muestra [{from_f:.2f} - {to_f:.2f} m] se sale físicamente de la corrida en LGG [{p_row['desde_m']:.2f} - {p_row['hasta_m']:.2f} m].",
                                    sev="ALERTA")
                        else:
                            reg_err("Corrida Desde (m)", "CAT_PLT_CORRIDA_NO_EXISTE_LGG",
                                    f"La corrida [{c_desde_f} - {c_hasta_f} m] o muestra [{from_f:.2f} - {to_f:.2f} m] no existe en el logueo oficial de LGG.",
                                    sev="ALERTA")
                    else:
                        m_lgg = matched_runs.iloc[0]
                        lgg_d = m_lgg["desde_m"]
                        lgg_h = m_lgg["hasta_m"]

                        # A) Límite de corrida no termina exactamente igual
                        if c_hasta_f is not None and abs(c_hasta_f - lgg_h) > 0.05:
                            reg_err("Corrida Hasta (m)", "CAT_PLT_CORRIDA_DIFIERE_LGG",
                                    f"La corrida anotada en PLT [{c_desde_f:.2f} - {c_hasta_f:.2f} m] no termina igual que en LGG [{lgg_d:.2f} - {lgg_h:.2f} m].",
                                    sev="ADVERTENCIA")

                        # B) Consistencia de Litología (Cascada Mutuamente Excluyente)
                        p_l1 = lito1_str.upper()
                        p_l2 = lito2_str.upper()
                        p_l3 = lito3_str.upper()

                        g_l1 = clean_str(m_lgg.get("lito1")).upper()
                        g_l2 = clean_str(m_lgg.get("lito2")).upper()
                        g_l3 = clean_str(m_lgg.get("lito3")).upper()

                        if p_l1 and g_l1:
                            # Prioridad 1: Macro discordancia
                            if p_l2 != g_l2 and p_l1 != g_l1:
                                reg_err("Litologia 1", "CAT_PLT_COMBINACION_LITO_DISCORDANTE_LGG",
                                        f"Roca distinta a logueo: PLT tiene '{p_l1}/{p_l2}/{p_l3}' pero LGG tiene '{g_l1}/{g_l2}/{g_l3}' para este tramo.",
                                        sev="ADVERTENCIA")
                            # Prioridad 2: Roca L2 distinta
                            elif p_l2 != g_l2:
                                reg_err("Litologia 2", "CAT_PLT_COMBINACION_LITO_DISCORDANTE_LGG",
                                        f"Roca L2 distinta a logueo: PLT tiene '{p_l1}/{p_l2}/{p_l3}' pero LGG tiene '{g_l1}/{g_l2}/{g_l3}' para este tramo.",
                                        sev="ADVERTENCIA")
                            # Prioridad 3: Detalle L3 distinto
                            elif p_l3 != g_l3 and p_l3 not in ("-", "NR", "VARIOS", "") and g_l3 not in ("-", "NR", "VARIOS", ""):
                                reg_err("Litologia 3", "CAT_PLT_COMBINACION_LITO_DISCORDANTE_LGG",
                                        f"Detalle L3 distinto a logueo: PLT tiene '{p_l1}/{p_l2}/{p_l3}' pero LGG tiene '{g_l1}/{g_l2}/{g_l3}' para este tramo.",
                                        sev="ADVERTENCIA")

                        # C) Dureza de campo vs ensayo de laboratorio
                        lgg_isrm = clean_str(m_lgg.get("isrm")).upper()
                        plt_isrm = expected_isrm or isrm_str
                        if lgg_isrm and plt_isrm and lgg_isrm in ("R0", "R1", "R2", "R3", "R4", "R5", "R6"):
                            if lgg_isrm != plt_isrm:
                                reg_err("ISRM Indice R", "CAT_PLT_DUREZA_DIFIERE_LGG",
                                        f"Dureza del ensayo ({plt_isrm}) difiere de la dureza estimada en campo en LGG ({lgg_isrm}).",
                                        sev="ADVERTENCIA")

        # =====================================================================
        # FASE 3: MÉTRICAS CONSOLIDADAS
        # =====================================================================
        invalid_count = sum(row_severity_flags)
        valid_count = total_rows - invalid_count
        quality_idx = (valid_count / total_rows * 100.0) if total_rows > 0 else 100.0

        return {
            "total_rows": total_rows,
            "valid_rows": valid_count,
            "invalid_rows": invalid_count,
            "quality_index": round(quality_idx, 2),
            "severity_counts": severity_counter,
            "category_counts": dict(category_counter),
            "column_error_counts": dict(column_error_counter),
            "drillhole_stats": dict(drillhole_stats),
            "campaign_stats": dict(campaign_stats),
            "anomalies": anomalies,
            "has_lgg_crosscheck": has_lgg
        }

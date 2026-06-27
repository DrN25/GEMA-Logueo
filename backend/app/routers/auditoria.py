import os
import io
import json
import shutil
import math
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
import time
from datetime import datetime
from collections import Counter, defaultdict
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, BackgroundTasks
from fastapi.responses import StreamingResponse, JSONResponse, FileResponse

from app.core.rules import MASTER_ERROR_RULES, get_rule_by_msg
from app.validator import validate_logueo_bulk_sheets, safe_float, safe_int, safe_str

router = APIRouter(prefix="/api", tags=["Auditoria"])
BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
uploads_dir = os.path.join(BASE_DIR, "uploads")
history_dir = os.path.join(uploads_dir, "history")
temp_dir = os.path.join(uploads_dir, "temp")

os.makedirs(history_dir, exist_ok=True)
os.makedirs(temp_dir, exist_ok=True)

# Helper to normalize/clean error messages for grouping
def simplify_message(msg):
    msg_clean = str(msg or "").strip()
    # Check matching rules in Master Error Rules catalog
    rule = get_rule_by_msg(msg_clean)
    if rule:
        return rule["msg"]
    # Fallback checking keywords
    msg_up = msg_clean.upper()
    if "CORRIDA PERFORADA" in msg_up or "LIMITE CRÍTICO" in msg_up:
        return "Longitud de corrida perforada excede el límite crítico de 1.6m."
    if "POSITIVA" in msg_up:
        return "Longitud de corrida perforada debe ser positiva."
    if "RECUPERADA" in msg_up and "AVANCE" in msg_up:
        return "Longitud recuperada es mayor que el avance perforado."
    if "RQD" in msg_up and "RECUPERADA" in msg_up:
        return "Metraje RQD es mayor que la longitud recuperada."
    if "SUMA DE FRAGMENTOS" in msg_up:
        return "La suma de fragmentos supera el avance perforado."
    if "BUZAMIENTO" in msg_up and "COINCIDE" in msg_up:
        return "La sumatoria de fracturas por buzamiento no coincide con el conteo general."
    if "ESPESOR DE RELLENO" in msg_up and "0MM" in msg_up:
        return "Se declaró espesor de relleno de junta pero la abertura es 0mm."
    if "ABERTURA" in msg_up and "REGISTRADO" in msg_up:
        return "La abertura de junta es mayor a 0mm pero no se ha registrado espesor de relleno."
    if "INCOMPATIBILIDAD GEOLÓGICA" in msg_up:
        if "PARED" in msg_up:
            return "Incompatibilidad geológica (Dureza de pared de junta supera la resistencia intacta de la corrida)."
        return "Incompatibilidad geológica (Resistencia vs Intemperismo de corrida)."
    if "HUÉRFANA" in msg_up or "HUERFANA" in msg_up:
        return "Profundidad huérfana de junta no corresponde a ningún tramo de corrida en LGG."
    if "ALFA" in msg_up:
        return "El ángulo Alfa es inválido. Debe estar entre 0° y 90° o ser -1."
    if "BETA" in msg_up:
        return "El ángulo Beta es inválido. Debe estar entre 0° y 360° o ser -1."
    if "JRC10" in msg_up:
        return "El valor de JRC10 es inválido. No se permiten valores mayores a 20."
    if "ESPESOR DE RELLENO" in msg_up and "MAYOR" in msg_up:
        return "El espesor de relleno no puede ser mayor que la abertura de junta."
    if "DEFINIR" in msg_up or "RELLENO ESTÁ SIN" in msg_up:
        return "Se declaró espesor de relleno pero el tipo de relleno está sin definir."
    if "DEFINIDO" in msg_up and "0MM" in msg_up:
        return "El tipo de relleno está definido pero la abertura de junta es 0mm."
    if "LITOLOGÍA" in msg_up or "LITOLOGIA" in msg_up:
        return "Incompatibilidad de litología entre la corrida y la junta."
    if "VACÍO" in msg_up or "VACIO" in msg_up:
        return "Campo obligatorio se encuentra vacío."
    return msg_clean

def get_safe_sheet_name(title, index):
    clean_title = "".join(c for c in title if c not in r':\/?*[]\'"').strip()
    suffix = f" ({index})"
    max_title_len = 31 - len(suffix)
    return f"{clean_title[:max_title_len].strip()}{suffix}"

def safe_replace(src: str, dst: str, retries: int = 5, delay: float = 0.2):
    for i in range(retries):
        try:
            os.replace(src, dst)
            return
        except (PermissionError, OSError) as e:
            if i == retries - 1:
                try:
                    shutil.copyfile(src, dst)
                    try: os.remove(src)
                    except: pass
                    return
                except:
                    raise e
            time.sleep(delay)

def generar_excel_reporte_core(diag: dict, compact: dict, filtered: list):
    font_title = Font(name="Segoe UI", size=16, bold=True, color="1B365D")
    font_subtitle = Font(name="Segoe UI", size=10, italic=True, color="555555")
    font_section = Font(name="Segoe UI", size=11, bold=True, color="1B365D")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Segoe UI", size=10, bold=True, color="000000")
    font_regular = Font(name="Segoe UI", size=10, color="000000")
    font_kpi_lbl = Font(name="Segoe UI", size=9, bold=True, color="555555")
    
    font_kpi_val_blue = Font(name="Segoe UI", size=18, bold=True, color="1B365D")
    font_kpi_val_green = Font(name="Segoe UI", size=18, bold=True, color="375623")
    font_kpi_val_red = Font(name="Segoe UI", size=18, bold=True, color="C00000")
    font_kpi_val_orange = Font(name="Segoe UI", size=18, bold=True, color="C65911")

    fill_primary = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    fill_accent_green = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fill_accent_yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fill_accent_orange = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    fill_accent_red = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
    fill_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    fill_kpi_gray = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")

    border_thin = Border(
        left=Side(style='thin', color='E2E8F0'), 
        right=Side(style='thin', color='E2E8F0'), 
        top=Side(style='thin', color='E2E8F0'), 
        bottom=Side(style='thin', color='E2E8F0')
    )
    border_kpi = Border(
        left=Side(style='thin', color='B0C4DE'),
        right=Side(style='thin', color='B0C4DE'),
        top=Side(style='thin', color='B0C4DE'),
        bottom=Side(style='thin', color='B0C4DE')
    )

    alignment_center = Alignment(horizontal="center", vertical="center")
    alignment_left = Alignment(horizontal="left", vertical="center")
    alignment_right = Alignment(horizontal="right", vertical="center")

    wb = openpyxl.Workbook()
    
    def write_kpi_card_opt(ws, start_row, start_col, label, value, bg_fill, val_font):
        c1 = ws.cell(row=start_row, column=start_col, value=label)
        c1.font = font_kpi_lbl
        c1.alignment = alignment_center
        
        c2 = ws.cell(row=start_row+1, column=start_col, value=value)
        c2.font = val_font
        c2.alignment = alignment_center
        
        for r in range(start_row, start_row+2):
            for c in range(start_col, start_col+2):
                cell = ws.cell(row=r, column=c)
                cell.fill = bg_fill
                cell.border = border_kpi
                
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col+1)
        ws.merge_cells(start_row=start_row+1, start_column=start_col, end_row=start_row+1, end_column=start_col+1)

    # --- HOJA 1: DASHBOARD EJECUTIVO ---
    ws_dash = wb.active
    ws_dash.title = "📊 Dashboard Ejecutivo"
    ws_dash.views.sheetView[0].showGridLines = True
    
    ws_dash.cell(row=2, column=2, value="SISTEMA DE AUDITORÍA DE LOGUEO").font = font_title
    ws_dash.cell(row=3, column=2, value="Dashboard de Control de Calidad y Consistencia de Logueo General y Estructural").font = font_subtitle
    
    total_filas = compact.get("familia1", {}).get("total_discontinuidades", 0)
    total_fields = compact.get("familia2", {}).get("total_fields", 0)
    total_vacios = sum(1 for i in filtered if i.get("tipo_incidencia") == "VACIO")
    total_advertencias = sum(1 for i in filtered if i.get("tipo_incidencia") == "ADVERTENCIA")
    total_alertas = sum(1 for i in filtered if i.get("tipo_incidencia") == "ALERTA")
    total_correctos = total_fields - (total_vacios + total_advertencias + total_alertas)
    pct_integridad = (total_correctos / max(1, total_fields)) * 100

    # KPI Cards
    write_kpi_card_opt(ws_dash, 5, 2, "TALADROS EVALUADOS", len(compact.get("resumen_por_celda_padre", {})), fill_kpi_gray, font_kpi_val_blue)
    write_kpi_card_opt(ws_dash, 5, 4, "FILAS DE LOGUEO EVALUADAS", total_filas, fill_kpi_gray, font_kpi_val_blue)
    write_kpi_card_opt(ws_dash, 5, 6, "INTEGRIDAD GLOBAL DE CAMPOS", f"{pct_integridad:.2f}%", fill_accent_green, font_kpi_val_green)
    write_kpi_card_opt(ws_dash, 5, 8, "ALERTAS CRÍTICAS", total_alertas, fill_accent_red, font_kpi_val_red)
    write_kpi_card_opt(ws_dash, 5, 10, "ADVERTENCIAS DE CONSISTENCIA", total_advertencias, fill_accent_orange, font_kpi_val_orange)

    # Tabla: Distribución por Campaña
    ws_dash.cell(row=9, column=2, value="DESEMPEÑO DE CONTROL POR CAMPAÑA").font = font_section
    headers_camp = ["Campaña", "Estructuras/Corridas", "Alertas (N)", "% Alertas", "Vacíos (N)", "% Vacíos"]
    for idx, col in enumerate(headers_camp, start=2):
        cell = ws_dash.cell(row=10, column=idx, value=col)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = alignment_center
        cell.border = border_thin

    r_camp = 11
    for row in compact.get("distribucion_campania", []):
        ws_dash.cell(row=r_camp, column=2, value=row.get("campania")).font = font_bold
        ws_dash.cell(row=r_camp, column=2).alignment = alignment_center
        
        ws_dash.cell(row=r_camp, column=3, value=safe_int(row.get("discontinuidades"))).number_format = '#,##0'
        ws_dash.cell(row=r_camp, column=3).alignment = alignment_right
        
        ws_dash.cell(row=r_camp, column=4, value=safe_int(row.get("alertas_cant"))).number_format = '#,##0'
        ws_dash.cell(row=r_camp, column=4).alignment = alignment_right
        
        ws_dash.cell(row=r_camp, column=5, value=safe_float(row.get("alertas_pct")) / 100.0).number_format = '0.00%'
        ws_dash.cell(row=r_camp, column=5).alignment = alignment_right
        
        ws_dash.cell(row=r_camp, column=6, value=safe_int(row.get("vacios_cant"))).number_format = '#,##0'
        ws_dash.cell(row=r_camp, column=6).alignment = alignment_right
        
        ws_dash.cell(row=r_camp, column=7, value=safe_float(row.get("vacios_pct")) / 100.0).number_format = '0.00%'
        ws_dash.cell(row=r_camp, column=7).alignment = alignment_right
        
        for col_idx in range(2, 8):
            ws_dash.cell(row=r_camp, column=col_idx).border = border_thin
            if r_camp % 2 == 0:
                ws_dash.cell(row=r_camp, column=col_idx).fill = fill_zebra
        r_camp += 1

    # Tabla: Distribución por Geólogos
    r_sect = r_camp + 2
    ws_dash.cell(row=r_sect, column=2, value="DESEMPEÑO DE CONTROL POR GEÓLOGO").font = font_section
    
    r_sect += 1
    headers_sect = ["Geólogo", "Registros", "Alertas (N)", "% Alertas", "Vacíos (N)", "% Vacíos"]
    for idx, col in enumerate(headers_sect, start=2):
        cell = ws_dash.cell(row=r_sect, column=idx, value=col)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = alignment_center
        cell.border = border_thin

    for row in compact.get("distribucion_geotecnico", []):
        r_sect += 1
        ws_dash.cell(row=r_sect, column=2, value=row.get("geotecnico")).font = font_bold
        ws_dash.cell(row=r_sect, column=2).alignment = alignment_center
        
        ws_dash.cell(row=r_sect, column=3, value=safe_int(row.get("discontinuidades"))).number_format = '#,##0'
        ws_dash.cell(row=r_sect, column=3).alignment = alignment_right
        
        ws_dash.cell(row=r_sect, column=4, value=safe_int(row.get("alertas_cant"))).number_format = '#,##0'
        ws_dash.cell(row=r_sect, column=4).alignment = alignment_right
        
        ws_dash.cell(row=r_sect, column=5, value=safe_float(row.get("alertas_pct")) / 100.0).number_format = '0.00%'
        ws_dash.cell(row=r_sect, column=5).alignment = alignment_right
        
        ws_dash.cell(row=r_sect, column=6, value=safe_int(row.get("vacios_cant"))).number_format = '#,##0'
        ws_dash.cell(row=r_sect, column=6).alignment = alignment_right
        
        ws_dash.cell(row=r_sect, column=7, value=safe_float(row.get("vacios_pct")) / 100.0).number_format = '0.00%'
        ws_dash.cell(row=r_sect, column=7).alignment = alignment_right
        
        for col_idx in range(2, 8):
            ws_dash.cell(row=r_sect, column=col_idx).border = border_thin
            if r_sect % 2 == 0:
                ws_dash.cell(row=r_sect, column=col_idx).fill = fill_zebra
        r_sect += 1

    # Tabla: Taladros más Afectados
    r_worst = r_sect + 2
    ws_dash.cell(row=r_worst-1, column=2, value="PEORES 5 TALADROS CON MAYOR DESVIACIÓN").font = font_section
    headers_worst = ["Taladro", "Registros", "Vacíos", "Advertencias", "Alertas", "Calificación"]
    for idx, col in enumerate(headers_worst, start=2):
        cell = ws_dash.cell(row=r_worst, column=idx, value=col)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = alignment_center
        cell.border = border_thin

    for row in compact.get("worst_cells", [])[:5]:
        r_worst += 1
        ws_dash.cell(row=r_worst, column=2, value=row.get("celda")).font = font_bold
        ws_dash.cell(row=r_worst, column=2).alignment = alignment_center
        
        ws_dash.cell(row=r_worst, column=3, value=safe_int(row.get("total_hijas"))).number_format = '#,##0'
        ws_dash.cell(row=r_worst, column=3).alignment = alignment_right
        
        ws_dash.cell(row=r_worst, column=4, value=safe_int(row.get("vacios"))).number_format = '#,##0'
        ws_dash.cell(row=r_worst, column=4).alignment = alignment_right
        
        ws_dash.cell(row=r_worst, column=5, value=safe_int(row.get("advertencias"))).number_format = '#,##0'
        ws_dash.cell(row=r_worst, column=5).alignment = alignment_right
        
        ws_dash.cell(row=r_worst, column=6, value=safe_int(row.get("alertas"))).number_format = '#,##0'
        ws_dash.cell(row=r_worst, column=6).alignment = alignment_right
        
        status = row.get("estado_celda", "OK")
        status_cell = ws_dash.cell(row=r_worst, column=7, value=status)
        status_cell.font = font_bold
        status_cell.alignment = alignment_center
        if status == "ALERTA": status_cell.fill = fill_accent_red
        elif status == "ADVERTENCIA": status_cell.fill = fill_accent_orange
        else: status_cell.fill = fill_accent_green
        
        for col_idx in range(2, 8):
            ws_dash.cell(row=r_worst, column=col_idx).border = border_thin
            if r_worst % 2 == 0:
                ws_dash.cell(row=r_worst, column=col_idx).fill = fill_zebra

    # Tabla para Gráfica Directa: Top 5 Alertas Críticas
    ws_dash.cell(row=9, column=9, value="PRINCIPALES ALERTAS CRÍTICAS").font = font_section
    headers_graph = ["Anomalía Geotécnica", "Frecuencia"]
    for idx, col in enumerate(headers_graph, start=9):
        cell = ws_dash.cell(row=10, column=idx, value=col)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = alignment_center
        cell.border = border_thin

    top_errs_list = Counter(simplify_message(i.get("mensaje")) for i in filtered if i.get("tipo_incidencia") == "ALERTA").most_common(5)
    r_graph = 11
    for msg, qty in top_errs_list:
        ws_dash.cell(row=r_graph, column=9, value=msg).font = font_regular
        ws_dash.cell(row=r_graph, column=9).border = border_thin
        
        c_qty = ws_dash.cell(row=r_graph, column=10, value=qty)
        c_qty.font = font_bold
        c_qty.alignment = alignment_right
        c_qty.number_format = '#,##0'
        c_qty.border = border_thin
        c_qty.fill = fill_accent_red
        r_graph += 1
        
    for dummy in range(r_graph, 16):
        ws_dash.cell(row=dummy, column=9, value="—").font = font_regular
        ws_dash.cell(row=dummy, column=9).border = border_thin
        ws_dash.cell(row=dummy, column=10, value=0).font = font_regular
        ws_dash.cell(row=dummy, column=10).border = border_thin
        ws_dash.cell(row=dummy, column=10).number_format = '#,##0'

    # Gráfica Nativa de Excel
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Distribución de Anomalías Críticas"
    chart.y_axis.title = "Frecuencia"
    chart.x_axis.title = "Regla de Consistencia"
    
    chart_data = Reference(ws_dash, min_col=10, min_row=10, max_row=15)
    chart_cats = Reference(ws_dash, min_col=9, min_row=11, max_row=15)
    chart.add_data(chart_data, titles_from_data=True)
    chart.set_categories(chart_cats)
    chart.legend = None
    chart.width = 15
    chart.height = 11
    ws_dash.add_chart(chart, "I17")

    # --- HOJA 2: REGISTRO MAESTRO DE ERRORES (CATÁLOGO / ÍNDICE) ---
    ws_cat = wb.create_sheet(title="❌ Catálogo de Errores")
    ws_cat.views.sheetView[0].showGridLines = True
    
    ws_cat.cell(row=2, column=2, value="CATÁLOGO DE REGLAS DE CONSISTENCIA").font = font_title
    ws_cat.cell(row=3, column=2, value="Índice maestro de validación geomecánica ordenado por frecuencia. Use los hipervínculos para navegar.").font = font_subtitle
    
    headers_cat = ["ID", "Gravedad", "Mensaje de Regla Evaluada", "Casos Hallados (N)", "Enlace Detallado"]
    for idx, col in enumerate(headers_cat, start=2):
        cell = ws_cat.cell(row=5, column=idx, value=col)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = alignment_center
        cell.border = border_thin

    incidencias_por_error = defaultdict(list)
    for inc in filtered:
        msg_simplificado = simplify_message(inc.get("mensaje", ""))
        incidencias_por_error[msg_simplificado].append(inc)

    catalog_frequencies = []
    for rule in MASTER_ERROR_RULES:
        rule_msg = rule["msg"]
        matches = incidencias_por_error[rule_msg]
        catalog_frequencies.append({
            "msg": rule_msg, "severity": rule["severity"], "matches": matches, "count": len(matches)
        })
        
    catalog_frequencies = sorted(catalog_frequencies, key=lambda x: x["count"], reverse=True)

    r_cat = 6
    active_sheets_mapping = {}
    
    for c_idx, rule in enumerate(catalog_frequencies, start=1):
        ws_cat.cell(row=r_cat, column=2, value=c_idx).font = font_regular
        ws_cat.cell(row=r_cat, column=2).alignment = alignment_center
        ws_cat.cell(row=r_cat, column=2).border = border_thin
        
        c_sev = ws_cat.cell(row=r_cat, column=3, value=rule["severity"])
        c_sev.font = font_bold
        c_sev.alignment = alignment_center
        c_sev.border = border_thin
        if rule["severity"] == "ALERTA": c_sev.fill = fill_accent_red
        elif rule["severity"] == "ADVERTENCIA": c_sev.fill = fill_accent_orange
        else: c_sev.fill = fill_accent_yellow
        
        ws_cat.cell(row=r_cat, column=4, value=rule["msg"]).font = font_bold if rule["count"] > 0 else font_regular
        ws_cat.cell(row=r_cat, column=4).border = border_thin
        
        c_count = ws_cat.cell(row=r_cat, column=5, value=rule["count"])
        c_count.font = font_bold
        c_count.alignment = alignment_right
        c_count.number_format = '#,##0'
        c_count.border = border_thin
        
        c_link = ws_cat.cell(row=r_cat, column=6)
        if rule["count"] > 0:
            tab_name = get_safe_sheet_name(rule["msg"], c_idx)
            active_sheets_mapping[rule["msg"]] = {"tab_name": tab_name, "records": rule["matches"]}
            
            c_link.value = f'=HYPERLINK("#\'{tab_name}\'!B2", "🔍 Navegar a Detalles")'
            c_link.font = Font(name="Segoe UI", size=10, bold=True, color="1B365D", underline="single")
            c_link.alignment = alignment_center
        else:
            c_link.value = "Limpio / 0 Incidencias"
            c_link.font = Font(name="Segoe UI", size=9, italic=True, color="7F8C8D")
            c_link.alignment = alignment_center
            c_link.fill = fill_accent_green
            
        c_link.border = border_thin
        r_cat += 1

    # --- HOJA 3: DETALLE COMPLETO DE INCIDENCIAS ---
    ws_detail = wb.create_sheet(title="📋 Detalle de Incidencias")
    ws_detail.views.sheetView[0].showGridLines = True
    
    ws_detail.cell(row=2, column=2, value="REGISTRO COMPLETO DE INCIDENCIAS").font = font_title
    ws_detail.cell(row=3, column=2, value="Listado plano consolidado de todas las desviaciones y vacíos detectados. Utilice filtros en las cabeceras.").font = font_subtitle
    
    headers_detail = [
        "Fila Excel", "Módulo", "Gravedad", "Taladro Padre", "ID/Prof Hija", "Campaña", 
        "Logger Geotécnico", "Columna de Falla", "Valor Actual", "Mensaje de Inconsistencia"
    ]
    
    ws_detail.append([]) 
    ws_detail.append([None] + headers_detail) 
    grid_heading_row = ws_detail.max_row
    
    for idx in range(2, 12):
        cell = ws_detail.cell(row=grid_heading_row, column=idx)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = alignment_center
        cell.border = border_thin
        
    start_detail_row = ws_detail.max_row + 1
    for inc_item in filtered:
        row_data = [
            None,
            safe_int(inc_item.get("fila_excel")),
            inc_item.get("modulo", "LGG"),
            inc_item.get("tipo_incidencia", "ALERTA"),
            inc_item.get("celda_padre"),
            inc_item.get("celda_hija"),
            inc_item.get("campania"),
            inc_item.get("geotecnico"),
            inc_item.get("columna"),
            inc_item.get("valor_actual") if inc_item.get("valor_actual") is not None else "—",
            simplify_message(inc_item.get("mensaje"))
        ]
        ws_detail.append(row_data)
        
    end_detail_row = ws_detail.max_row
    
    for r_idx in range(start_detail_row, end_detail_row + 1):
        if r_idx <= start_detail_row + 300:
            ws_detail.cell(row=r_idx, column=2).alignment = alignment_center
            ws_detail.cell(row=r_idx, column=3).alignment = alignment_center
            ws_detail.cell(row=r_idx, column=4).alignment = alignment_center
            ws_detail.cell(row=r_idx, column=5).alignment = alignment_center
            ws_detail.cell(row=r_idx, column=6).alignment = alignment_center
            ws_detail.cell(row=r_idx, column=7).alignment = alignment_center
            ws_detail.cell(row=r_idx, column=8).alignment = alignment_left
            ws_detail.cell(row=r_idx, column=9).alignment = alignment_center
            ws_detail.cell(row=r_idx, column=10).alignment = alignment_center
            ws_detail.cell(row=r_idx, column=11).alignment = alignment_left
            
            if r_idx % 2 == 0:
                for col_idx in range(2, 12):
                    if col_idx != 4:
                        ws_detail.cell(row=r_idx, column=col_idx).fill = fill_zebra
        else:
            ws_detail.cell(row=r_idx, column=2).alignment = alignment_center
            ws_detail.cell(row=r_idx, column=4).alignment = alignment_center
            
        cell_sev = ws_detail.cell(row=r_idx, column=4)
        sev = cell_sev.value
        if sev == "ALERTA": cell_sev.fill = fill_accent_red
        elif sev == "ADVERTENCIA": cell_sev.fill = fill_accent_orange
        else: cell_sev.fill = fill_accent_yellow
        cell_sev.font = font_bold
        
        for col_idx in range(2, 12):
            ws_detail.cell(row=r_idx, column=col_idx).border = border_thin
            
    ws_detail.auto_filter.ref = f"B{grid_heading_row}:K{end_detail_row}"

    # --- HOJAS 4+: DETALLES ESPECÍFICOS POR REGLA ---
    for orig_msg, mapping_data in active_sheets_mapping.items():
        sh_name = mapping_data["tab_name"]
        err_records = mapping_data["records"]
        
        ws_err = wb.create_sheet(title=sh_name)
        ws_err.views.sheetView[0].showGridLines = True
        
        c_back = ws_err.cell(row=2, column=2)
        c_back.value = '=HYPERLINK("#\'❌ Catálogo de Errores\'!B2", "⬅ Volver al Catálogo de Errores")'
        c_back.font = Font(name="Segoe UI", size=10, bold=True, color="1B365D", underline="single")
        c_back.alignment = alignment_left
        
        ws_err.cell(row=4, column=2, value="ANÁLISIS DE ANOMALÍA ESPECÍFICA").font = font_section
        cell_err_desc = ws_err.cell(row=5, column=2, value=f"Regla: {orig_msg.upper()}")
        cell_err_desc.font = Font(name="Segoe UI", size=10, bold=True, color="7F1D1D")
        cell_err_desc.fill = fill_accent_red
        cell_err_desc.border = border_thin
        ws_err.merge_cells(start_row=5, start_column=2, end_row=5, end_column=7)
        
        st_affected = len(set(x.get("celda_padre", "N/A") for x in err_records))
        tot_affected = len(err_records)
        
        write_kpi_card_opt(ws_err, 7, 2, "TALADROS AFECTADOS", st_affected, fill_kpi_gray, font_kpi_val_blue)
        write_kpi_card_opt(ws_err, 7, 4, "REGISTROS AFECTADOS", tot_affected, fill_kpi_gray, font_kpi_val_blue)
        
        # Distribución por Campaña
        ws_err.cell(row=10, column=2, value="DISTRIBUCIÓN POR CAMPAÑA").font = font_section
        for idx, col in enumerate(["Campaña / Año", "Ocurrencias", "% Contribución"], start=2):
            cell = ws_err.cell(row=11, column=idx, value=col)
            cell.font = font_header
            cell.fill = fill_primary
            cell.alignment = alignment_center
            cell.border = border_thin
            
        r_dist_yr = defaultdict(int)
        for r in err_records:
            r_dist_yr[str(r.get("campania", "N/A"))] += 1
            
        curr_y_r = 12
        for yr, y_qty in sorted(r_dist_yr.items()):
            ws_err.cell(row=curr_y_r, column=2, value=yr).font = font_bold
            ws_err.cell(row=curr_y_r, column=2).alignment = alignment_center
            ws_err.cell(row=curr_y_r, column=2).border = border_thin
            
            c_yq = ws_err.cell(row=curr_y_r, column=3, value=y_qty)
            c_yq.font = font_regular
            c_yq.alignment = alignment_right
            c_yq.number_format = '#,##0'
            c_yq.border = border_thin
            
            c_yp = ws_err.cell(row=curr_y_r, column=4, value=y_qty / max(1, tot_affected))
            c_yp.font = font_regular
            c_yp.alignment = alignment_right
            c_yp.number_format = '0.00%'
            c_yp.border = border_thin
            curr_y_r += 1
            
        # Distribución por Geólogo
        ws_err.cell(row=10, column=6, value="DISTRIBUCIÓN POR REGISTRADOR (GEÓLOGO)").font = font_section
        for idx, col in enumerate(["Logger Geotécnico", "Ocurrencias", "% Contribución"], start=6):
            cell = ws_err.cell(row=11, column=idx, value=col)
            cell.font = font_header
            cell.fill = fill_primary
            cell.alignment = alignment_center
            cell.border = border_thin
            
        r_dist_sc = defaultdict(int)
        for r in err_records:
            r_dist_sc[str(r.get("geotecnico", "N/A"))] += 1
            
        curr_s_r = 12
        for sc, s_qty in sorted(r_dist_sc.items()):
            ws_err.cell(row=curr_s_r, column=6, value=sc).font = font_bold
            ws_err.cell(row=curr_s_r, column=6).alignment = alignment_center
            ws_err.cell(row=curr_s_r, column=6).border = border_thin
            
            c_sq = ws_err.cell(row=curr_s_r, column=7, value=s_qty)
            c_sq.font = font_regular
            c_sq.alignment = alignment_right
            c_sq.number_format = '#,##0'
            c_sq.border = border_thin
            
            c_sp = ws_err.cell(row=curr_s_r, column=8, value=s_qty / max(1, tot_affected))
            c_sp.font = font_regular
            c_sp.alignment = alignment_right
            c_sp.number_format = '0.00%'
            c_sp.border = border_thin
            curr_s_r += 1

        # Detalle de Registros
        ws_err.append([])
        ws_err.append([None, "REGISTROS INDIVIDUALES AFECTADOS"])
        ws_err.cell(row=ws_err.max_row, column=2).font = font_section
        
        headers_inc = [
            "Fila Excel", "Módulo", "Taladro Padre", "ID/Prof Hija", "Campaña", 
            "Logger Geotécnico", "Columna de Falla", "Valor Actual", "Mensaje de Regla"
        ]
        ws_err.append([None] + headers_inc)
        header_row_idx = ws_err.max_row
        
        for col_idx in range(2, 11):
            cell = ws_err.cell(row=header_row_idx, column=col_idx)
            cell.font = font_header
            cell.fill = fill_primary
            cell.alignment = alignment_center
            cell.border = border_thin
            
        start_data_row = ws_err.max_row + 1
        for inc_item in err_records:
            row_data = [
                None,
                safe_int(inc_item.get("fila_excel")),
                inc_item.get("modulo", "LGG"),
                inc_item.get("celda_padre"),
                inc_item.get("celda_hija"),
                inc_item.get("campania"),
                inc_item.get("geotecnico"),
                inc_item.get("columna"),
                inc_item.get("valor_actual") if inc_item.get("valor_actual") is not None else "—",
                inc_item.get("mensaje")
            ]
            ws_err.append(row_data)
            
        end_data_row = ws_err.max_row
        
        for r_idx in range(start_data_row, end_data_row + 1):
            if r_idx <= start_data_row + 150:
                ws_err.cell(row=r_idx, column=2).alignment = alignment_center
                ws_err.cell(row=r_idx, column=3).alignment = alignment_center
                ws_err.cell(row=r_idx, column=4).alignment = alignment_center
                ws_err.cell(row=r_idx, column=5).alignment = alignment_center
                ws_err.cell(row=r_idx, column=6).alignment = alignment_center
                ws_err.cell(row=r_idx, column=7).alignment = alignment_center
                ws_err.cell(row=r_idx, column=8).alignment = alignment_left
                ws_err.cell(row=r_idx, column=9).alignment = alignment_center
                ws_err.cell(row=r_idx, column=10).alignment = alignment_left
                
                if r_idx % 2 == 0:
                    for col_idx in range(2, 11):
                        ws_err.cell(row=r_idx, column=col_idx).fill = fill_zebra
            else:
                ws_err.cell(row=r_idx, column=2).alignment = alignment_center
                ws_err.cell(row=r_idx, column=4).alignment = alignment_center
                
            for col_idx in range(2, 11):
                ws_err.cell(row=r_idx, column=col_idx).border = border_thin
                
        ws_err.auto_filter.ref = f"B{header_row_idx}:J{end_data_row}"

    # Auto-ajuste de Columnas
    for ws in wb.worksheets:
        ws.column_dimensions['A'].width = 3
        for col_idx in range(2, ws.max_column + 1):
            vals = []
            for row_idx in range(1, min(15, ws.max_row + 1)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    val_str = str(val)
                    if val_str.startswith("=HYPERLINK"):
                        vals.append("Navegar")
                    else:
                        vals.append(val_str)
            if not vals: continue
            max_len = max(len(v) for v in vals)
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 11), 52)

    return wb

def run_logueo_audit_pipeline(file_path: str, lgg_sheet: str, est_sheet: str, audit_id: str):
    raw_json_out = os.path.join(history_dir, f"{audit_id}_diagnostico.json")
    compact_json_out = os.path.join(history_dir, f"{audit_id}_compact.json")
    excel_pregenerated_out = os.path.join(history_dir, f"{audit_id}_reporte_completo.xlsx")
    
    start_time = time.time()
    t_str = lambda: datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    print(f"[*] [{t_str()}] Inicio de validación geotécnica física y cruzada para el reporte {audit_id}")
    print(f"[*] [{t_str()}] Hojas de trabajo: LGG='{lgg_sheet}', Estructural='{est_sheet}'")
    
    # 1. Ejecutar validación y guardar diagnóstico
    validate_logueo_bulk_sheets(file_path, lgg_sheet, est_sheet, raw_json_out)
    
    elapsed_val = round(time.time() - start_time, 2)
    print(f"[+] [{t_str()}] Finalización de validación y guardado de JSON diagnóstico en ({elapsed_val}s)")
    
    # Copiar diagnóstico público
    shutil.copyfile(raw_json_out, os.path.join(uploads_dir, "diagnostico_geomecanico.json"))
    
    print(f"[*] [{t_str()}] Inicio de compilación de KPIs y compactado del reporte para {audit_id}")
    # 2. Generar archivo compacto resumen
    with open(raw_json_out, "r", encoding="utf-8") as f:
        diag = json.load(f)
        
    compact = {k: v for k, v in diag.items() if k != "incidencias"}
    incidencias = diag.get("incidencias", [])
    total_filas = diag.get("total_filas_procesadas", 0)
    
    resumen_celdas = diag.get("resumen_por_celda_padre", {})
    num_celdas_padre = len(resumen_celdas)
    promedio_hijas = sum(x["total_hijas"] for x in resumen_celdas.values()) / max(1, num_celdas_padre)
    total_metros = sum(safe_float(x.get("dist_celda", 0.0)) for x in resumen_celdas.values())
    
    total_fields = total_filas * 20  # estimado
    total_vacios = sum(1 for i in incidencias if i.get("tipo_incidencia") == "VACIO")
    total_advertencias = sum(1 for i in incidencias if i.get("tipo_incidencia") == "ADVERTENCIA")
    total_alertas = sum(1 for i in incidencias if i.get("tipo_incidencia") == "ALERTA")
    total_correctos = total_fields - (total_vacios + total_advertencias + total_alertas)
    
    row_errors = defaultdict(set)
    for i in incidencias:
        row_errors[f"{i['modulo']}_{i['fila_excel']}"].add(i["tipo_incidencia"])
        
    discs_con_alerta = sum(1 for row, errs in row_errors.items() if "ALERTA" in errs)
    discs_con_advertencia = sum(1 for row, errs in row_errors.items() if "ADVERTENCIA" in errs and "ALERTA" not in errs)
    discs_con_vacio = sum(1 for row, errs in row_errors.items() if "VACIO" in errs)
    discs_correctas = total_filas - len(row_errors)
    
    camp_stats = defaultdict(lambda: {"vacios": 0, "advertencias": 0, "alertas": 0, "filas": set()})
    geo_stats = defaultdict(lambda: {"vacios": 0, "advertencias": 0, "alertas": 0, "filas": set()})
    sector_stats = defaultdict(lambda: {"vacios": 0, "advertencias": 0, "alertas": 0, "filas": set()})
    
    observaciones_por_año = defaultdict(lambda: defaultdict(lambda: {"incidents": 0, "stations": set()}))
    top_stations_por_año = defaultdict(lambda: defaultdict(lambda: Counter()))
    
    for i in incidencias:
        c = i.get("campania", "N/A")
        if c == "N/A": continue
        obs_key = simplify_message(i.get("mensaje", ""))
        celda = i.get("celda_padre", "N/A")
        
        observaciones_por_año[c][obs_key]["incidents"] += 1
        observaciones_por_año[c][obs_key]["stations"].add(celda)
        top_stations_por_año[c][obs_key][celda] += 1
        
        camp_stats[c]["filas"].add(f"{i['modulo']}_{i['fila_excel']}")
        geo_stats[g := i.get("geotecnico", "N/A")]["filas"].add(f"{i['modulo']}_{i['fila_excel']}")
        sector_stats[s := i.get("sector_geotecnico", "N/A")]["filas"].add(f"{i['modulo']}_{i['fila_excel']}")
        
        tipo = i.get("tipo_incidencia")
        if tipo == "VACIO":
            camp_stats[c]["vacios"] += 1
            geo_stats[g]["vacios"] += 1
            sector_stats[s]["vacios"] += 1
        elif tipo == "ADVERTENCIA":
            camp_stats[c]["advertencias"] += 1
            geo_stats[g]["advertencias"] += 1
            sector_stats[s]["advertencias"] += 1
        elif tipo == "ALERTA":
            camp_stats[c]["alertas"] += 1
            geo_stats[g]["alertas"] += 1
            sector_stats[s]["alertas"] += 1
            
    consolidado_tabla = {}
    for year, types in observaciones_por_año.items():
        consolidado_tabla[year] = {}
        total_inc_año = sum(v["incidents"] for k, v in types.items())
        severity = "LEVE" if total_inc_año < 50 else ("MODERADO" if total_inc_año < 250 else "CRÍTICO")
        consolidado_tabla[year]["severity"] = severity
        consolidado_tabla[year]["total_incidents"] = total_inc_año
        
        for obs_key, stats in types.items():
            worst = [{"celda": k, "count": v} for k, v in top_stations_por_año[year][obs_key].most_common(3)]
            consolidado_tabla[year][obs_key] = {
                "incidents": stats["incidents"],
                "affected_stations": len(stats["stations"]),
                "top_stations": worst
            }
            
    distribucion_campania = []
    for c, stats in camp_stats.items():
        rows_count = len(stats["filas"])
        total_fields_group = rows_count * 20
        distribucion_campania.append({
            "campania": c, "discontinuidades": rows_count, "vacios_cant": stats["vacios"],
            "vacios_pct": (stats["vacios"] / max(1, total_fields_group)) * 100,
            "advertencias_cant": stats["advertencias"], "advertencias_pct": (stats["advertencias"] / max(1, total_fields_group)) * 100,
            "alertas_cant": stats["alertas"], "alertas_pct": (stats["alertas"] / max(1, total_fields_group)) * 100
        })
        
    distribucion_geotecnico = []
    for g, stats in geo_stats.items():
        rows_count = len(stats["filas"])
        total_fields_group = rows_count * 20
        distribucion_geotecnico.append({
            "geotecnico": g, "discontinuidades": rows_count, "vacios_cant": stats["vacios"],
            "vacios_pct": (stats["vacios"] / max(1, total_fields_group)) * 100,
            "advertencias_cant": stats["advertencias"], "advertencias_pct": (stats["advertencias"] / max(1, total_fields_group)) * 100,
            "alertas_cant": stats["alertas"], "alertas_pct": (stats["alertas"] / max(1, total_fields_group)) * 100
        })
        
    distribucion_sector = []
    for s, stats in sector_stats.items():
        rows_count = len(stats["filas"])
        total_fields_group = rows_count * 20
        distribucion_sector.append({
            "sector": s, "discontinuidades": rows_count, "vacios_cant": stats["vacios"],
            "vacios_pct": (stats["vacios"] / max(1, total_fields_group)) * 100,
            "advertencias_cant": stats["advertencias"], "advertencias_pct": (stats["advertencias"] / max(1, total_fields_group)) * 100,
            "alertas_cant": stats["alertas"], "alertas_pct": (stats["alertas"] / max(1, total_fields_group)) * 100
        })
        
    msg_alertas = Counter(simplify_message(i.get("mensaje")) for i in incidencias if i.get("tipo_incidencia") == "ALERTA")
    msg_advertencias = Counter(simplify_message(i.get("mensaje")) for i in incidencias if i.get("tipo_incidencia") == "ADVERTENCIA")
    
    top_5_alertas = [{"mensaje": k, "cantidad": v, "pct": (v / max(1, total_alertas)) * 100} for k, v in msg_alertas.most_common(5)]
    lista_alertas = [{"mensaje": k, "cantidad": v, "pct": (v / max(1, total_alertas)) * 100} for k, v in msg_alertas.most_common()]
    lista_advertencias = [{"mensaje": k, "cantidad": v, "pct": (v / max(1, total_advertencias)) * 100} for k, v in msg_advertencias.most_common()]
    
    compact["audit_id"] = audit_id
    compact["fecha_auditoria"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    compact["nombre_archivo"] = os.path.basename(file_path)
    compact["consolidado_observaciones"] = consolidado_tabla
    
    compact["familia1"] = {
        "num_celdas_padre": num_celdas_padre,
        "promedio_hijas": round(promedio_hijas, 2),
        "total_discontinuidades": total_filas,
        "total_metros": round(total_metros, 2)
    }
    compact["familia2"] = {"total_fields": total_fields, "total_vacios": total_vacios, "total_advertencias": total_advertencias, "total_alertas": total_alertas, "total_correctos": total_correctos}
    compact["familia3"] = {"total_discontinuidades": total_filas, "discontinuidades_alertas": discs_con_alerta, "discontinuidades_advertencias": discs_con_advertencia, "discontinuidades_vacios": discs_con_vacio, "discontinuidades_correctas": discs_correctas}
    compact["distribucion_campania"] = distribucion_campania
    compact["distribucion_sector"] = distribucion_sector
    compact["distribucion_geotecnico"] = distribucion_geotecnico
    compact["top_5_alertas"] = top_5_alertas
    compact["error_types_detailed"] = {"alertas": lista_alertas, "advertencias": lista_advertencias}
    
    sorted_worst = sorted(resumen_celdas.items(), key=lambda x: (x[1].get("alertas", 0), x[1].get("vacios", 0), x[1].get("advertencias", 0)), reverse=True)[:20]
    compact["worst_cells"] = [{"celda": k, **v} for k, v in sorted_worst]
    col_counter = Counter(i.get("columna", "Desconocido") for i in incidencias)
    compact["top_column_errors"] = [{"columna": k, "cantidad": v} for k, v in col_counter.most_common(15)]
    
    # Guardar compact JSON
    compact_json_tmp = compact_json_out + ".tmp"
    with open(compact_json_tmp, "w", encoding="utf-8") as f:
        json.dump(compact, f, ensure_ascii=False)
    safe_replace(compact_json_tmp, compact_json_out)
    
    # Copiar resumen público
    public_compact = os.path.join(uploads_dir, "resumen_geomecanico_ligero.json")
    public_compact_tmp = public_compact + ".tmp"
    shutil.copyfile(compact_json_out, public_compact_tmp)
    safe_replace(public_compact_tmp, public_compact)

    print(f"[+] [{t_str()}] Finalización de compactado y guardado del resumen JSON en {compact_json_out}")

    # 3. Pre-generar Reporte Excel en segundo plano
    print(f"[*] [{t_str()}] Inicio de pre-generación del libro Excel para {audit_id}")
    excel_start = time.time()
    try:
        wb_rep = generar_excel_reporte_core(diag, compact, incidencias)
        rep_tmp = excel_pregenerated_out + ".tmp"
        wb_rep.save(rep_tmp)
        safe_replace(rep_tmp, excel_pregenerated_out)
        
        # Copiar reporte público
        public_excel = os.path.join(uploads_dir, "reporte_completo_ultimo.xlsx")
        public_excel_tmp = public_excel + ".tmp"
        shutil.copyfile(excel_pregenerated_out, public_excel_tmp)
        safe_replace(public_excel_tmp, public_excel)
        elapsed_excel = round(time.time() - excel_start, 2)
        print(f"[+] [{t_str()}] Libro Excel generado y guardado en disco con éxito ({elapsed_excel}s)")
    except Exception as e:
        print(f"[-] [{t_str()}] Error al pre-generar Excel de Logueo: {e}")

# --- API ENDPOINTS ---

@router.get("/logueo/estado-reporte")
def verificar_estado_reporte(audit_id: str):
    """Verifica si el reporte Excel pre-generado ya existe en disco."""
    file_path = os.path.join(history_dir, f"{audit_id}_reporte_completo.xlsx")
    return {"excel_ready": os.path.exists(file_path)}

@router.post("/logueo/cancelar-auditoria")
def cancelar_auditoria(audit_id: str):
    """Cancela la auditoría eliminando archivos del historial."""
    print(f"[*] [{t_str()}] Petición de cancelación recibida para {audit_id}")
    for ext in [".xlsx", "_diagnostico.json", "_compact.json", "_reporte_completo.xlsx"]:
        path = os.path.join(history_dir, f"{audit_id}{ext}")
        if os.path.exists(path):
            try:
                os.remove(path)
                print(f"[+] [{t_str()}] Archivo eliminado: {path}")
            except Exception as e:
                print(f"[-] [{t_str()}] No se pudo eliminar {path}: {e}")
    return {"status": "cancelado"}

@router.post("/logueo/sheets")
async def obtener_nombres_hojas(file: UploadFile = File(...)):
    """Sube un excel y extrae sus nombres de hoja."""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Formato de archivo no soportado. Debe ser Excel.")
        
    temp_filename = f"temp_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}"
    file_path = os.path.join(temp_dir, temp_filename)
    
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
        
    wb = None
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheet_names = wb.sheetnames
        return {"filename": temp_filename, "sheets": sheet_names}
    except Exception as e:
        if os.path.exists(file_path):
            try: os.remove(file_path)
            except: pass
        raise HTTPException(status_code=500, detail=f"Error leyendo el archivo Excel: {str(e)}")
    finally:
        if wb:
            try: wb.close()
            except: pass

@router.post("/logueo/importar-excel-bulk")
async def importar_excel_bulk(
    background_tasks: BackgroundTasks,
    payload: dict
):
    """Dispara el pipeline de auditoría de fondo con las hojas mapeadas."""
    filename = payload.get("filename")
    lgg_sheet = payload.get("lgg_sheet")
    est_sheet = payload.get("est_sheet")
    
    if not filename or not lgg_sheet or not est_sheet:
        raise HTTPException(status_code=400, detail="Faltan parámetros obligatorios.")
        
    temp_path = os.path.join(temp_dir, filename)
    if not os.path.exists(temp_path):
        raise HTTPException(status_code=404, detail="El archivo subido ya no existe en el servidor temporal.")
        
    audit_id = f"audit_logueo_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    file_path = os.path.join(history_dir, f"{audit_id}.xlsx")
    
    # Mover archivo a su ubicación permanente
    shutil.move(temp_path, file_path)
    
    background_tasks.add_task(run_logueo_audit_pipeline, file_path, lgg_sheet, est_sheet, audit_id)
    return {"status": "procesando", "audit_id": audit_id}

@router.get("/logueo/auditorias")
def listar_auditorias():
    if not os.path.exists(history_dir):
        return []
    audits = []
    for f in os.listdir(history_dir):
        if f.endswith("_compact.json"):
            audit_id = f.replace("_compact.json", "")
            compact_file = os.path.join(history_dir, f)
            try:
                with open(compact_file, "r", encoding="utf-8") as file_content:
                    meta = json.load(file_content)
                    audits.append({
                        "audit_id": audit_id,
                        "fecha": meta.get("fecha_auditoria", "Desconocida"),
                        "archivo": meta.get("nombre_archivo", "Desconocido.xlsx"),
                        "total_filas": meta.get("familia1", {}).get("total_discontinuidades", 0),
                        "total_vacios": meta.get("familia2", {}).get("total_vacios", 0),
                        "total_advertencias": meta.get("familia2", {}).get("total_advertencias", 0),
                        "total_alertas": meta.get("familia2", {}).get("total_alertas", 0)
                    })
            except:
                pass
    return sorted(audits, key=lambda x: x["fecha"], reverse=True)

@router.get("/logueo/resumen-ligero")
def obtener_resumen_ligero(audit_id: str = None, years: str = None):
    if audit_id:
        raw_file = os.path.join(history_dir, f"{audit_id}_diagnostico.json")
        compact_file = os.path.join(history_dir, f"{audit_id}_compact.json")
        excel_file = os.path.join(history_dir, f"{audit_id}.xlsx")
        
        if not os.path.exists(compact_file) or not os.path.exists(raw_file):
            if os.path.exists(excel_file):
                return JSONResponse(status_code=202, content={"status": "procesando", "message": "Auditoría geotécnica en proceso..."})
            raise HTTPException(status_code=404, detail="La auditoría no existe o fue eliminada.")
    else:
        raw_file = os.path.join(uploads_dir, "diagnostico_geomecanico.json")
        compact_file = os.path.join(uploads_dir, "resumen_geomecanico_ligero.json")
        if not os.path.exists(raw_file) or not os.path.exists(compact_file):
            jsons = [f for f in os.listdir(history_dir) if f.endswith("_diagnostico.json")]
            if jsons:
                jsons.sort(key=lambda x: os.path.getmtime(os.path.join(history_dir, x)), reverse=True)
                latest_id = jsons[0].replace("_diagnostico.json", "")
                raw_file = os.path.join(history_dir, f"{latest_id}_diagnostico.json")
                compact_file = os.path.join(history_dir, f"{latest_id}_compact.json")
            else:
                return JSONResponse(status_code=202, content={"status": "procesando", "message": "Esperando inicialización de datos de auditoría..."})

    with open(raw_file, "r", encoding="utf-8") as f:
        diag = json.load(f)
        
    incidencias = diag.get("incidencias", [])
    
    if years and years != "TODOS" and years != "":
        years_list = [y.strip() for y in years.split(",") if y.strip()]
        incidencias = [i for i in incidencias if str(i.get("campania")) in years_list]
        resumen_celdas_raw = diag.get("resumen_por_celda_padre", {})
        resumen_celdas = {k: v for k, v in resumen_celdas_raw.items() if str(v.get("campania")) in years_list}
        total_filas = len(incidencias)
    else:
        resumen_celdas = diag.get("resumen_por_celda_padre", {})
        total_filas = diag.get("total_filas_procesadas", 0)

    num_celdas_padre = len(resumen_celdas)
    promedio_hijas = sum(x["total_hijas"] for x in resumen_celdas.values()) / max(1, num_celdas_padre)
    total_metros = sum(safe_float(x.get("dist_celda", 0.0)) for x in resumen_celdas.values())
    
    total_fields = total_filas * 20
    total_vacios = sum(1 for i in incidencias if i.get("tipo_incidencia") == "VACIO")
    total_advertencias = sum(1 for i in incidencias if i.get("tipo_incidencia") == "ADVERTENCIA")
    total_alertas = sum(1 for i in incidencias if i.get("tipo_incidencia") == "ALERTA")
    total_correctos = total_fields - (total_vacios + total_advertencias + total_alertas)
    
    row_errors = defaultdict(set)
    for i in incidencias:
        row_errors[f"{i['modulo']}_{i['fila_excel']}"].add(i["tipo_incidencia"])
        
    discs_con_alerta = sum(1 for row, errs in row_errors.items() if "ALERTA" in errs)
    discs_con_advertencia = sum(1 for row, errs in row_errors.items() if "ADVERTENCIA" in errs and "ALERTA" not in errs)
    discs_con_vacio = sum(1 for row, errs in row_errors.items() if "VACIO" in errs)
    discs_correctas = total_filas - len(row_errors)
    
    camp_stats = defaultdict(lambda: {"vacios": 0, "advertencias": 0, "alertas": 0, "filas": set()})
    geo_stats = defaultdict(lambda: {"vacios": 0, "advertencias": 0, "alertas": 0, "filas": set()})
    sector_stats = defaultdict(lambda: {"vacios": 0, "advertencias": 0, "alertas": 0, "filas": set()})
    
    observaciones_por_año = defaultdict(lambda: defaultdict(lambda: {"incidents": 0, "stations": set()}))
    top_stations_por_año = defaultdict(lambda: defaultdict(lambda: Counter()))
    
    for i in incidencias:
        c = i.get("campania", "N/A")
        obs_key = simplify_message(i.get("mensaje", ""))
        celda = i.get("celda_padre", "N/A")
        
        observaciones_por_año[c][obs_key]["incidents"] += 1
        observaciones_por_año[c][obs_key]["stations"].add(celda)
        top_stations_por_año[c][obs_key][celda] += 1
        
        camp_stats[c]["filas"].add(f"{i['modulo']}_{i['fila_excel']}")
        geo_stats[i.get("geotecnico", "N/A")]["filas"].add(f"{i['modulo']}_{i['fila_excel']}")
        sector_stats[i.get("sector_geotecnico", "N/A")]["filas"].add(f"{i['modulo']}_{i['fila_excel']}")
        
        tipo = i.get("tipo_incidencia")
        if tipo == "VACIO":
            camp_stats[c]["vacios"] += 1
            geo_stats[i.get("geotecnico", "N/A")]["vacios"] += 1
            sector_stats[i.get("sector_geotecnico", "N/A")]["vacios"] += 1
        elif tipo == "ADVERTENCIA":
            camp_stats[c]["advertencias"] += 1
            geo_stats[i.get("geotecnico", "N/A")]["advertencias"] += 1
            sector_stats[i.get("sector_geotecnico", "N/A")]["advertencias"] += 1
        elif tipo == "ALERTA":
            camp_stats[c]["alertas"] += 1
            geo_stats[i.get("geotecnico", "N/A")]["alertas"] += 1
            sector_stats[i.get("sector_geotecnico", "N/A")]["alertas"] += 1
            
    consolidado_tabla = {}
    for year, types in observaciones_por_año.items():
        consolidado_tabla[year] = {}
        total_inc_año = sum(v["incidents"] for k, v in types.items())
        severity = "LEVE" if total_inc_año < 50 else ("MODERADO" if total_inc_año < 250 else "CRÍTICO")
        consolidado_tabla[year]["severity"] = severity
        consolidado_tabla[year]["total_incidents"] = total_inc_año
        for obs_key, stats in types.items():
            worst = [{"celda": k, "count": v} for k, v in top_stations_por_año[year][obs_key].most_common(3)]
            consolidado_tabla[year][obs_key] = {
                "incidents": stats["incidents"],
                "affected_stations": len(stats["stations"]),
                "top_stations": worst
            }
            
    distribucion_campania = []
    for c, stats in camp_stats.items():
        rows_count = len(stats["filas"])
        total_fields_group = rows_count * 20
        distribucion_campania.append({
            "campania": c, "discontinuidades": rows_count, "vacios_cant": stats["vacios"],
            "vacios_pct": (stats["vacios"] / max(1, total_fields_group)) * 100,
            "advertencias_cant": stats["advertencias"], "advertencias_pct": (stats["advertencias"] / max(1, total_fields_group)) * 100,
            "alertas_cant": stats["alertas"], "alertas_pct": (stats["alertas"] / max(1, total_fields_group)) * 100
        })
        
    distribucion_geotecnico = []
    for g, stats in geo_stats.items():
        rows_count = len(stats["filas"])
        total_fields_group = rows_count * 20
        distribucion_geotecnico.append({
            "geotecnico": g, "discontinuidades": rows_count, "vacios_cant": stats["vacios"],
            "vacios_pct": (stats["vacios"] / max(1, total_fields_group)) * 100,
            "advertencias_cant": stats["advertencias"], "advertencias_pct": (stats["advertencias"] / max(1, total_fields_group)) * 100,
            "alertas_cant": stats["alertas"], "alertas_pct": (stats["alertas"] / max(1, total_fields_group)) * 100
        })
        
    distribucion_sector = []
    for s, stats in sector_stats.items():
        rows_count = len(stats["filas"])
        total_fields_group = rows_count * 20
        distribucion_sector.append({
            "sector": s, "discontinuidades": rows_count, "vacios_cant": stats["vacios"],
            "vacios_pct": (stats["vacios"] / max(1, total_fields_group)) * 100,
            "advertencias_cant": stats["advertencias"], "advertencias_pct": (stats["advertencias"] / max(1, total_fields_group)) * 100,
            "alertas_cant": stats["alertas"], "alertas_pct": (stats["alertas"] / max(1, total_fields_group)) * 100
        })
        
    msg_alertas = Counter(simplify_message(i.get("mensaje")) for i in incidencias if i.get("tipo_incidencia") == "ALERTA")
    msg_advertencias = Counter(simplify_message(i.get("mensaje")) for i in incidencias if i.get("tipo_incidencia") == "ADVERTENCIA")
    
    top_5_alertas = [{"mensaje": k, "cantidad": v, "pct": (v / max(1, total_alertas)) * 100} for k, v in msg_alertas.most_common(5)]
    lista_alertas = [{"mensaje": k, "cantidad": v, "pct": (v / max(1, total_alertas)) * 100} for k, v in msg_alertas.most_common()]
    lista_advertencias = [{"mensaje": k, "cantidad": v, "pct": (v / max(1, total_advertencias)) * 100} for k, v in msg_advertencias.most_common()]
    
    res_compact = {
        "audit_id": audit_id or "default",
        "fecha_auditoria": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "nombre_archivo": os.path.basename(raw_file),
        "consolidado_observaciones": consolidado_tabla,
        "resumen_por_celda_padre": resumen_celdas,
        "familia1": {
            "num_celdas_padre": num_celdas_padre,
            "promedio_hijas": round(promedio_hijas, 2),
            "total_discontinuidades": total_filas,
            "total_metros": round(total_metros, 2)
        },
        "familia2": {"total_fields": total_fields, "total_vacios": total_vacios, "total_advertencias": total_advertencias, "total_alertas": total_alertas, "total_correctos": total_correctos},
        "familia3": {"total_discontinuidades": total_filas, "discontinuidades_alertas": discs_con_alerta, "discontinuidades_advertencias": discs_con_advertencia, "discontinuidades_vacios": discs_con_vacio, "discontinuidades_correctas": discs_correctas},
        "distribucion_campania": distribucion_campania,
        "distribucion_sector": distribucion_sector,
        "distribucion_geotecnico": distribucion_geotecnico,
        "top_5_alertas": top_5_alertas,
        "error_types_detailed": {"alertas": lista_alertas, "advertencias": lista_advertencias}
    }
    
    sorted_worst = sorted(resumen_celdas.items(), key=lambda x: (x[1].get("alertas", 0), x[1].get("vacios", 0), x[1].get("advertencias", 0)), reverse=True)[:20]
    res_compact["worst_cells"] = [{"celda": k, **v} for k, v in sorted_worst]
    col_counter = Counter(i.get("columna", "Desconocido") for i in incidencias)
    res_compact["top_column_errors"] = [{"columna": k, "cantidad": v} for k, v in col_counter.most_common(15)]
    
    return res_compact

@router.get("/logueo/incidencias-paginadas")
def obtener_incidencias_paginadas(
    page: int = 1, limit: int = 50, tipo: str = None, celda: str = None, columna: str = None,
    campania: str = None, geotecnico: str = None, search: str = None, audit_id: str = None
):
    if audit_id: 
        raw_file = os.path.join(history_dir, f"{audit_id}_diagnostico.json")
        if not os.path.exists(raw_file):
            return {"data": [], "page": 1, "total_pages": 1, "total_records": 0}
    else:
        raw_file = os.path.join(uploads_dir, "diagnostico_geomecanico.json")
        if not os.path.exists(raw_file):
            jsons = [f for f in os.listdir(history_dir) if f.endswith("_diagnostico.json")]
            if jsons:
                jsons.sort(key=lambda x: os.path.getmtime(os.path.join(history_dir, x)), reverse=True)
                raw_file = os.path.join(history_dir, jsons[0])
            else:
                return {"data": [], "page": 1, "total_pages": 1, "total_records": 0}

    with open(raw_file, "r", encoding="utf-8") as f:
        diag = json.load(f)
        
    incidencias = diag.get("incidencias", [])
    
    # Aplicar filtros
    if tipo:
        incidencias = [i for i in incidencias if i.get("tipo_incidencia") == tipo]
    if celda:
        incidencias = [i for i in incidencias if i.get("celda_padre") == celda]
    if columna:
        incidencias = [i for i in incidencias if i.get("columna") == columna]
    if campania:
        c_list = [c.strip() for c in campania.split(",") if c.strip()]
        incidencias = [i for i in incidencias if str(i.get("campania")) in c_list]
    if geotecnico:
        incidencias = [i for i in incidencias if i.get("geotecnico") == geotecnico]
    if search:
        search_lower = search.lower()
        incidencias = [
            i for i in incidencias 
            if search_lower in str(i.get("mensaje", "")).lower() 
            or search_lower in str(i.get("columna", "")).lower()
            or search_lower in str(i.get("celda_padre", "")).lower()
            or search_lower in str(i.get("celda_hija", "")).lower()
        ]
        
    total_records = len(incidencias)
    total_pages = max(1, math.ceil(total_records / limit))
    page = max(1, min(page, total_pages))
    
    start_idx = (page - 1) * limit
    end_idx = start_idx + limit
    paginated = incidencias[start_idx:end_idx]
    
    return {
        "data": paginated,
        "page": page,
        "total_pages": total_pages,
        "total_records": total_records
    }

@router.get("/logueo/reporte-excel")
def descargar_reporte_excel(audit_id: str = None):
    if audit_id:
        file_path = os.path.join(history_dir, f"{audit_id}_reporte_completo.xlsx")
    else:
        file_path = os.path.join(uploads_dir, "reporte_completo_ultimo.xlsx")
        
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="El reporte Excel solicitado no se encuentra en el servidor. Espere a que termine el procesamiento.")
        
    return FileResponse(
        path=file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="Reporte_Auditoria_Geotecnica_Logueo.xlsx"
    )

@router.get("/logueo/reporte-markdown")
def descargar_reporte_markdown(audit_id: str = None, years: str = None):
    resumen = obtener_resumen_ligero(audit_id, years)
    if isinstance(resumen, JSONResponse):
        return resumen
        
    # Construir reporte markdown a mano con datos actualizados
    title = resumen.get("nombre_archivo", "Archivo de Logueo")
    m = resumen.get("metricas_globales", {})
    correct_pct = ((m.get("total_ok", 0) / max(1, m.get("total_celdas_hija_procesadas", 0))) * 100)
    
    md_content = f"""# REPORTE DE AUDITORÍA GEOTÉCNICA Y CONTROL DE CALIDAD (QA/QC)
Generado el: {resumen.get("fecha_auditoria", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))}
Archivo evaluado: {title}
Código de Auditoría: {resumen.get("audit_id", "default")}

---

## 1. RESUMEN EJECUTIVO DE INTEGRIDAD
El sistema de auditoría ha evaluado la consistencia física de las corridas de logueo general (LGG) y la relación espacial de las discontinuidades estructurales (Logueo Estructural), cruzando durezas, litologías, e intervalos.

*   **Total de Taladros Evaluados:** {resumen.get("familia1", {}).get("num_celdas_padre", 0)}
*   **Total de Registros de Logueo Procesados:** {resumen.get("familia1", {}).get("total_discontinuidades", 0)}
*   **Mapeo Lineal Equivalente:** {resumen.get("familia1", {}).get("total_metros", 0.0)} metros perforados.
*   **Integridad Global de Registros:** {correct_pct:.2f}% de filas sin desviaciones críticas.
*   **Total de Desviaciones Críticas (Alertas):** {m.get("total_alertas", 0)}
*   **Total de Advertencias de Consistencia:** {m.get("total_advertencias", 0)}
*   **Campos Obligatorios Vacíos:** {m.get("total_vacios", 0)}

---

## 2. PRINCIPALES ANOMALÍAS CRÍTICAS DETECTADAS
A continuación se listan las reglas que más frecuentemente se han incumplido:

"""
    for idx, item in enumerate(resumen.get("top_5_alertas", []), start=1):
        md_content += f"{idx}. **{item['mensaje']}** - {item['cantidad']} casos ({item['pct']:.2f}% de las alertas críticas).\n"
        
    md_content += """
---

## 3. DESEMPEÑO POR CAMPAÑA Y RESPONSABLE
A continuación se detalla la cantidad de desviaciones encontradas agrupadas por campaña de perforación:

| Campaña / Año | Registros | Alertas (N) | % Alertas | Vacíos (N) | % Vacíos |
|---|---|---|---|---|---|
"""
    for c in resumen.get("distribucion_campania", []):
        md_content += f"| {c['campania']} | {c['discontinuidades']} | {c['alertas_cant']} | {c['alertas_pct']:.2f}% | {c['vacios_cant']} | {c['vacios_pct']:.2f}% |\n"
        
    md_content += """
---
*Fin del Reporte Técnico de Auditoría. Geolog Pro 2.0.*
"""
    headers = {
        'Content-Disposition': f'attachment; filename="Reporte_Auditoria_{resumen.get("audit_id", "logueo")}.md"'
    }
    return StreamingResponse(io.BytesIO(md_content.encode("utf-8")), media_type="text/markdown", headers=headers)

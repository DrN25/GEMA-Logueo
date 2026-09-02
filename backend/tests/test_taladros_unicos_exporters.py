import unittest
import os
import sys
import openpyxl

# Add backend directory to sys.path
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)

from app.services.plt_excel_exporter_regulares import export_plt_regulares_to_excel
from app.routers.auditoria import generar_excel_reporte_core

class TestTaladrosUnicosExporters(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.temp_dir = os.path.join(BASE_DIR, "uploads", "test_temp")
        os.makedirs(cls.temp_dir, exist_ok=True)

    @classmethod
    def tearDownClass(cls):
        import shutil
        if os.path.exists(cls.temp_dir):
            shutil.rmtree(cls.temp_dir, ignore_errors=True)

    def test_plt_exporter_discrepancies(self):
        """Verifica que en el reporte PLT vs LGG solo se listen taladros exclusivos."""
        diag = {
            "total_rows": 2,
            "drillhole_stats": {"DH-01": {}, "DH-ONLY-PLT": {}},
            "unique_samples_plt": [
                {"taladro": "DH-01", "from_m": 10, "to_m": 10.5, "campana": "2023", "p_kn": 5.2, "d_mm": 50, "is50_mpa": 4.1, "ucs_mpa": 82, "lito1": "AND", "incidencias_cant": 0, "estado": "CONFORME"},
                {"taladro": "DH-ONLY-PLT", "from_m": 20, "to_m": 20.5, "campana": "2024", "p_kn": 6.1, "d_mm": 50, "is50_mpa": 4.8, "ucs_mpa": 96, "lito1": "AND", "incidencias_cant": 0, "estado": "CONFORME"},
            ],
            "unique_runs_lgg": [
                {"taladro": "DH-01", "de": 0, "a": 100, "campana": "2023"},
                {"taladro": "DH-ONLY-LGG", "de": 0, "a": 50, "campana": "2023"},
            ],
            "anomalies": []
        }
        out_path = os.path.join(self.temp_dir, "test_plt_discrepancies.xlsx")
        export_plt_regulares_to_excel(diag, out_path)

        wb = openpyxl.load_workbook(out_path)
        ws_plt = wb["🗂️ Taladros Únicos PLT"]
        ws_lgg = wb["🗂️ Taladros Únicos LGG"]

        # Solo debe listar DH-ONLY-PLT en la hoja PLT
        plt_dh_rows = [ws_plt.cell(r, 3).value for r in range(6, ws_plt.max_row + 1) if ws_plt.cell(r, 3).value]
        self.assertEqual(plt_dh_rows, ["DH-ONLY-PLT"])
        self.assertEqual(ws_plt.cell(6, 16).value, "❌ NO REGISTRADO EN LGG")

        # Solo debe listar DH-ONLY-LGG en la hoja LGG
        lgg_dh_rows = [ws_lgg.cell(r, 3).value for r in range(6, ws_lgg.max_row + 1) if ws_lgg.cell(r, 3).value]
        self.assertEqual(lgg_dh_rows, ["DH-ONLY-LGG"])
        self.assertEqual(ws_lgg.cell(6, 11).value, "⚠️ SIN ENSAYOS PLT")

    def test_plt_exporter_zero_discrepancies(self):
        """Verifica banner verde cuando todos los taladros coinciden 100%."""
        diag = {
            "total_rows": 1,
            "drillhole_stats": {"DH-MATCH": {}},
            "unique_samples_plt": [
                {"taladro": "DH-MATCH", "from_m": 10, "to_m": 10.5, "campana": "2023", "p_kn": 5.2, "d_mm": 50, "is50_mpa": 4.1, "ucs_mpa": 82, "lito1": "AND", "incidencias_cant": 0, "estado": "CONFORME"},
            ],
            "unique_runs_lgg": [
                {"taladro": "DH-MATCH", "de": 0, "a": 100, "campana": "2023"},
            ],
            "anomalies": []
        }
        out_path = os.path.join(self.temp_dir, "test_plt_zero_discrepancies.xlsx")
        export_plt_regulares_to_excel(diag, out_path)

        wb = openpyxl.load_workbook(out_path)
        ws_plt = wb["🗂️ Taladros Únicos PLT"]
        ws_lgg = wb["🗂️ Taladros Únicos LGG"]

        self.assertIn("CONFORME", str(ws_plt.cell(6, 2).value))
        self.assertIn("CONFORME", str(ws_lgg.cell(6, 2).value))

    def test_auditoria_lgg_est_exporter_discrepancies(self):
        """Verifica que en el reporte LGG vs Estructural solo se listen taladros exclusivos."""
        diag = {
            "unique_lgg_runs": [
                {"taladro": "DH-BOTH", "de": 0, "a": 10, "longitud": 10, "rec_m": 10, "rqd_m": 8, "estado": "CONFORME"},
                {"taladro": "DH-ONLY-LGG", "de": 0, "a": 10, "longitud": 10, "rec_m": 10, "rqd_m": 8, "estado": "CONFORME"},
            ],
            "unique_est_structures": [
                {"taladro": "DH-BOTH", "profundidad": 5, "estado": "CONFORME"},
                {"taladro": "DH-ONLY-EST", "profundidad": 8, "estado": "CONFORME"},
            ],
            "incidencias": []
        }
        compact = {
            "resumen_por_celda_padre": {},
            "distribucion_campania": [],
            "distribucion_geotecnico": []
        }
        wb = generar_excel_reporte_core(diag, compact, [])

        ws_lgg = wb["🗂️ Taladros Únicos LGG"]
        ws_est = wb["🗂️ Taladros Únicos Estructural"]

        lgg_dh_rows = [ws_lgg.cell(r, 3).value for r in range(6, ws_lgg.max_row + 1) if ws_lgg.cell(r, 3).value]
        self.assertEqual(lgg_dh_rows, ["DH-ONLY-LGG"])
        self.assertEqual(ws_lgg.cell(6, 15).value, "❌ SIN ESTRUCTURAS")

        est_dh_rows = [ws_est.cell(r, 3).value for r in range(6, ws_est.max_row + 1) if ws_est.cell(r, 3).value]
        self.assertEqual(est_dh_rows, ["DH-ONLY-EST"])
        self.assertEqual(ws_est.cell(6, 11).value, "❌ NO REGISTRADO EN LGG")

    def test_auditoria_lgg_est_exporter_zero_discrepancies(self):
        """Verifica banner verde cuando todos los taladros coinciden 100% entre LGG y Estructural."""
        diag = {
            "unique_lgg_runs": [
                {"taladro": "DH-BOTH", "de": 0, "a": 10, "longitud": 10, "rec_m": 10, "rqd_m": 8, "estado": "CONFORME"},
            ],
            "unique_est_structures": [
                {"taladro": "DH-BOTH", "profundidad": 5, "estado": "CONFORME"},
            ],
            "incidencias": []
        }
        compact = {
            "resumen_por_celda_padre": {},
            "distribucion_campania": [],
            "distribucion_geotecnico": []
        }
        wb = generar_excel_reporte_core(diag, compact, [])

        ws_lgg = wb["🗂️ Taladros Únicos LGG"]
        ws_est = wb["🗂️ Taladros Únicos Estructural"]

        self.assertIn("CONFORME", str(ws_lgg.cell(6, 2).value))
        self.assertIn("CONFORME", str(ws_est.cell(6, 2).value))


if __name__ == "__main__":
    unittest.main()
